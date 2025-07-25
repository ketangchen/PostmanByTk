# -*- coding: utf-8 -*-
import tkinter as tk  # /usr/local/bin/python3.7 pip -m install treelib  -i https://pypi.tuna.tsinghua.edu.cn/simple
from tkinter import ttk, \
    scrolledtext  # D:\Python\Python37\python.exe pip install pillow  -i https://pypi.tuna.tsinghua.edu.cn/simple
import requests
import datetime
from faker import Faker
import json
import string
import random
import time
from tkinter import messagebox
import glob

from PIL import Image, \
    ImageTk  # /usr/local/bin/python3.7 -m pip install Pillow -i https://pypi.tuna.tsinghua.edu.cn/simple
from PIL import Image, ImageTk  # pip install pillow

import os
from os.path import isfile, splitext
import tkinter as tk
from tkinter import filedialog
import threading
# import win32con
import ctypes
from xmindparser import xmind_to_dict
from treelib import Tree  # /usr/local/bin/python3.7 -m pip install treelib
from uuid import uuid4
from openpyxl import Workbook

from time import sleep
import os
import paddle
from textblob import TextBlob
import onnxruntime
import numpy as np
from paddleocr import PaddleOCR
import openai  # /usr/local/bin/python3.7 -m pip install openai -i https://pypi.tuna.tsinghua.edu.cn/simple

# 配置API（注意：使用的是非官方代理端点，可能存在风险）
openai.api_key = "your_api_key"  # 建议改用环境变量
openai.base_url = "https://free.v36.cm/v1/"  # 非官方端点
openai.default_headers = {"x-foo": "true"}

from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Border, Side, PatternFill

# 确保存储日志的目录存在
current_script_path = os.getcwd()
logFileName = 'requests_log'
logFilePath = f'{current_script_path}/{logFileName}'
if not os.path.exists(logFilePath):
    os.makedirs(logFilePath)

fileName = 'file'
filePath = f'{current_script_path}/{fileName}'
if not os.path.exists(filePath):
    os.makedirs(filePath)

caseExcelFileName = 'caseExcel'
caseExcelfilePath = f'{current_script_path}/{caseExcelFileName}'
if not os.path.exists(caseExcelfilePath):
    os.makedirs(caseExcelfilePath)

encodingType = 'gbk'





def write_xmind_data_to_excel(casePathDataList, caseType):
    """
    根据不同公司标准选择不同的用例格式来将用例的路径写进Excel
    """
    print(casePathDataList)
    if caseType == '版本用例' or caseType == '回归用例':
        ###中国平安用例格式
        versionCaseLevelDict = {'0': 'P0 冒烟测试', '1': 'P1 正向流程', '2': 'P2 异常反向'}
        regressionCaseLevelDict = {'0': 'P0 重要且频繁（致命，回归必跑）', '1': 'P1 重要不频繁（严重，评审选跑）', '2': 'P2 非重要功能（一般，评审选跑）'}
        defaultList = ['M 功能测试', 'P1 正向流程', 'A 需求分析']
        if caseType == '回归用例':
            defaultList = ['P2 非重要功能（一般，评审选跑）', 'APP类', ' ', '否', '不可以', 'A 需求分析']
        if len(casePathDataList) != 0:
            for d in casePathDataList:
                temp = d
                if len(d) == 3:
                    d.extend(defaultList)
                elif len(d) == 4:
                    level = d[3]
                    d.pop(-1)
                    if caseType != '回归用例':
                        if level in versionCaseLevelDict:
                            d.extend(['M 功能测试', versionCaseLevelDict[level], 'A 需求分析'])
                        else:
                            d.extend(['M 功能测试', versionCaseLevelDict['1'], 'A 需求分析'])
                    else:
                        if level in regressionCaseLevelDict:
                            d.extend([regressionCaseLevelDict[level], 'APP类', ' ', '否', '不可以', 'A 需求分析'])
                        else:
                            d.extend([regressionCaseLevelDict['2'], 'APP类', ' ', '否', '不可以', 'A 需求分析'])

                print(casePathDataList)
        # 创建一个新的工作薄
        wb = Workbook()
        # 选择默认的工作表
        ws = wb.active
        # 要写入的数据"M 功能测试"PO 冒烟测试'A 需求分析
        rows = [["用例名称", "用例步骤", "预期结果", "用例类型", "用例级别", "用例来源"]]
        if caseType == '回归用例':
            rows = [["用例名称", "用例步骤", "预期结果", "用例级别", "用例类型", "用例描述", "是否可自动化", "不可自动化原因", "用例来源"]]
        rows.extend(casePathDataList)
        print(rows)
        # 逐行写入数据
        for row in rows:
            ws.append(row)
        # 保存工作簿
        # wb.save('result.xlsx')
        wb.save(rf'{current_script_path}\\{caseExcelFileName}\\{getCurrentTimeInfo1()}-result.xlsx')
    elif caseType == '普通用例':
        ###广交所用例格式
        print(casePathDataList)
        # 创建一个新的工作薄
        wb = Workbook()
        # 选择默认的工作表
        ws = wb.active
        rows = [
                   '测试用例ID,系统名称,功能模块,一级功能,二级功能,三级功能,设计名称,前提条件,测试数据,验证点,验证内容,测试步骤,期望结果,实际结果,是否通过,测试版本,测试人员,复核人员,测试日期,优先级'.split(
                       ',')] * 2
        rows.extend(casePathDataList)
        # 逐行写入数据
        for row in rows:
            ws.append(row)

        # 定义黄色填充
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        # 定义边框样式
        border_style = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000")
        )

        # 遍历数据行，设置边框并根据条件染色
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            # 检查是否需要染色（假设第15列是“是否通过”列）
            if "优先级：" in row[19].value:  # row[4] 是第5列的单元格
                for cell in row:
                    cell.fill = yellow_fill  # 染色为黄色
            # 为每个单元格设置边框
            for cell in row:
                cell.border = border_style

        # 删除最后一列
        # wb.max_column 返回的是最后一列的索引，例如最后一列是第10列，则返回10
        ws.delete_cols(ws.max_column)

        # print(getCurrentTimeInfo1())
        # print(f'{current_script_path}\\{caseExcelFileName}\\{getCurrentTimeInfo1()}-result.xlsx')

        # 保存工作簿
        wb.save(f'{current_script_path}\\{caseExcelFileName}\\{getCurrentTimeInfo1()}-result.xlsx')


def xmind_to_tree(xmind_file_path):
    # 将Xmind文件解析为字典
    xmind_data = xmind_to_dict(xmind_file_path)
    # 创建Tree对象
    tree = Tree()

    # 递归函数，用于构建多叉树
    def build_tree_from_topic(topic, parent):
        # 为每个主题创建一个树节点
        node_id = str(uuid4())  # 生成唯一的节点ID
        tree.create_node(topic['title'], node_id, parent=parent)

        # 遍历子主题并递归调用
        for sub_topic in topic.get('topics', []):
            build_tree_from_topic(sub_topic, node_id)

    # 从Xmind数据中找到根主题，并开始构建树
    for sheet in xmind_data:
        root_topic = sheet['topic']
        build_tree_from_topic(root_topic, None)

    return tree


def get_all_leaf_paths_data_list(tree, caseType):
    """
    根据不同公司标准选择不同的用例格式来获取用例的路径
    """
    if caseType == '版本用例' or caseType == '回归用例':
        ###中国平安用例格式
        res = []
        d = {}  # 存tag和identifier到字典做映射
        all_nodes_info = tree.all_nodes()
        for a in all_nodes_info:
            d[a.identifier] = a.tag

        all_init_paths = tree.paths_to_leaves()  # 加工初始化路径，返回标题形式
        for a in all_init_paths:

            temp = [d[i] for i in a]
            print(f'casePath:{temp}')
            if len(temp) < 2:
                continue
            case_1 = temp[-1]  # 叶子结点
            case_2 = temp[-2]  # 叶子结点的父节点

            if '名称:' not in case_1 and '名称:' not in case_2:
                continue
            if '名称:' in case_1 and '名称:' in case_2:
                continue

            if '名称:' in case_1 and '名称:' not in case_2:
                try:
                    caseName = case_1.split('名称:')[1].split('步骤:')[0].replace('\r', '').replace('\n', '')
                    caseStep = case_1.split('步骤:')[1].split('预期:')[0].replace('\r', '').replace('\n', '')
                    caseExpect = case_1.split('预期:')[1].replace('\r', '').replace('\n', '')
                    temp.pop(-1)
                    casePath = '/'.join(temp) + '/' + caseName
                    res.append([casePath, caseStep, caseExpect])
                except Exception as e:
                    print(f'错误案例为：{case_1}')
                    print(f'转化异常：{e}')

            if '名称:' not in case_1 and '名称:' in case_2:
                try:
                    caseName = case_2.split('名称:')[1].split('步骤:')[0].replace('\r', '').replace('\n', '')
                    caseStep = case_2.split('步骤:')[1].split('预期:')[0].replace('\r', '').replace('\n', '')
                    caseExpect = case_2.split('预期:')[1].replace('\r', '').replace('\n', '')
                    temp.pop(-1)
                    casePath = '/'.join(temp) + '/' + caseName
                    res.append([casePath, caseStep, caseExpect, case_1])
                except Exception as e:
                    print(f'错误案例为：{case_2}')
                    print(f'转化异常：{e}')
        return res
    elif caseType == '普通用例':
        ###广交所用例格式
        rows = '测试用例ID,系统名称,功能模块,一级功能,二级功能,三级功能,设计名称,前提条件,测试数据,验证点,验证内容,测试步骤,期望结果,实际结果,是否通过,测试版本,测试人员,复核人员,测试日期,优先级'.split(
            ',')
        d = {'测试用例ID': 0, '系统名称': 1, '功能模块': 2, '一级功能': 3, '二级功能': 4, '三级功能': 5, '设计名称': 6, '前提条件': 7, '测试数据': 8,
             '验证点': 9, '验证内容': 10, '测试步骤': 11, '期望结果': 12, '实际结果': 13, '是否通过': 14, '测试版本': 15, '测试人员': 16, '复核人员': 17,
             '测试日期': 18, '优先级': 19}
        print(d)
        res = []
        all_nodes_info = tree.all_nodes()
        for a in all_nodes_info:
            d[a.identifier] = a.tag

        all_init_paths = tree.paths_to_leaves()  # 加工初始化路径，返回标题形式
        count = 0
        characterCountLen = len(str(len(all_init_paths)))
        for a in all_init_paths:
            restemp = [' '] * len(rows)
            restemp[d['测试用例ID']] = '用例' + '0' * (characterCountLen - len(f'{count + 1}')) + f'{count + 1}'
            restemp[d['功能模块']] = '功能模块'
            count += 1
            print(restemp)
            temp = [d[i] for i in a]
            print(temp)
            for k, t in enumerate(temp):
                print(t)
                if k == 1 and ('设计名称' not in t):
                    restemp[d['系统名称']] = t
                elif k == 2 and ('设计名称' not in t) and ('：' not in t):
                    restemp[d['一级功能']] = t
                elif k == 3 and ('设计名称' not in t) and ('：' not in t):
                    restemp[d['二级功能']] = t
                elif k == 4 and ('设计名称' not in t) and ('：' not in t):
                    restemp[d['三级功能']] = t
                elif '设计名称' in t:
                    restemp[d['设计名称']] = t
                elif '验证点' in t:
                    restemp[d['验证点']] = t
                elif '前提条件' in t:
                    restemp[d['前提条件']] = t
                elif '测试数据' in t:
                    restemp[d['测试数据']] = t
                elif '验证内容' in t:
                    restemp[d['验证内容']] = t
                    restemp[d['测试步骤']] = t.replace('验证内容', '测试步骤')
                    restemp[d['期望结果']] = t.replace('验证内容', '期望结果')
                elif '测试步骤' in t:
                    restemp[d['测试步骤']] = t
                elif '期望结果' in t:
                    restemp[d['期望结果']] = t
                elif '优先级' in t:
                    restemp[d['优先级']] = t
            res.append(restemp)
            print(restemp)
        return res


# 读取JSON文件
def read_json_file1(file_path) -> dict:
    temp = []
    res = ''
    with open(file_path, 'r', encoding=encodingType) as f:
        lines = f.readlines()
        for line in lines:
            temp.append(line.strip())
    print(temp)
    try:
        res = json.loads(''.join(temp))
        print(res)
    except Exception as e:
        res = {"error": f"{e}"}
    return res


def read_json_file(file_path) -> dict:
    with open(file_path, 'r', encoding=encodingType) as file:  # 记得把对应json文件编码转化为gbk编码
        data = json.load(file)
        print(data)
        return data


def getCurrentTimeInfo():
    currentTimeStamp = time.time()
    timestamp = currentTimeStamp
    print(f'timestamp:{timestamp}')
    formatted_time = time.strftime('%Y-%m-%d-%H-%M-%S', time.localtime(timestamp))
    # print(currentTimeStamp,formatted_time)
    return (currentTimeStamp, formatted_time)


def getCurrentTimeInfo1():
    currentTimeStamp = time.time()
    timestamp = currentTimeStamp
    formatted_time = time.strftime('%Y-%m-%d-%H-%M-%S', time.localtime(timestamp))
    # print(currentTimeStamp,formatted_time)
    return formatted_time


def save_request_data(url, headers, method, request_data, response_data):
    """保存请求及其响应到文件"""
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    filename = f'requests_log/{timestamp}.json'
    data = {
        'url': url,
        'headers': headers,
        'method': method,
        'request': request_data,
        'response': response_data
    }
    try:
        with open(filename, 'w') as file:
            json.dump(data, file, ensure_ascii=False, indent=4)
    except Exception as e:
        print(e)


def list_request_records():
    """列出所有保存的请求记录文件"""
    files = os.listdir('requests_log')
    files.sort(reverse=True)
    return files[::-1]


def find_all_key(dictionary):
    """
    在复杂宁典中找到所有的键。
    :param dictionary:输入的复杂宁典。
    :return:所有的健列表。
    """
    stack = [((), dictionary)]
    res = []
    while stack:
        path, current_dict = stack.pop()
        for key, value in current_dict.items():
            res.append(key)
            new_path = path + (key,)
            if isinstance(value, dict):
                stack.append((new_path, value))
    return res


def find_key_and_update_value(dictionary, key_to_find, new_value) -> None:
    """
    在复杂字典中找到指定键的路径，并更新其值
    :paramdictionary:输入的复杂宁典
    :param key_to_find:需要查找的键。
    :param new value: 新的值.
    :return: 如果找到键，返回True;否则返回FaLse。
    """
    stack = [((), dictionary)]
    flag = True
    while stack and flag:
        path, current_dict = stack.pop()
        for key, value in current_dict.items():
            new_path = path + (key,)
            if key == key_to_find:
                current_dict[key] = new_value
                flag = False
                break
            elif isinstance(value, dict):
                stack.append((new_path, value))


def find_value_of_key_in_nested_dict(d, target_key):
    """
    非递归版本的搜索炭套宇典中的键对应值。
    :param d: 要搜索的宁典或宁典列表。
    :param target_key:要搜索的日标键
    :return: 包含找到指定键对应值。
    """

    if isinstance(d, str):
        d = json.loads(d)  # <class 'dict'>
    if isinstance(d, dict):
        stack = [(d, ())]
    elif isinstance(d, list):
        stack = [(item, ()) for item in d]
    else:
        return []

    result = []

    while stack:
        current, path = stack.pop()

        if isinstance(current, dict):
            for key, value in current.items():
                current_path = path + (key,)
                if key == target_key:
                    result.append((current_path, value))
                if isinstance(value, (dict, list)):
                    stack.append((value, current_path))
        elif isinstance(current, list):
            for idx, item in enumerate(current):
                current_path = path + (idx,)
                if isinstance(item, (dict, list)):
                    stack.append((item, current_path))

    print(result)
    for r in result:
        if len(r) != 0 and r[1] != None:
            return r[1]
    return 'None'


def generate_vin(length=17):
    valid_chars = string.digits + string.ascii_uppercase
    valid_chars = ''.join(filter(lambda x: x not in 'QIO', valid_chars))
    return ''.join(random.choices(valid_chars, k=length))


def generate_reqular_plate():
    province_codes = '京津沪渝冀豫云辽黑湘院鲁新苏浙赣鄂桂甘晋蒙陕吉闽贵粤青藏川宁琼'
    region_codes = ''.join((c for c in string.ascii_uppercase if c not in 'I0'))
    sequence = ''.join((c for c in string.ascii_uppercase if c not in 'I0'))
    sequence += ''.join((str(i) for i in range(10)))
    province = random.choice(province_codes)
    region = random.choice(region_codes)
    sequence = ''.join(random.choices(sequence, k=5))
    return f'{province}{region}{sequence}'


def generate_new_energy_plate(type='small'):
    province_codes = '京津沪渝冀豫云辽黑湘皖鲁新苏浙赣鄂桂甘蒙陕吉闽贵粤青川宁琼'
    province = random.choice(province_codes)
    prefix_letter = 'DF' if type == 'small' else random.choice('DF')
    sequence = ''.join(random.choices(string.digits, k=4)) if type == 'small' else ''.join(
        random.choices(string.digits + 'DF', k=5))
    return f"{province}{prefix_letter}{sequence}"


def generate_police_plate():
    province_codes = '京津沪渝冀豫云辽黑湘皖鲁新苏浙赣鄂桂甘晋蒙陕吉贵粤青藏川宁琼'
    region_codes = ''.join((c for c in string.ascii_uppercase if c not in 'I0'))
    province = random.choice(province_codes)
    region = random.choice(region_codes)
    number = ''.join(random.choices(string.digits, k=4))
    return f'{province}{region}{number}警'


def generate_consulate_plate():
    province_codes = '沪粤川云桂鄂闽鲁陕蒙藏黑辽渝'
    region_codes = 'ABCD'
    province = random.choice(province_codes)
    region = random.choice(region_codes)
    number = ''.join(random.choices(string.digits, k=4))
    return f'{province}{region}{number}领'


def generate_army_plate():
    army_codes = ''.join((c for c in string.ascii_uppercase if c in 'VKHBSLJNGCE'))
    region_codes = ''.join((c for c in string.ascii_uppercase if c not in 'I0'))
    army = random.choice(army_codes)
    region = random.choice(region_codes)
    number = ''.join(random.choices(string.digits, k=5))
    return f'{army}{region}{number}'


def generate_army_head_plate():
    head = random.choice(string.ascii_uppercase)
    head_number = random.choice('ADJRTVY')
    sequence = ''.join(random.choices(string.digits, k=5))
    return f'{head}{head_number}{sequence}'


def generate_ramdon_plate():
    # 随机选择车牌类型生成
    plate_types1 = {
        'regular': generate_reqular_plate,
        'new_energy_small': lambda: generate_new_energy_plate('small'),
        'new_energy_large': lambda: generate_new_energy_plate('large'),
        'police': generate_police_plate,
        'consulate': generate_consulate_plate,
        'army': generate_army_plate,
        'army_head': generate_army_head_plate
    }
    plate_types = {
        '常规车牌号': generate_reqular_plate,
        '小型新能源车牌号': lambda: generate_new_energy_plate('small'),
        '大型新能源车牌号': lambda: generate_new_energy_plate('large'),
        '警车车牌号': generate_police_plate,
        '领事馆车牌号': generate_consulate_plate,
        '武警车牌号': generate_army_plate,
        '军用车牌号': generate_army_head_plate
    }
    # 随机生成车牌
    plate_type = random.choice(list(plate_types.keys()))
    plate_number = plate_types[plate_type]()
    print(f"车牌号: {plate_number}，车牌类型: {plate_type}")
    return (plate_number, plate_type)


def generate_id_card():
    idCardList = ['450903199003131632', '450206195206027256', '4512231997090568394', '451321196403185690']
    return random.choice(idCardList)  # id cand number


## 在嵌套字典中查找指定键对应的值。
def find_value_in_nested_dict(dict_obj, key):
    if key in dict_obj:
        # 如果键在当前字典层级中，直接返口值
        return dict_obj[key]
    for k, v in dict_obj.items():
        # 如果值是字典，则递归查找
        if isinstance(v, dict):
            result = find_value_in_nested_dict(v, key)
            if result is not None:
                return result
    # 如果键在所有嵌套字典中都未找到，返回None
    return None


class SimplePostmanApp(tk.Tk):
    def __init__(self):
        super().__init__()
        # 显式调用update方法
        self.update()

        # 设置窗口背景颜色
        self.configure(bg='lightskyblue')

        try:
            # 设置桌面图标，这里需要提供一个ico件的路径
            self.iconbitmap(f'{current_script_path}/happy.ico')
        except:
            pass

        try:
            # 设置窗口图标
            self.wm_iconbitmap(bitmap=f'{current_script_path}/happy.ico')
        except:
            pass

        # 获取电脑分辨率
        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()
        # 设置坐标
        x = (screen_width - 900) // 2
        y = (screen_height - 800) // 2
        # 设置应用名
        self.title("simplepostman1.0")
        # 设置展示位置
        self.geometry(f"{900}x{800}+{x}+{y}")

        # 设置窗口自适应
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(5, weight=1)

        # 界面布局设置
        self.create_mainWin_widgets()

        # 初始化加载历史记录
        self.load_records()

        # 绑定快捷键《CtrL+Z)到撒操作的函数
        self.bind("<Control-z>", lambda event: self.undo_operation())
        # # 绑定快捷《CtrL+)到撤销操作的函数#
        # self.bind("<Control-z>", lambda event: self.undo_operation())

    def goto_line(self, text_widget, line_number):
        """"""
        # 创建一个消息提示子窗口
        text_widget.mark_set('insert', f'{line_number}.0')
        text_widget.see('insert')
        text_widget.focus()

    def messageInformInWin(self, informContent, lastingTime):
        """
        # 定义一个函数，用于创建子窗口并在2秒后关闭它
        """
        # 创建一个消息提示子窗口
        popup = tk.Toplevel(self)
        popup.title("消息提示")

        # 获取电脑分辨率
        screen_width = popup.winfo_screenwidth()
        screen_height = popup.winfo_screenheight()
        # 设置坐标
        x = (screen_width - 200) // 2
        y = (screen_height - 100) // 2
        # 设置展示位置
        popup.geometry(f"{200}x{100}+{x}+{y}")
        # 设置窗口自适应
        popup.grid_columnconfigure(1, weight=1)
        popup.grid_rowconfigure(5, weight=1)

        # 在子窗口中添加一些内容
        # label = tk.Label(popup, text=informContent)
        # label.pack(pady=20)
        popup.content_text = tk.Text(popup, bg='lightskyblue')
        popup.content_text.pack(fill=tk.BOTH, expand=True)

        popup.content_text.insert('insert', informContent)

        # 定义一个函数，用于关闭子窗口
        def close_window():
            popup.destroy()

        # lastingTime后调用关闭窗口的函数
        popup.after(lastingTime, close_window)

    def undo_operation(self):
        print("微销操作")
        # 在这里添加撒销操作的具体实现代码

    def askAI(self, qtext, atext, model="gpt-4o-mini"):
        """单次提问函数"""
        question=qtext.get("1.0", tk.END)
        delay = 1.5
        self.messageInformInWin("提问中......", 3000)
        try:
            print(f"\n正在处理问题: {question[:30]}...")

            completion = openai.chat.completions.create(
                model=model,
                messages=[{"role": "user", "content": question}],
                timeout=20  # 增加超时限制
            )
            answer = completion.choices[0].message.content
            # 打印当前结果
            print(f"【问题】{question}\n【回答】{answer}\n")
            self.clearContent(atext)
            atext.insert(tk.END,f'{answer}')
            sleep(delay)  # 避免频繁请求
        except Exception as e:
            answer = f"ERROR: {str(e)}"
            print(answer)
            atext.insert(tk.END, f'{answer}')
        #return answer

    def batch_askAI(self, questions, qtext, atext, model="gpt-4o-mini", delay=1.5):
        """批量提问函数"""
        results = []
        for idx, question in enumerate(questions, 1):
            try:
                print(f"\n正在处理问题 {idx}/{len(questions)}: {question[:30]}...")

                completion = openai.chat.completions.create(
                    model=model,
                    messages=[{"role": "user", "content": question}],
                    timeout=20  # 增加超时限制
                )

                answer = completion.choices[0].message.content
                results.append((question, answer))

                # 打印当前结果
                print(f"【问题】{question}\n【回答】{answer[:100]}...\n")
                sleep(delay)  # 避免频繁请求

            except Exception as e:
                print(f"问题 {idx} 处理失败: {str(e)}")
                results.append((question, f"ERROR: {str(e)}"))

        #return results

    def write_xmind_to_excel(self, which_combobox):
        # 设置文件类型过滤器，只显示XMind文件(.xmind)
        filetypes = [('XMind files', '*.xmind')]
        xmind_file_path = filedialog.askopenfilename(title="选择xmind文件", filetypes=filetypes)
        # 打开一个文件选择对话框，用户选择xmind文件
        if not isfile(xmind_file_path):
            # 检查所选文件是否存在
            print(f"文件不存在:{xmind_file_path}")
        else:
            pass

        caseType = which_combobox.get()

        # 将Xmind解析为多叉树
        xmind_tree = xmind_to_tree(xmind_file_path)

        # 获取所有叶子节点的路径
        all_leaf_paths_list = get_all_leaf_paths_data_list(xmind_tree, caseType)
        # print(all leaf_paths_list)

        write_xmind_data_to_excel(all_leaf_paths_list, caseType)  ###

        # 清空输入框的内容
        self.response_text.delete("1.0", "end-1c")
        # 输入框插入内容
        self.response_text.insert('insert', f'生成excel路径为: {current_script_path}/{caseExcelFileName}')

    def create_SetBodyKey_sub_window(self):
        # 创建设置入参子窗口
        sub_window = tk.Toplevel(self)
        try:
            # 设置子窗口图标
            sub_window.wm_iconbitmap(bitmap=f'{current_script_path}/happy.ico')
        except:
            pass
        # 设置窗口背景颜色
        sub_window.configure(bg='lightskyblue')
        # 获取电脑分辨率
        screen_width = sub_window.winfo_screenwidth()
        screen_height = sub_window.winfo_screenheight()
        # 设置坐标
        x = (screen_width - 400) // 2
        y = (screen_height - 300) // 2
        # 设置应用名
        sub_window.title("SetHeadersAndBodyKey")
        # 设置展示位置
        sub_window.geometry(f"{400}x{300}+{x + 655}+{y}")
        # 设置窗口自适应
        sub_window.grid_columnconfigure(1, weight=1)
        sub_window.grid_rowconfigure(5, weight=1)

        self.create_setBodyKey_sub_widgets(sub_window)

    def create_tools_sub_window(self):

        # 创建工具子窗口
        sub_window = tk.Toplevel(self)
        try:
            # 设置子窗口图标
            sub_window.wm_iconbitmap(bitmap=f'{current_script_path}/happy.ico')
        except:
            pass
        # 设置窗口背景颜色
        sub_window.configure(bg='lightskyblue')
        # 获取电脑分辨率
        screen_width = sub_window.winfo_screenwidth()
        screen_height = sub_window.winfo_screenheight()
        # 设置坐标
        x = (screen_width - 800) // 2
        y = (screen_height - 400) // 2
        # 设置应用名
        sub_window.title("tools")
        # 设置展示位置
        sub_window.geometry(f"{800}x{400}+{x}+{y}")
        # 设置窗口自适应
        sub_window.grid_columnconfigure(1, weight=1)
        sub_window.grid_rowconfigure(5, weight=1)
        # 在子窗口中添加标签
        # Label = tk.Label(sub_window，text="这是一个子窗口")
        # label.pack()
        self.create_tools_sub_widgets(sub_window)

    def create_tools_sub_window1(self):

        # 创建工具子窗口
        sub_window = tk.Toplevel(self)
        try:
            # 设置子窗口图标
            sub_window.wm_iconbitmap(bitmap=f'{current_script_path}/happy.ico')
        except:
            pass
        # 设置窗口背景颜色
        sub_window.configure(bg='lightskyblue')
        # 获取电脑分辨率
        screen_width = sub_window.winfo_screenwidth()
        screen_height = sub_window.winfo_screenheight()
        # 设置坐标
        x = (screen_width - 800) // 2
        y = (screen_height - 400) // 2
        # 设置应用名
        sub_window.title("ChangeImg")
        # 设置展示位置
        sub_window.geometry(f"{800}x{400}+{x}+{y}")
        # 设置窗口自适应
        sub_window.grid_columnconfigure(1, weight=1)
        sub_window.grid_rowconfigure(5, weight=1)
        # 在子窗口中添加标签
        # Label = tk.Label(sub_window，text="这是一个子窗口")
        # label.pack()
        self.create_tools_sub_widgets1(sub_window)

    def create_tools_sub_window2(self):

        # 创建工具子窗口
        sub_window = tk.Toplevel(self)
        try:
            # 设置子窗口图标
            sub_window.wm_iconbitmap(bitmap=f'{current_script_path}/happy.ico')
        except:
            pass
        # 设置窗口背景颜色
        sub_window.configure(bg='lightskyblue')
        # 获取电脑分辨率
        screen_width = sub_window.winfo_screenwidth()
        screen_height = sub_window.winfo_screenheight()
        # 设置坐标
        x = (screen_width - 1000) // 2
        y = (screen_height - 500) // 2
        # 设置应用名
        sub_window.title("SeeImgToTxt")
        # 设置窗口展示位置和大小
        sub_window.geometry(f"{1000}x{500}+{x}+{y}")
        # 设置窗口自适应
        sub_window.grid_columnconfigure(1, weight=1)
        sub_window.grid_rowconfigure(5, weight=1)
        # 在子窗口中添加标签
        # Label = tk.Label(sub_window，text="这是一个子窗口")
        # label.pack()
        self.create_tools_sub_widgets2(sub_window)

    def create_tools_sub_window3(self):

        # 创建工具子窗口
        sub_window = tk.Toplevel(self)
        try:
            # 设置子窗口图标
            sub_window.wm_iconbitmap(bitmap=f'{current_script_path}/happy.ico')
        except:
            pass
        # 设置窗口背景颜色
        sub_window.configure(bg='lightskyblue')
        # 获取电脑分辨率
        screen_width = sub_window.winfo_screenwidth()
        screen_height = sub_window.winfo_screenheight()
        # 设置坐标
        x = (screen_width - 1000) // 2
        y = (screen_height - 500) // 2
        # 设置应用名
        sub_window.title("QaByAI")
        # 设置窗口展示位置和大小
        sub_window.geometry(f"{1000}x{500}+{x}+{y}")
        # 设置窗口自适应
        sub_window.grid_columnconfigure(1, weight=1)
        sub_window.grid_rowconfigure(5, weight=1)
        # 在子窗口中添加标签
        # Label = tk.Label(sub_window，text="这是一个子窗口")
        # label.pack()
        self.create_tools_sub_widgets3(sub_window)

    def change_picture_format(self):

        # # 创建设置子窗口
        # sub_window = tk.Toplevel(self)
        # # 不显示子窗口
        # sub_window.withdraw()
        # 设置只打开图片文件
        filetypes = [('Image files', ',jpg ,jpeg .png ,gif .bmp')]
        img_path = filedialog.askopenfilename(title="选择图片文件", filetypes=filetypes)
        # 打开一个文件选择对话框，用户选择图片文件
        if not isfile(img_path):
            # 检查所选文件是否存在
            print(f"文件不存在:{img_path}")
        else:
            img = Image.open(img_path)
            # 打开图片文件
            icon_sizes = [(64, 64)]
            # 定义图标大小
            base_name = splitext(img_path)[0]
            # 提取图像文件基本名称《不包括扩展名)
            icon_filename = f'{base_name}.ico'
            img.save(icon_filename, sizes=icon_sizes)
            print(f"图标文件已保存为:{icon_filename}")

    def create_mainWin_widgets(self):
        # URL输入
        self.url_label = tk.Label(self, text="URL:")
        self.url_label.grid(column=0, row=0)

        self.url_entry = tk.Entry(self, bg="lightskyblue")
        self.url_entry.grid(column=1, row=0, sticky='EW')  # , columnspan=2)

        self.set_button = ttk.Button(self, text="Setkey", command=self.create_SetBodyKey_sub_window)  # 组件按钮格式化功能
        self.set_button.grid(column=2, row=0)

        # 请求方法选择
        self.method_label = ttk.Label(self, text="Method:")
        self.method_label.grid(column=0, row=1)

        self.method_combobox = ttk.Combobox(self, values=["GET", "POST", "PUT", "DELETE"])
        self.method_combobox.grid(column=1, row=1, sticky='EW')
        self.method_combobox.current(0)
        # 设置ttk.Combobox的背景色
        # self.method_combobox.configure(style='TCombobox.Tooltip.TkDefault')

        # 选择前置登录接口记录的路径
        self.before_login_buton = ttk.Button(self, text="BeforeLoginpath",
                                             command=self.findLoginRecordFilePath)  # 组件按钮格式化功能
        self.before_login_buton.grid(column=2, row=1)

        self.before_login_combobox = ttk.Combobox(self,
                                                  values=[f'{current_script_path}/reguests_log/post数字员工请求登录demo.json'])
        self.before_login_combobox.grid(column=3, row=1, sticky='EW')
        self.before_login_combobox.current(0)

        # Headers输入
        self.headers_label = ttk.Label(self, text="Headers (JSON):")
        self.headers_label.grid(column=0, row=2)

        self.headers_text = scrolledtext.ScrolledText(self, width=45, height=10, bg="lightskyblue")
        self.headers_text.grid(column=1, row=2, rowspan=2, sticky='NSEW')

        self.headers_operation_combobox = ttk.Combobox(self, values=['Format', 'Copy', 'Paste', 'Clear'],
                                                       state='readonly')  # 组件按钮格式化功能
        self.headers_operation_combobox.grid(column=2, row=2)
        self.headers_operation_combobox.current(0)
        self.operation_headers_button = ttk.Button(self, text="Do",
                                                   command=lambda: self.combinateCommonOperation(self,self.headers_text,
                                                                                                 self.headers_operation_combobox))  # 组件按钮格式化功能
        self.operation_headers_button.grid(column=3, row=2)

        # self.format_headers_button = ttk.Button(self, text="Format Headers", command=lambda:self.format_content(self.headers_text))# 组件按钮格式化功能
        # self.format_headers_button.grid(column=2, row=2)
        #
        # self.copy_headers_button = ttk.Button(self,text='Copy', command=lambda:self.copy_content(self.headers_text))#组件按钮格式化功能
        # self.copy_headers_button.grid(column=3,row=2)
        #
        # # 清空响应结果
        # self.clear_headers_btn = tk.Button(self, text="Clear", command=lambda: self.clear(self.headers_text)) # 带参
        # self.clear_headers_btn.grid(column=2,row=3)
        #
        # # 粘贴板内容带入并清空响应结果
        # self.paste_body_btn = tk.Button(self, text="Paste", command=lambda: self.paste(self, self.headers_text))  # 带
        # self.paste_body_btn.grid(column=3, row=3)

        # Body输入
        self.body_label = ttk.Label(self, text="Body (JSON) :")
        self.body_label.grid(column=0, row=4)

        self.body_text = scrolledtext.ScrolledText(self, width=45, height=15, bg="lightskyblue")
        self.body_text.grid(column=1, row=4, rowspan=2, sticky='NSEW')

        self.body_operation_combobox = ttk.Combobox(self, values=['Format', 'Copy', 'Paste', 'Clear'],
                                                    state='readonly')  # 组件按钮格式化功能
        self.body_operation_combobox.grid(column=2, row=4)
        self.body_operation_combobox.current(0)
        self.operation_body_button = ttk.Button(self, text="Do",
                                                command=lambda: self.combinateCommonOperation(self, self.body_text,
                                                                                              self.body_operation_combobox))  # 组件按钮格式化功能
        self.operation_body_button.grid(column=3, row=4)

        # self.fommat_body_button = ttk.Button(self, text="Format Body", command=lambda:self.format_content(self.body_text)) # 组件按钮格式化功能
        # self.fommat_body_button.grid(column=2,row=4)
        #
        # self.copy_body_button = ttk.Button(self, text="Copy", command=lambda:self.copy_content (self.body_text))  # 组件按钮格式化功能
        # self.copy_body_button.grid(column=3,row=4)
        #
        # # 清空响应结果
        # self.clear_body_btn = tk.Button(self, text="Clear", command=lambda: self.clear(self.body_text)) # 带参
        # self.clear_body_btn.grid(column=2,row=5)
        #
        # # 粘贴板内容带入并清空响应结果
        # self.paste_body_btn = tk.Button(self, text="paste",command=lambda: self.paste(self, self.body_text)) # 带参
        # self.paste_body_btn.grid(column=3, row=5)

        from tkinter.font import Font
        bold_font = Font(family='Helvetica', size=10, weight='bold')

        # 发送按钮
        self.send_label = ttk.Label(self, text="Send Request:")
        self.send_label.grid(column=0, row=6)

        # self,send_button = ttk,Button(self,text="Send", command=self.send_request)不加线程的话会一直不回弹,导致主线程卡死
        self.send_button = tk.Button(self, text="Send", bg='#FFB6C1', fg='black', font=bold_font,
                                     command=lambda: self.thread_it(self.send_request_by_userInputData))  # 加线程
        self.send_button.grid(column=1, row=6, sticky='EW')

        # #复制案件号
        # self.copy_response_key_btn = tk.Button(self, texta"Copy ReportNo", anchor="c", command=lambda: self.copy_value_of_key('reportlo')#
        # self.copy_response_key_btn.grid(column=2, row=4)

        # 复制关键字对应的值
        self.select_key_btn = tk.Button(self, text="CopyResKey", anchor="c",
                                        command=lambda: self.copy_value_of_key(self, self.response_text,
                                                                               self.select_key_combobox.get()))
        self.select_key_btn.grid(column=2, row=7)

        # 设置关键字输入框
        self.select_key_combobox = ttk.Combobox(self, state="NORMAL")
        self.select_key_combobox.grid(column=3, row=7, sticky='EW')
        self.select_key_combobox["values"] = ['reportNo']
        self.select_key_combobox["values"] = find_value_in_nested_dict(
            read_json_file(f'{current_script_path}/configini.json'), "CopyKey")
        self.select_key_combobox.current(0)  # 设置默认值为列表中的第一个元素

        self.refresh_response_text_key_button = ttk.Button(self, text="RefreshKey",
                                                           command=lambda: self.refreshKey(self.response_text,
                                                                                           self.select_key_combobox))
        self.refresh_response_text_key_button.grid(column=2, row=6)

        # Response响应输出
        self.response_label = ttk.Label(self, text="Response (JSON):")
        self.response_label.grid(column=0, row=7)

        self.response_text = scrolledtext.ScrolledText(self, width=45, height=30, bg="lightskyblue")
        self.response_text.grid(column=1, row=7, rowspan=13, sticky='NSEW')  # ,columnspan=3

        self.response_operation_combobox = ttk.Combobox(self, values=['Format', 'Copy', 'Paste', 'Clear'],
                                                        state='readonly')  # 组件按钮格式化功能
        self.response_operation_combobox.grid(column=2, row=8)
        self.response_operation_combobox.current(3)
        self.operation_response_button = ttk.Button(self, text="Do", command=lambda: self.combinateCommonOperation(self,
                                                                                                                   self.response_text,
                                                                                                                   self.response_operation_combobox))  # 组件按钮格式化功能
        self.operation_response_button.grid(column=3, row=8)

        # self.format_response_button = tk.Button(self, text='Format Response', anchor='c', command=lambda: self.format_content(self.response_text))  # 组件按钮格式化功能
        # self.format_response_button.grid(column=2,row=8)
        #
        # self.format_response_button = tk.Button(self, text="Copy",anchor='c',command=lambda:self.copy_content(self.response_text))# 组件按钮格式化功能
        # self.format_response_button.grid(column=3,row=8)
        #
        # # 清空响应结果
        # self.clear_response_btn = tk.Button(self, text="Clear", command=lambda: self.clear(self.response_text))  # 带参
        # #self.clear_response_btn = tk.Button(self, text="Clear", command=self.clear)#不带参数
        # self.clear_response_btn.grid(column=2,row=9)
        #
        # # 粘贴板内容带入并清空响应结果
        # self.paste_response_btn = tk.Button(self, text="paste", command=lambda: self.paste(self, self.response_text))  # 带参
        # self.paste_response_btn.grid(column=3,row=9)

        # 搜索按钮
        self.search_btn = tk.Button(self, text="FindReskey", anchor="c",
                                    command=lambda: self.search_text(self.search_combobox.get()))
        self.search_btn.grid(column=2, row=10)
        # 搜索response关键字结果
        self.search_combobox = ttk.Combobox(self, state='NORMAL')
        self.search_combobox.grid(column=3, row=10, sticky='EW')
        self.search_combobox['values'] = ['CollegeName']
        # self.search_combobox['values'] = find_value_of_key_in_nested_dict(read_json_file(f'{current_script_path}/configini.json'), "Searchkey")# list
        self.search_combobox.current(0)  # 设置默认值为列表中的第一个元素

        self.set_button_Tools = tk.Button(self, text="Tools", command=self.create_tools_sub_window)  # 组件按钮格式化功能
        self.set_button_Tools.grid(column=2,
                             row=11)  # self,set button = tk.Button(self, text="json",command=self,find_file_to_fill_record) # 组件按纽定将式化函教功能# self.set button.grid(column=2，row=1)

        self.set_button_ChangeImg = tk.Button(self, text="ChangeImg", command=self.create_tools_sub_window1)  # 组件按钮格式化功能
        self.set_button_ChangeImg.grid(column=2,
                             row=12)  # self,set button = tk.Button(self, text="json",command=self,find_file_to_fill_record) # 组件按纽定将式化函教功能# self.set button.grid(column=2，row=1)

        self.set_button_SeeImgToTxt = tk.Button(self, text="SeeImgToTxt", command=self.create_tools_sub_window2)  # 组件按钮格式化功能
        self.set_button_SeeImgToTxt.grid(column=2,
                             row=13)  # self,set button = tk.Button(self, text="json",command=self,find_file_to_fill_record) # 组件按纽定将式化函教功能# self.set button.grid(column=2，row=1)

        self.set_button_QaByAI = tk.Button(self, text="QaByAI", command=self.create_tools_sub_window3)  # 组件按钮格式化功能
        self.set_button_QaByAI.grid(column=2,
                             row=14)  # self,set button = tk.Button(self, text="json",command=self,find_file_to_fill_record) # 组件按纽定将式化函教功能# self.set button.grid(column=2，row=1)

        self.select_record_button = tk.Button(self, text="Select Records:", anchor='c',
                                              command=self.find_file_to_fill_record)  #
        self.select_record_button.grid(column=0, row=20)
        self.record_combobox = ttk.Combobox(self, state='readonly', width=50)
        self.record_combobox.grid(column=1, row=20, sticky='EW')
        self.record_combobox.bind('<<ComboboxSelected>>', self.fill_record)

        self.delete_record_button = tk.Button(self, text="Delete Records", anchor='c',
                                              command=lambda: self.thread_it(self.delete_records))
        self.delete_record_button.grid(column=2, row=20)

        # 为匹配文本定义样式
        self.response_text.tag_config("match", background="yellow")  #

    def create_setBodyKey_sub_widgets(self, sub_win):
        # surveyUm标签
        sub_win.surveyUm_label = ttk.Label(sub_win, text="surveyUm:")
        sub_win.surveyUm_label.grid(column=0, row=0)
        # 选择surveyUm
        sub_win.surveyUm_combobox = ttk.Combobox(sub_win, state='NORMAL')
        sub_win.surveyUm_combobox.grid(column=1, row=0, sticky='Ew')
        # sub_win.surveyUm_combobox['values'] =['CHENKETAN6096','HUJUN464']

        # 获取配置里的UM账号
        UMList = read_json_file(f'{current_script_path}/configini.json')['UM']
        sub_win.surveyUm_combobox['values'] = UMList

        sub_win.surveyUm_combobox.current(0)  # 设置默认值为列表中的第一个元素
        print(sub_win.surveyUm_combobox.get())

        # 保存surveyUm
        sub_win.set_surveyUm_button = ttk.Button(sub_win, text="save",
                                                 command=lambda: self.changeKeyValueInJsonFile('surveyUm',
                                                                                               sub_win.surveyUm_combobox.get()))  # 组件按期定格式
        sub_win.set_surveyUm_button.grid(column=2, row=0)

        # surveyUmphone标签
        sub_win.surveyUmphone_label = ttk.Label(sub_win, text="surveyUmphone:")
        sub_win.surveyUmphone_label.grid(column=0, row=1)
        # 选择surveyUmphone
        sub_win.surveyUmphone_combobox = ttk.Combobox(sub_win, state='NORMAL')
        sub_win.surveyUmphone_combobox.grid(column=1, row=1, sticky='EW')
        sub_win.surveyUmphone_combobox['values'] = ['17374899426']
        sub_win.surveyUmphone_combobox.current(0)  # 设置默认值为列表中的第一个元素
        # 保存surveyUmphone
        sub_win.set_surveyumphone_button = ttk.Button(sub_win, text="save",
                                                      command=lambda: self.changeKeyValueInJsonFile("surveyumphone",
                                                                                                    sub_win.surveyUmphone_combobox.get()))
        sub_win.set_surveyumphone_button.grid(column=2, row=1)

        # mobileNo标签
        sub_win.mobileNo_label = ttk.Label(sub_win, text="mobileNo:")
        sub_win.mobileNo_label.grid(column=0, row=2)
        # 选择mobileNo
        sub_win.mobileNo_combobox = ttk.Combobox(sub_win, state='NORMAL')
        sub_win.mobileNo_combobox.grid(column=1, row=2, sticky='EW')
        sub_win.mobileNo_combobox['values'] = ['17374899426']
        sub_win.mobileNo_combobox.current(0)  # 设置默认值为列表中的第一个元素
        # 保存mobileNo
        sub_win.set_mobileNo_button = ttk.Button(sub_win, text="save",
                                                 command=lambda: self.changeKeyValueInJsonFile('mobileNo',
                                                                                               sub_win.mobileNo_combobox.get()))  # 组件按钮绑定格式化函数功能
        sub_win.set_mobileNo_button.grid(column=2, row=2)

        # 选取修改headers里的键值
        sub_win.headers_text_key_label = ttk.Label(sub_win, text="selectHeadersKey:")
        sub_win.headers_text_key_label.grid(column=0, row=10)
        # 选择headers修改的键
        sub_win.headers_text_key_combobox = ttk.Combobox(sub_win, state='NORMAL')
        sub_win.headers_text_key_combobox.grid(column=1, row=10, sticky='EW')
        sub_win.headers_text_key_combobox["values"] = ["Authorization", "Content-Type", "X-Portal-Token", "Cookie"]
        sub_win.headers_text_key_combobox.current(0)  # 设置默认值为列表中的第一个元素
        # 刷新键值
        sub_win.refresh_headers_text_keyValue_button = ttk.Button(sub_win, text="refreshKey",
                                                                  command=lambda: self.refreshKey(self.headers_text,
                                                                                                  sub_win.headers_text_key_combobox))
        sub_win.refresh_headers_text_keyValue_button.grid(column=2, row=10)

        # 修改headers里的键值
        sub_win.headers_text_keyValue_label = ttk.Label(sub_win, text="updateHeadersKey:")
        sub_win.headers_text_keyValue_label.grid(column=0, row=11)
        # 选择headers修改
        sub_win.headers_text_keyValue_combobox = ttk.Combobox(sub_win, state='NORMAL')
        sub_win.headers_text_keyValue_combobox.grid(column=1, row=11, sticky='EW')
        sub_win.headers_text_keyValue_combobox['values'] = ['456']
        sub_win.headers_text_keyValue_combobox.current(0)  # 设置默认值为列表中的第一个元素
        # 更改键值
        sub_win.set_headers_text_keyValue_button = ttk.Button(sub_win, text="change",
                                                              command=lambda: self.changeKeyValueInJsonstr(
                                                                  self.headers_text,
                                                                  sub_win.headers_text_key_combobox.get(),
                                                                  sub_win.headers_text_keyValue_combobox.get()))
        sub_win.set_headers_text_keyValue_button.grid(column=2, row=11)

        # 选取修改body里的键值
        sub_win.body_text_key_label = ttk.Label(sub_win, text="selectBodyKey:")
        sub_win.body_text_key_label.grid(column=0, row=12)
        # 选择body修改的键
        sub_win.body_text_key_combobox = ttk.Combobox(sub_win, state="NORMAL")
        sub_win.body_text_key_combobox.grid(column=1, row=12, sticky='EW')
        sub_win.body_text_key_combobox['values'] = ["reportNo", "businessKey"]
        sub_win.body_text_key_combobox.current(0)  # 设置默认值为列表中的第一个元素
        # 刷新键值
        sub_win.refresh_body_text_keyValue_button = ttk.Button(sub_win, text="refreshkey",
                                                               command=lambda: self.refreshKey(self.body_text,
                                                                                               sub_win.body_text_key_combobox))
        sub_win.refresh_body_text_keyValue_button.grid(column=2, row=12)

        # 修改body里的键值
        sub_win.body_text_keyValue_label = ttk.Label(sub_win, text="updateBodyKey:")
        sub_win.body_text_keyValue_label.grid(column=0, row=13)
        # 选择body修改的键
        sub_win.body_text_keyValue_combobox = ttk.Combobox(sub_win, state='NORMAL')
        sub_win.body_text_keyValue_combobox.grid(column=1, row=13, sticky='EW')
        sub_win.body_text_keyValue_combobox['values'] = ['456']
        sub_win.body_text_keyValue_combobox.current(0)  # 设置默认值为列表中的第一个元素
        # 更改键值
        sub_win.set_body_text_keyValue_button = ttk.Button(sub_win, text="change",
                                                           command=lambda: self.changeKeyValueInJsonstr(self.body_text,
                                                                                                        sub_win.body_text_key_combobox.get(),
                                                                                                        sub_win.body_text_keyValue_combobox.get()))
        sub_win.set_body_text_keyValue_button.grid(column=2, row=13)

    def create_tools_sub_widgets(self, sub_win):
        sub_win.output_text = scrolledtext.ScrolledText(sub_win, width=30, height=20, bg="lightskyblue")
        sub_win.output_text.grid(column=0, row=0, rowspan=15, sticky='NSEW')

        sub_win.getHtmlData_button = tk.Button(sub_win, text="GetHtmlData", anchor="c", command=lambda: self.thread_it(
            self.getHtmlData(sub_win.output_text, sub_win.output_text.get("1.0", "end-1c"))))
        sub_win.getHtmlData_button.grid(column=1, row=0)

        sub_win.getPythonCode_button = tk.Button(sub_win, text="GetPyCode", anchor="c", command=lambda: self.thread_it(
            self.restore_python_code_by_AI(sub_win.output_text, sub_win.output_text.get("1.0", "end-1c"))))
        sub_win.getPythonCode_button.grid(column=1, row=1)

        sub_win.getTestData_button = tk.Button(sub_win, text="GetTestData", anchor="c", command=lambda: self.thread_it(
            self.getTestData(sub_win.output_text, sub_win.output_text.get("1.0", "end-1c"))))
        sub_win.getTestData_button.grid(column=1, row=2)

        # 设置加密按钮
        sub_win.encrypt_btn = tk.Label(sub_win, text="EncryptChar", anchor='c')
        sub_win.encrypt_btn.grid(column=1, row=3)

        # 设置加密符
        sub_win.encrypt_combobox = ttk.Combobox(sub_win, state='NORMAL')
        sub_win.encrypt_combobox.grid(column=2, row=3)
        sub_win.encrypt_combobox['values'] = ['测', '测试']
        # sub_win.encrypt_combobox["values"] = find_value_of_key_in_nested_dict((read_json_file(f'{current_script_path}/configini.json')),  "Encrypt")
        # print(read_json_file(f'current_script_path}/configini.json'))
        sub_win.encrypt_combobox.current(0)  # 设置默认值为列表中的第一个元素 format_json

        # 复制关键字对应的值
        sub_win.select_key_btn1 = tk.Button(sub_win, text='GetTxtKeyCopy', anchor='c',
                                            command=lambda: self.copy_value_of_key(sub_win, sub_win.output_text,
                                                                                   sub_win.select_key_combobox1.get()))
        sub_win.select_key_btn1.grid(column=1, row=4)

        # 设置关键字输入框
        sub_win.select_key_combobox1 = ttk.Combobox(sub_win, state='NORMAL')
        sub_win.select_key_combobox1.grid(column=2, row=4, sticky='EW')
        sub_win.select_key_combobox1['values'] = [
            '123']  # find_value_of_key_in_nested_dict((read_json_file(f'{current_script_path}/configini.json')),"CopyKey")
        sub_win.select_key_combobox1.current(0)  # 设置默认值为列表中的第一个元素

        sub_win.output_operation_combobox = ttk.Combobox(sub_win, values=['Format', 'Copy', 'Paste', 'Clear'],
                                                         state='readonly')  # 组件按钮格式化功能
        sub_win.output_operation_combobox.grid(column=1, row=5)
        sub_win.output_operation_combobox.current(0)
        sub_win.operation_output_button = ttk.Button(sub_win, text="Do",
                                                     command=lambda: self.combinateCommonOperation(sub_win,
                                                                                                   sub_win.output_text,
                                                                                                   sub_win.output_operation_combobox))  # 组件按钮格式化功能
        sub_win.operation_output_button.grid(column=2, row=5)

        # # Format格式化json展示
        # sub_win.format_btn = ttk.Button(sub_win, text="Format", command=lambda: self.format_content(sub_win.outout_text))
        # sub_win.format_btn.grid(column=1, row=5)
        #
        # # 复制结果
        # sub_win.copy_output_btn = ttk.Button(sub_win,text="Copy",command=lambda: self.copy_content(sub_win.output_text)) #带参数
        # sub_win.copy_output_btn.grid(column=2, row=5)
        #
        # # 清空结果
        # sub_win.clear_output_btn = ttk.Button(sub_win,text="Clear", command=lambda: self.clear(sub_win.output_text)) #带参数
        # sub_win.clear_output_btn.grid(column=1, row=6)
        #
        # # 粘贴板内容带入并先前清空结果
        # sub_win.paste_output_btn = ttk.Button(sub_win,text="Paste", command=lambda: self.paste(sub_win, sub_win.output_text)) # 带参数
        # sub_win.paste_output_btn.grid(column=2, row=6)

        # 转化照片格式为ico格式
        sub_win.change_picture_format_btn = ttk.Button(sub_win, text="ChangePictureFormat",
                                                       command=self.change_picture_format)  # 带参数
        sub_win.change_picture_format_btn.grid(column=1, row=7)

        # 选择用例
        sub_win.choose_case_combobox = ttk.Combobox(sub_win, values=['普通用例', '版本用例', '回归用例'], state='readonly')  # 带参数
        sub_win.choose_case_combobox.grid(column=1, row=8)
        sub_win.choose_case_combobox.current(0)

        # xmind转化excel
        sub_win.WriteXmindToExcel_btn = ttk.Button(sub_win, text="ChooseAndWriteXmindToExcel",
                                                   command=lambda: self.write_xmind_to_excel(
                                                       sub_win.choose_case_combobox))  # 带参数
        sub_win.WriteXmindToExcel_btn.grid(column=2, row=8)

    def create_tools_sub_widgets1(self, sub_win):
        # 准备一些图片路径作为示例
        image_paths = [
            "test"
        ]

        # 创建一个 Combobox 组件用于选择图片路径
        combobox = ttk.Combobox(sub_win, values=image_paths, state="readonly")
        combobox.grid(column=2, row=8)
        combobox.current(0)  # 默认选择第一项

        # 创建一个 Label 组件用于显示图片
        image_label = tk.Label(sub_win, width=1200, height=800)
        image_label.grid(column=0, row=0, columnspan=2, rowspan=8, sticky='NSEW')

        # 测试选择标签
        sub_win.test_label = ttk.Button(sub_win, text="test:")
        sub_win.test_label.grid(column=2, row=0)
        sub_win.test_combobox = ttk.Combobox(sub_win, state="NORMAL")
        sub_win.test_combobox.grid(column=2, row=1, sticky='EW')
        sub_win.test_combobox['values'] = ['test', '123']
        sub_win.test_combobox.current(0)

        # 用于存储当前显示的图片的 PhotoImage 对象
        current_photo = None

        # 加载图片并调整大小以适应 Label
        def load_and_resize_image(path, label):
            try:
                # 使用 PIL 库加载图片
                image = Image.open(path)
                # 调整图片大小以适应 Label
                image = image.resize((label.winfo_width(), label.winfo_height()), Image.ANTIALIAS)
                # 创建 PhotoImage 对象
                photo = ImageTk.PhotoImage(image)
                # 更新 Label 显示图片
                label.configure(image=photo)
                global current_photo
                current_photo = photo
            except IOError:
                print(f"无法加载图片：{path}")

        # 根据下拉框的选择更新图片
        def update_image(event):
            # 获取下拉框的值，即图片路径
            selected_path = f'{current_script_path}/{fileName}/{combobox.get()}.png'

            new_selected_path = selected_path

            # 如果路径有效，则加载并显示图片
            load_and_resize_image(new_selected_path, image_label)

        # 绑定下拉框的事件
        combobox.bind("<<ComboboxSelected>>", update_image)

    def resize_image_to_fit(self, image_path, max_width, max_height):
        """将图片等比缩放至适应Label尺寸"""
        img = Image.open(image_path)
        img.thumbnail((max_width, max_height), Image.LANCZOS)  # 保持比例缩放
        return ImageTk.PhotoImage(img)

    def create_tools_sub_widgets2(self, sub_win):
        """创建图片OCR识别界面布局（左图右文）- 支持图片自适应填充"""
        # --- 主框架（保持不变）---
        main_frame = tk.Frame(sub_win)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # ===== 左侧图片展示区域（关键修改） =====
        img_frame = tk.Frame(main_frame, bd=2, relief=tk.SUNKEN)
        img_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5, pady=5)

        self.img_label = tk.Label(img_frame, text="图片预览区域", bg='#f0f0f0')
        self.img_label.pack(fill=tk.BOTH, expand=True) # tk.BOTH

        # --- 图片自适应逻辑 ---
        def load_image_to_label(image_path):
            """加载图片并自适应填充Label"""
            try:
                # 1. 用PIL打开图片并保存原始对象
                self.original_image = Image.open(image_path)
                # 2. 立即更新显示（处理初始加载）
                self.update_image_display()
            except Exception as e:
                self.img_label.config(text=f"  图片加载失败: {str(e)}", image='')

        def update_image_display(event=None):
            """动态调整图片尺寸以适应Label（强制填满区域，允许裁剪）"""
            if hasattr(self, 'original_image'):
                # 获取Label当前有效尺寸
                label_width = max(1, self.img_label.winfo_width())
                label_height = max(1, self.img_label.winfo_height())

        # 绑定Label尺寸变化事件
        self.img_label.bind("<Configure>", update_image_display)

        # ===== 右侧功能区域 =====
        right_frame = tk.Frame(main_frame)
        right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)

        # ===== 右侧OCR结果区域（保持不变） =====
        text_frame = tk.Frame(main_frame)
        text_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=5, pady=5)

        scrollbar = tk.Scrollbar(text_frame)
        self.text_output = tk.Text(text_frame, wrap=tk.WORD, yscrollcommand=scrollbar.set,
                                   font=('Consolas', 10), padx=5, pady=5)
        scrollbar.config(command=self.text_output.yview)

        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.text_output.pack(fill=tk.BOTH, expand=True)
        self.text_output.insert(tk.END, "OCR识别结果将显示在这里...")

        # --- 新增操作按钮组 ---
        btn_group = tk.Frame(right_frame, bd=1, relief=tk.RAISED)
        btn_group.pack(fill=tk.X, pady=5)

        # 操作类型选择
        self.output_operation = ttk.Combobox(
            btn_group,
            values=['Format', 'Copy', 'Paste', 'Clear'],
            state='readonly'
        )
        self.output_operation.pack(side=tk.LEFT, padx=5, pady=2)
        self.output_operation.current(1)

        # 执行按钮
        ttk.Button(
            btn_group, text="Execute",
            command=lambda: self.combinateCommonOperation(sub_win,
                                                          self.text_output,
                                                          self.output_operation)
        ).pack(side=tk.LEFT, padx=5)

        # ===== 底部按钮区域（保持不变） =====
        btn_frame = tk.Frame(sub_win)
        btn_frame.pack(fill=tk.X, padx=5, pady=5)

        tk.Label(btn_frame, text="识别语言:").pack(side=tk.LEFT, padx=(0, 5))
        self.lang_combo = ttk.Combobox(btn_frame, values=["中文", "英文", "日文", "韩文", "阿拉伯文"])
        self.lang_combo.pack(side=tk.LEFT, padx=(0, 10))
        self.lang_combo.current(0)

        tk.Label(btn_frame, text="识别内容:").pack(side=tk.LEFT, padx=(0, 5))
        self.lang_combo1 = ttk.Combobox(btn_frame, values=["文本", "定位"])
        self.lang_combo1.pack(side=tk.LEFT, padx=(0, 10))
        self.lang_combo1.current(0)

        # 修改原按钮命令：调用load_image_to_label而非直接操作Label
        select_btn = ttk.Button(btn_frame, text="选择图片",
                                command=lambda: self.select_image_for_ocr_tk())
        select_btn.pack(side=tk.RIGHT)

    def select_image_for_ocr_tk(self):
        """Tkinter版本的选择图片方法"""
        img_path = filedialog.askopenfilename(
            title="选择图片",
            filetypes=[("图片文件", "*.png *.jpg *.jpeg *.bmp")]
        )

        if img_path:
            try:
                # 显示图片
                img = Image.open(img_path)
                img.thumbnail((400, 400))  # 限制显示大小
                photo = ImageTk.PhotoImage(img)

                self.img_label.config(image=photo)
                self.img_label.image = photo  # 保持引用

                # 调用OCR识别  param lang must in dict_keys(['ch', 'en', 'korean', 'japan', 'chinese_cht', 'ta', 'te', 'ka', 'latin', 'arabic', 'cyrillic', 'devanagari']), but got chi_sim
                lang_map = {"中文": "ch", "英文": "en",
                            "日文": "japan", "韩文": "korean", "阿拉伯文": "arabic"}
                selected_lang = self.lang_combo.get()
                lang = lang_map[selected_lang]
                pattern = self.lang_combo1.get()

                # 调用OCR方法
                self.seeImgToTxtByPaddleOcr(self.text_output, img_path, lang, pattern)

            except Exception as e:
                messagebox.showerror(" 错误", f"图片处理失败: {str(e)}")
    def select_image_for_ocr_qa(self):
        """Tkinter版本的选择图片方法"""
        img_path = filedialog.askopenfilename(
            title="选择图片",
            filetypes=[("图片文件", "*.png *.jpg *.jpeg *.bmp")]
        )

        if img_path:
            try:
                # # 显示图片
                # img = Image.open(img_path)
                # img.thumbnail((400, 400))  # 限制显示大小
                # photo = ImageTk.PhotoImage(img)
                #
                # self.img_label.config(image=photo)
                # self.img_label.image = photo  # 保持引用

                # 调用OCR识别  param lang must in dict_keys(['ch', 'en', 'korean', 'japan', 'chinese_cht', 'ta', 'te', 'ka', 'latin', 'arabic', 'cyrillic', 'devanagari']), but got chi_sim
                lang_map = {"中文": "ch", "英文": "en",
                            "日文": "japan", "韩文": "korean", "阿拉伯文": "arabic"}
                selected_lang = self.lang_combo.get()
                lang = lang_map[selected_lang]

                # 调用OCR方法
                self.seeImgToTxtByPaddleOcr1(self.left_text_output, img_path, lang)

                self.askAI(self.left_text_output, self.right_text_output,"gpt-4o-mini")

            except Exception as e:
                messagebox.showerror(" 错误", f"图片处理失败: {str(e)}")

    """
    上述代码不符合要求，选择插入的图片未能自适应填充满img_frame区域，要求不改变代码的基本功能进行修改
    """

    def create_tools_sub_widgets3(self, sub_win):
        """创建图片OCR识别界面布局（左图右文）- 支持图片自适应填充"""
        # --- 第一行主框架 ---
        main_frame = tk.Frame(sub_win)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # ===== 左侧输入问题区域 =====
        left_text_frame = tk.Frame(main_frame)
        left_text_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5, pady=5)

        scrollbar = tk.Scrollbar(left_text_frame)
        self.left_text_output = tk.Text(left_text_frame, wrap=tk.WORD, yscrollcommand=scrollbar.set,
                                   font=('Consolas', 10), padx=5, pady=5)
        scrollbar.config(command=self.left_text_output.yview)

        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.left_text_output.pack(fill=tk.BOTH, expand=True)
        self.left_text_output.insert(tk.END, " ") #请输入问题.......

        # ===== 右侧RIGHT输出结果区域 =====
        right_text_frame = tk.Frame(main_frame)
        right_text_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=5, pady=5)

        scrollbar = tk.Scrollbar(right_text_frame)
        self.right_text_output = tk.Text(right_text_frame, wrap=tk.WORD, yscrollcommand=scrollbar.set,
                                   font=('Consolas', 10), padx=5, pady=5)
        scrollbar.config(command=self.right_text_output.yview)

        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.right_text_output.pack(fill=tk.BOTH, expand=True)
        self.right_text_output.insert(tk.END, "")

        # --- 第二行主框架 ---
        main_frame1 = tk.Frame(sub_win)
        main_frame1.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # ===== 第二行左侧功能区域 =====
        left_frame = tk.Frame(main_frame1)
        left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # --- 新增操作按钮组 ---
        btn_group = tk.Frame(left_frame, bd=1, relief=tk.RAISED)
        btn_group.pack(fill=tk.X, pady=5)

        # 操作类型选择
        self.output_operation = ttk.Combobox(
            btn_group,
            values=['Format', 'Copy', 'Paste', 'Clear'],
            state='readonly'
        )
        self.output_operation.pack(side=tk.LEFT, padx=5, pady=2)
        self.output_operation.current(2)

        # 执行按钮
        ttk.Button(
            btn_group, text="Execute",
            command=lambda: self.combinateCommonOperation(sub_win,
                                                          self.left_text_output,
                                                          self.output_operation)
        ).pack(side=tk.LEFT, padx=5)

        # --- 提问按钮 ---
        btn_group1 = tk.Frame(left_frame, bd=1, relief=tk.RAISED)
        btn_group1.pack(fill=tk.X, pady=5)

        # 提问按钮
        ttk.Button(btn_group1, text="提问",command=lambda: self.askAI(self.left_text_output,self.right_text_output,"gpt-4o-mini")).pack(side=tk.LEFT, padx=5)

        # ===== 第二行右侧功能区域 =====
        right_frame = tk.Frame(main_frame1)
        right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)

        # --- 新增操作按钮组 ---
        btn_group2 = tk.Frame(right_frame, bd=1, relief=tk.RAISED)
        btn_group2.pack(fill=tk.X, pady=5)

        # 操作类型选择
        self.output_operation2 = ttk.Combobox(
            btn_group2,
            values=['Format', 'Copy', 'Paste', 'Clear'],
            state='readonly'
        )
        self.output_operation2.pack(side=tk.RIGHT, padx=5, pady=2)
        self.output_operation2.current(1)

        # 执行按钮
        ttk.Button(
            btn_group2, text="Execute",
            command=lambda: self.combinateCommonOperation(sub_win,
                                                          self.right_text_output,
                                                          self.output_operation2)
        ).pack(side=tk.RIGHT, padx=5)

        # ===== 底部按钮区域 =====
        btn_frame = tk.Frame(sub_win)
        btn_frame.pack(fill=tk.X, padx=5, pady=5)

        tk.Label(btn_frame, text="识别语言:").pack(side=tk.LEFT, padx=(0, 5))
        self.lang_combo = ttk.Combobox(btn_frame, values=["中文", "英文", "日文", "韩文", "阿拉伯文"])
        self.lang_combo.pack(side=tk.LEFT, padx=(0, 10))
        self.lang_combo.current(0)

        # ask_btn = ttk.Button(btn_frame, text="提问",command=lambda: self.askAI(self.left_text_output,self.right_text_output,"gpt-4o-mini"))
        # ask_btn.pack(side=tk.LEFT)

        # 修改原按钮命令：调用load_image_to_label而非直接操作Label
        select_btn = ttk.Button(btn_frame, text="选择图片提问", command=lambda: self.select_image_for_ocr_qa())
        select_btn.pack(side=tk.RIGHT)

    def nested_to_string(self, nested_list):
        """
        嵌套列表转字符串
        nested_list = [[1, [2, 3]], ['a', ['b', 'c']]]
        result = nested_to_string(nested_list)  # 输出: "1 2 3 a b c"
        """
        if isinstance(nested_list, list):
            return ' '.join(self.nested_to_string(item) for item in nested_list)
        else:
            return str(nested_list)

    def seeImgToTxtByPaddleOcr(self, which_text, img_path, lang, pattern):
        # 创建PaddleOCR对象，指定语言模型，默认为中文英文模型
        ocr = PaddleOCR(use_angle_cls=True, lang=lang)  # 'ch'表示中文，'en'表示中文
        # 使用OCR进行文字识别
        result = ocr.ocr(img_path, cls=True)
        print(f"识别结果是：{result}")
        print(f"识别结果是：{result[0]}")
        print(type(result))
        d = {}
        res = []
        # 结果展示
        for line in result[0]:
            coords = line[0]  # 文本坐标
            # 获取四个角坐标
            x1, y1 = coords[0]
            x2, y2 = coords[1]
            x3, y3 = coords[2]
            x4, y4 = coords[3]

            # 计算中心点坐标
            center_x = (x1 + x2 + x3 + x4) / 4
            center_y = (y1 + y2 + y3 + y4) / 4
            textResult = " "
            # print(f"检测到文本：{line[1][0]}，置信度：{line[1][1]}，该行文本的中心坐标是: ({center_x}, {center_y})")
            # print(f"{modifyTxt(line[1][0])}")
            if pattern == '文本':
                textResult = f"{line[1][0]}"
            elif pattern == '定位':
                textResult = f"检测到文本：{line[1][0]}，置信度：{line[1][1]}，该行文本的中心坐标是: ({center_x}, {center_y})"
            # 修正方案1：确保result为字符串类型
            if not isinstance(textResult, str):
                textResult = str(result)
            res.append(textResult)
            # print(f"{line[1][0]}")
            # d[line[1][0]] = (center_x, center_y)
        # return d
        self.clearContent(which_text)
        which_text.insert('insert', '\n'.join(res))

    def seeImgToTxtByPaddleOcr1(self, which_text, img_path, lang):
        # 创建PaddleOCR对象，指定语言模型，默认为中文英文模型
        ocr = PaddleOCR(use_angle_cls=True, lang=lang)  # 'ch'表示中文，'en'表示中文
        # 使用OCR进行文字识别
        result = ocr.ocr(img_path, cls=True)
        print(f"识别结果是：{result}")
        print(f"识别结果是：{result[0]}")
        print(type(result))
        d = {}
        res = []
        # 结果展示
        for line in result[0]:
            # coords = line[0]  # 文本坐标
            # # 获取四个角坐标
            # x1, y1 = coords[0]
            # x2, y2 = coords[1]
            # x3, y3 = coords[2]
            # x4, y4 = coords[3]
            # # 计算中心点坐标
            # center_x = (x1 + x2 + x3 + x4) / 4
            # center_y = (y1 + y2 + y3 + y4) / 4
            textResult = f"{line[1][0]}"
            # 修正方案1：确保result为字符串类型
            if not isinstance(textResult, str):
                textResult = str(result)
            res.append(textResult)
        self.clearContent(which_text)
        which_text.insert('insert', '\n'.join(res))

    # 删除日志
    def delete_records(self):
        # 设置目录路径
        directory = f'{current_script_path}/requests_log'
        # 获取所有以"20"开头的文件路径
        for filename in glob.glob(os.path.join(directory, '20*')):
            os.remove(filename)  # 移除"20"开头的文件
        self.load_records()

    # 一定程度上缩进复制的网页python代码字符串
    def restore_python_code(self, which_text, code_str):  # 定义一个栈来跟踪代码块的缩进
        indent_stack = [0]
        restored_lines = []
        current_line = ''
        for line in code_str.splitlines():
            # 移除行首和行尾的空白字符
            stripped_line = line.strip()
            if stripped_line:
                # 计算当前行的缩进级别
                current_indent = len(line) - len(stripped_line)
                # 调整缩进级别
                while current_indent < indent_stack[-1]:
                    indent_stack.pop()
                if current_indent > indent_stack[-1]:
                    indent_stack.append(current_indent)
                # 添加缩进空格
                indent_spaces = ' ' * indent_stack[-1]
                # 将处理后的行添加到当前行
                current_line += '' + stripped_line
            else:
                # 如果行是空的，输出当前行并重置
                restored_lines.append(current_line)
                current_line = ''
                indent_stack = [0]
        # 添加最后一行
        if current_line:
            restored_lines.append(current_line)

        # 将恢复后的代码行合并为单一字符串
        print(restored_lines)
        restored_code = '\n'.join(restored_lines).strip()

        if 'if __name__' in restored_code:
            restored_code = restored_code.replace('if __name__', '\nif __name__')
        if 'def' in restored_code:
            restored_code = restored_code.replace('def', '\ndef')
        if 'class' in restored_code:
            restored_code = restored_code.replace('class', '\nclass')
        if 'from' in restored_code:
            restored_code = restored_code.replace('from', '\nfrom')
        if 'import' in restored_code:
            restored_code = restored_code.replace('import', '\nimport')
        if ',\n' in restored_code or ', \n' in restored_code:
            restored_code = restored_code.replace(',\n', ',')
        if '\n ' in restored_code:
            restored_code = restored_code.replace('\n ', '\n  ')
        if '(\n' in restored_code:
            restored_code = restored_code.replace('(\n', '(')
        if '=\n' in restored_code:
            restored_code = restored_code.replace('=\n', '=')
        self.clearContent(which_text)
        which_text.insert('insert', restored_code)

    def restore_python_code_by_AI(self, which_text, text_code_str):  # 用chatgpt4mini解决
        question = f"{text_code_str}这段代码格式和语法正确吗？分析包括程序的语法分析和格式分析两个步骤，回答要求如下：" \
                   f"第一步：语法不正确的话简单输出错误解释，不需要进行第二步的处理；第二步：格式不正确的话请将其修改为正确的格式输出来。要求只输出代码"
        try:
            completion = openai.chat.completions.create(
                model='gpt-4o-mini',
                messages=[{"role": "user", "content": question}],
                timeout=20  # 增加超时限制
            )
            answer = completion.choices[0].message.content
            formatted = answer
            sleep(1.5)  # 避免频繁请求
            self.clearContent(which_text)
            which_text.insert('insert', formatted.replace('```python', '').replace('```', ''))
        except Exception as e:
            print(f"问题 {question} 处理失败: {str(e)}")
            error_msg = f"ERROR: {str(e)}"
            formatted = error_msg
            self.clearContent(which_text)
            which_text.insert('insert', formatted)
        # return formatted

    def refreshKey(self, which_text, which_combobox):
        try:
            jsonStr = which_text.get("1.0", "end-1c")
            jsonDict = json.loads(jsonStr)
            print(type(jsonDict), jsonDict)
            keyList = find_all_key(jsonDict)
            print(type(keyList), keyList)
            which_combobox['values'] = keyList
        except:
            which_combobox['values'] = []

    def changeKeyValueInJsonstr(self, which_text, key_to_modify, new_value):
        try:
            jsonStr = which_text.get("1.0", "end-1c")
            jsonDict = json.loads(jsonStr)
            # 更新键对应的值
            find_key_and_update_value(jsonDict, key_to_modify, new_value)
            which_text.delete('1.0', tk.END)
            # 展示修改后的值
            display = json.dumps(jsonDict, ensure_ascii=False, indent=4)
            which_text.delete('1.0', tk.END)
            which_text.insert(tk.END, display)
            print(display)
        except Exception as e:
            which_text.delete('1.0', tk.END)
            which_text.insert(tk.END, f'更新异常: {e}')
            pass

    def changeKeyValueInJsonFile(self, key_to_modify, new_value):
        try:
            # JSON文件名为configin.json
            json_file_path = f'{current_script_path}/configini.json'

            # 读取JSON文件
            data = read_json_file(json_file_path)

            print(type(data))
            # print(data)

            if not isinstance(data, dict):
                data = json.loads(data)

            # print(data)
            # print(type(data))

            # 修改键值，假设要修改"key_to_modify"的值，注意还没写进去文件里
            if key_to_modify in data['lobalVariable']:
                data['GlobalVariable'][key_to_modify] = new_value
            # 写回JSON文件
            with open(json_file_path, 'w', encoding=encodingType) as file:
                json.dump(data, file, ensure_ascii=False, indent=4)
            # 每调用一次函数就更新入参值
            bodyDict = json.loads(self.body_text.get("1.0", "end-1c"))
            # 修改后获取配置的全局变量值
            setKeyList = read_json_file(f'{current_script_path}/configini.json')['GlobalVariable']
            # print('setKeyList',setKeyList)#<class 'dict'>

            parameterKeyList = ["reporterCaliNo", "reporterRegisterTel", "taskHandleUM",
                                "taskHandleUlMohile""surveyorMobile", "surveyorName", "surveyorUm"]
            # 获取配置文件里的键值对
            for key in parameterKeyList:
                if 'reporterCallNo' == key or 'reporterRegisterTel' == key:
                    find_key_and_update_value(bodyDict, key, setKeyList['mobileNo'])
                if 'taskHandleUMMobile' == key or 'surveyorMobile' == key:
                    find_key_and_update_value(bodyDict, key, setKeyList['surveyumphone'])
                if 'supveyorName' == key or 'surveyorUm' == key or 'taskHandleUM' == key:
                    find_key_and_update_value(bodyDict, key, setKeyList['surveyUm'])
                    # bodyDict["survey"][key] = setKeyList['surveyUm']
            self.body_text.delete("1,0", tk.END)
            # 展示修改后的值
            display = json.dumps(bodyDict, ensure_ascii=False, indent=4)
            self.body_text.insert(tk.END, display)
            print(display)
        except Exception as e:
            self.body_text.insert(tk.END, f'更新异常: {e}')
            pass

    # 设置线程，避免窗口回弹崩溃
    def thread_it(self, func, *args):
        """将函数打包进线程"""
        self.myThread = threading.Thread(target=func, args=args)
        self.myThread.setDaemon(True)  # 主线程退出就直接让子线程跟随退出 ，不论是否运行完成
        self.myThread.start()

    def send_request_by_userInputData(self):  # 获取用户输入
        url = self.url_entry.get()
        method = self.method_combobox.get()
        headers_input = self.headers_text.get("1.0", tk.END)
        body_input = self.body_text.get("1.0", tk.END)
        if method == 'GET' and '&' in body_input:
            td = {}
            for i in body_input.split('&'):
                j = i.split('=')
                td[j[0]] = j[1]
            body_input = td
            print(td)

        # 尝试将输入转换为字典
        headers = self.try_parse_json(headers_input)  # <class 'dict'>
        body = self.try_parse_json(body_input)  # <class 'dict!>

        # 前置依赖鉴权接口
        if 'Cookie' in headers:
            loginResponseInfo = self.send_request_by_recordFile(self.before_login_combobox.get())
            # print(type(loginResponseInfo), loginResponseInfo)
            if loginResponseInfo != '':
                newestAuthCookie = find_value_of_key_in_nested_dict(loginResponseInfo, "Cookie")
                # print('newestAuthCookie',newestAuthCookie)
                headers['Cookie'] = newestAuthCookie
                # print(headers)
                # 更新Cookie到headers输入框里
                self.headers_text.delete('1.0', tk.END)
                self.headers_text.insert(tk.END, self.format_json(json.dumps(headers)))
        # 发送请求
        try:
            # messagebox.showinfo('Sending...')
            # self.response_text.delete('1.0'.tk.END)
            # self.response_text.insert(tk.END，'Sending...")
            # messageInformInWin('Sending...',400)
            self.messageInformInWin('Sending...', 700)
            if method == 'GET':
                response = requests.request(method, url=url, params=body, headers=headers)
                # response = requests.get(unl=unl,params=body,headers=headers)
                # 展示响应
                self.response_text.delete("1.0", tk.END)
                self.response_text.insert(tk.END, self.format_json(response.text))
                save_request_data(url, headers, method, body, self.try_parse_json(response.text))
                self.load_records()
            elif method == 'POST':
                response = requests.post(url=url, data=json.dumps(body), headers=headers)
                # print(response.headens.get( set-Cookie!))
                Cookie = response.headers.get('Set-Cookie')
                responseDict = json.loads(response.text)
                if not isinstance(responseDict, dict):
                    responseDict = json.loads(responseDict)
                    print(type(responseDict))
                if 'data' in responseDict:
                    if isinstance(responseDict['data'], list):  # 判断数据类型是否为列表
                        responseDict['data'].append({'Cookie': Cookie})  # <class 'list'>
                    elif isinstance(responseDict['data'], dict):  # 判断数据类型是否为字典
                        responseDict['data']['Cookie'] = Cookie  # <class 'dict'>'
                    elif responseDict['data'] is None:  # 判断数据类型是否为字典
                        responseDict['data'] = dict()
                        responseDict['data']['data'] = None
                        responseDict['data']['Cookie'] = Cookie  # <class 'dict'>
                else:
                    responseDict['Cookie'] = Cookie  # <class 'dict'>pass
                responseJson = json.dumps(responseDict)
                # 展示响应
                self.response_text.delete('1.0', tk.END)
                self.response_text.insert(tk.END, self.format_json(responseJson))
                save_request_data(url, headers, method, body, self.try_parse_json(responseJson))
                self.load_records()
            else:
                response = requests.request(method, url=url, params=json.dumps(body), headers=headers)  # 展示响应
                self.response_text.delete('1.0', tk.END)
                self.response_text.insert(tk.END, self.format_json(response.text))
                save_request_data(url, headers, method, body, self.try_parse_json(response.text))
                self.load_records()
            # messagebox.showinfo(Successful!!)
        except Exception as e:
            eDict = {}
            eDict['e'] = f'Error: {e}'
            eJson = json.dumps(eDict)
            self.response_text.delete('1.0', tk.END)
            self.response_text.insert(tk.END, self.format_json(eJson))
            save_request_data(url, headers, method, body, self.try_parse_json(eJson))  # 异常结果也保存self.load_records()

    def send_request_by_recordFile(self, jsonRecordFilePath):
        if jsonRecordFilePath == '':
            return ''
        jsonRecordDict = read_json_file(jsonRecordFilePath)  # dict
        print('jsonRecordDict', jsonRecordDict)
        # 获取用户输入
        url = jsonRecordDict['ur']
        method = jsonRecordDict['method']
        headers_request = jsonRecordDict['headers']
        body_request = jsonRecordDict['request']
        # 尝试将输入转换为字典
        headers = headers_request  # <class 'str'>
        body = json.dumps(body_request)  # <class 'str'>
        # 发送请求
        try:
            self.messageInformInWin('Sending . .', 400)
            if method == 'GET':
                response = requests.request(method, url=url, params=body,
                                            headers=headers)  # response = reguests.get(unl=unl,params=body, headers=headers)
                return json.loads(response.text)
            elif method == 'POST':
                # response = neguests.request(method, unlzunl , .params=json .dumps(body), headers=headens)
                response = requests.post(url=url, data=body, headers=headers)
                print(response)
                Cookie = response.headers.get('Set-Cookie')
                responseDict = json.loads(response.text)  # <class idict'>
                if isinstance(responseDict['data'], list):  # 判断数据类型是否为列表
                    responseDict['data'].append({'Cookie': Cookie})  # <class"list'>
                if isinstance(responseDict['data'], dict):  # 判断数据类型是否为字典
                    responseDict['data']['Cookie'] = Cookie  # <class 'dict'>
                return responseDict
            else:
                response = requests.request(method, url=url, params=json.dumps(body), headers=headers)
                return json.loads(response.text)
        except Exception as e:
            self.response_text.delete('1.0', tk.END)
            self.response_text.insert(tk.END, f"Error: {e}")
            # messagebox.showinfo( Unsuccessful!

    # 尝试解析JSON字符串
    def try_parse_json(self, json_input):
        try:
            return json.loads(json_input)  #
        except Exception as e:
            return {'error': f'Error: {e}'}

    def combinateCommonOperation(self, which_win, which_text, which_combobox):
        operationName = which_combobox.get()
        print(operationName)
        if operationName == 'Format':
            self.thread_it(self.format_content(which_text))
        elif operationName == 'Copy':
            self.thread_it(self.copy_content(which_text))
        elif operationName == 'Paste':
            self.thread_it(self.paste(which_win, which_text))
        elif operationName == 'Clear':
            self.thread_it(self.clearContent(which_text))
        pass

    def format_content(self, which_text):
        formatted = self.format_json(which_text.get("1.0", tk.END))
        which_text.delete('1.0', tk.END)
        which_text.insert(tk.END, formatted)

    # 格式化JSON字符串
    def format_json(self, json_input) -> str:
        try:
            parsed = json.loads(json_input)
            formatted = json.dumps(parsed, ensure_ascii=False, indent=4)
            return formatted  # jsonStr
        except json.JSONDecodeError:
            return ""  # 返回一个空的JSON对象，如果输入无法解析

    def findLoginRecordFilePath(self):
        filetypes = [("JSON Files", "*,json")]
        path = filedialog.askopenfilename(title="选择登录接口文件", filetypes=filetypes)
        # 打开一个文件选择对话框，用户选择json文件
        if not isfile(path):  # 检查所选文件是否存在
            print(f"文件不存在:{path}")
        else:
            pass
        try:
            self.before_login_combobox.insert(tk.END, path)
        except:
            defaultPath = f'{current_script_path}/requests_log/post数字员工请求登录demo.json'
            self.before_login_combobox.insert(tk.END, defaultPath)

    def find_file_to_fill_record(self):
        """通过选择的记录填充表单"""
        record_file = self.record_combobox.get()
        # path = f'{curnent_script-pathl/(logFileNamel/irecord_filel
        filetypes = [("JSON Files", "*.json")]
        path = filedialog.askopenfilename(title="选择json文件", filetypes=filetypes)
        # 打开一个文件选择对话框，用户选择json文件
        if not isfile(path):  # 检查所选文件是否存在
            print(f"文件不存在:{path}")
        else:
            pass
        try:
            with open(path, 'r') as file:
                data = json.load(file)
                request_data = data['request']
                response_data = data['response']

                self.url_entry.delete(0, tk.END)
                self.url_entry.insert(0, data['url'])
                self.method_combobox.set(data['method'])
                self.headers_text.delete(1.0, tk.END)
                self.headers_text.insert(
                    tk.END.json.dumps(data["headers"], ensure_ascii=False, indent=4).replace('\'', '\"'))
                self.body_text.delete(1.0, tk.END)
                self.body_text.insert(tk.END,
                                      json.dumps(request_data, ensure_ascii=False, indent=4).replace('\'', '\"'))
                self.response_text.delete(1, 0, tk.END)
                self.response_text.insert(tk.END,
                                          json.dumps(response_data, ensure_ascii=False, indent=4).replace('\'', '\"'))
        except:
            pass

    def copy_content(self, which_text):  # 获取输入框的内容
        input_content = which_text.get("1.0", "end-1c")  # 将内容复制到剪贴板
        self.clipboard_clear()
        self.clipboard_append(input_content)  # 显示复制成功的消息《可选》
        print("内容已复制到剪贴板。")  # messagebox.showinfo(input_content)
        self.messageInformInWin(input_content, 2800)
        time.sleep(2)

    def encryptResult(self, stringValue, secretChar):
        if not isinstance(stringValue, str):
            return "输入类型错误"
        if len(stringValue) < 2:
            return stringValue
        newstringValue = secretChar + stringValue[0: len(stringValue) // 2] + secretChar + stringValue[len(
            stringValue) // 2:] + secretChar
        return newstringValue

    def getTestData(self, which_text, encryptChar_text):
        encryptChar = encryptChar_text  # .get()#测试
        # 生成测试数据
        f = Faker(["zh_CN"])  # 默认en_Us，支持中文本地化zh_Ch
        fake_name = f.name()
        fake_id = self.encryptResult(generate_id_card(), encryptChar)
        fake_phone = self.encryptResult(f.phone_number(), encryptChar)
        fake_email = self.encryptResult(f.email(), encryptChar)
        plate_info = generate_ramdon_plate()
        plate_number = self.encryptResult(plate_info[0], encryptChar)
        plate_type = self.encryptResult(plate_info[1], encryptChar)
        fake_postcode = self.encryptResult(f.postcode(), encryptChar)
        vin = self.encryptResult(generate_vin(), encryptChar)
        currentTimeInfo = getCurrentTimeInfo()
        timestamp = self.encryptResult(str(currentTimeInfo[0]), encryptChar)
        formatted_time = self.encryptResult(currentTimeInfo[1], encryptChar)
        bankUser = '平安测试六零零零三四零一二四二零'
        bankName = '平安银行'
        acountCard = self.encryptResult("11006545830302", encryptChar)
        # 以JSON格式返回结果
        data = {
            "姓名": fake_name,
            "身份证号": fake_id,
            "手机号": fake_phone,
            "邮箱": fake_email,
            "车牌号": plate_number,
            "车牌类型": plate_type,
            "VIN': vin,"
            "邮编": fake_postcode,
            "开户名": bankUser,
            "银行": bankName,
            "银行卡": acountCard,
            "时间": timestamp,
            "格式化时间": formatted_time
        }
        # 打印JSON格式数据
        print(json.dumps(data, ensure_ascii=False, indent=4))
        which_text.delete(1.0, tk.END)
        which_text.insert(tk.END, json.dumps(data, ensure_ascii=False, indent=4).replace("\'", '\"'))

    def getHtmlData(self, which_text, html_content):
        tag = False
        output = ""
        res = []
        for char in html_content:
            if char == '<':
                if len(output) != 0:
                    res.append(output)
                    output = ""
                tag = True
            elif char == '>':
                tag = False
            elif not tag:
                output = output + char
        res = [r for r in res if r != '' or r != ' ' or r != '\n' or r != '\t']
        self.clearContent(which_text)
        which_text.insert('insert', '\n'.join(res))

    def clearContent(self, which_text):
        # 清空输入框的内容
        which_text.delete("1.0", "end-1c")

    def paste(self, which_win, which_text):
        # 清空输入框的内容
        which_text.delete("1.0", "end-1c")
        # 获取煎贴板的内容
        clipboard_content = which_win.clipboard_get()
        # 将内容插入到输入框
        which_text.insert('insert', clipboard_content)

    def copy_value_of_key(self, which_win, which_text, targetkey):  # 获取输入框的内容
        input_content = which_text.get("1.0", "end-1c")
        # 将内容复制到剪贴板
        which_win.clipboard_clear()
        try:
            res = self.try_parse_json(input_content)  # <class  dict'>print(type(res))
            value = find_value_of_key_in_nested_dict(res, targetkey)
            which_win.clipboard_append(value)
            # 显示复制成功的消息《可选)
            print("内容已复制到剪贴板。")
            # messagebox.showinfo(value)messageInformInWin(value,lastingTime: 500)
            self.messageInformInWin(value, 2800)
        except Exception as e:
            #which_win.clipboard_append('0000')  # 显示复制成功的消息《可选》
            print(f"内容复制到剪贴板失败，具体原因为:{e}")
            which_win.clipboard_append(f"内容复制到剪贴板失败，具体原因为:{e}")  # 显示复制成功的消息《可选》
            # messagebox.showinfo(f"内容复制到剪贴板失败，具体原因为:{e}")messaneinformTnWincintorncontimplePostmanApp, find file_to_fill record()>try>with open(path, 'r") as file中"肉客更制到前贴板牛哈，且休厦因头
            self.messageInformInWin(f"内容复制到剪贴板失败，具体原因为:{e}", 2800)

            # messagebox.showinfo(f"内容复制到剪贴板失败，具体原因为:{e}")
            # messageInformInWin(infomContent,f"内容复制到剪贴板失败，具体原因为: {e}",500)

    def load_records(self):
        records = list_request_records()
        self.record_combobox['values'] = records

    ## 发送请求的函数
    def send_request(self):
        # 获取用户输入
        url = self.url_entry.get()
        method = self.method_combobox.get()
        headers_input = self.headers_text.get("1.0", tk.END)
        body_input = self.body_text.get("1.0", tk.END)

        # 尝试将输入转换为字典
        headers = self.try_parse_json(headers_input)
        body = self.try_parse_json(body_input)

        # 发送请求
        try:
            if method == 'GET':
                response = requests.request(method, url=url, params=body, headers=headers)
                # 展示响应
                self.response_text.delete('1.0', tk.END)
                self.response_text.insert(tk.END, self.format_json(response.text))
                save_request_data(url, headers, method, body, self.try_parse_json(response.text))
                self.load_records()
            elif method == 'POST':
                response = requests.post(url=url, data=json.dumps(body), headers=headers)
                # 展示响应
                self.response_text.delete('1.0', tk.END)
                self.response_text.insert(tk.END, self.format_json(response.text))
                save_request_data(url, headers, method, body, self.try_parse_json(response.text))
                self.load_records()
            else:
                response = requests.request(method, url=url, params=body, headers=headers)
                # 展示响应
                self.response_text.delete('1.0', tk.END)
                self.response_text.insert(tk.END, self.format_json(response.text))
                save_request_data(url, headers, method, body, self.try_parse_json(response.text))
                self.load_records()

        except Exception as e:
            self.response_text.delete('1.0', tk.END)
            self.response_text.insert(tk.END, f"Error: {e}")

    # 尝试解析JSON字符串
    def try_parse_json(self, json_input):
        try:
            return json.loads(json_input)
        except json.JSONDecodeError:
            return {}

    # 格式化Headers输入
    def format_headers(self):
        formatted = self.format_json(self.headers_text.get("1.0", tk.END))
        self.headers_text.delete('1.0', tk.END)
        self.headers_text.insert(tk.END, formatted)

    # 格式化Body输入
    def format_body(self):
        formatted = self.format_json(self.body_text.get("1.0", tk.END))
        self.body_text.delete('1.0', tk.END)
        self.body_text.insert(tk.END, formatted)

    # 格式化Response输入
    def format_response(self):
        formatted = self.format_json(self.response_text.get("1.0", tk.END))
        self.response_text.delete('1.0', tk.END)
        self.response_text.insert(tk.END, formatted)

    # 格式化JSON字符串
    def format_json(self, json_input):
        try:
            parsed = json.loads(json_input)
            formatted = json.dumps(parsed, ensure_ascii=False, indent=4)
            return formatted
        except json.JSONDecodeError:
            return "{}"  # 返回一个空的JSON对象，如果输入无法解析

    def fill_record(self, event):
        """通过选择的记录填充表单"""
        record_file = self.record_combobox.get()
        path = f'{current_script_path}/{logFileName}/{record_file}'
        with open(path, 'r') as file:
            data = json.load(file)
            request_data = data['request']
            response_data = data['response']

            self.url_entry.delete(0, tk.END)
            self.url_entry.insert(0, data['url'])
            self.method_combobox.set(data['method'])
            self.headers_text.delete(1.0, tk.END)
            self.headers_text.insert(tk.END,
                                     json.dumps(data['headers'], ensure_ascii=False, indent=4).replace('\'', '\"'))
            self.body_text.delete(1.0, tk.END)
            self.body_text.insert(tk.END, json.dumps(request_data, ensure_ascii=False, indent=4).replace('\'', '\"'))
            self.response_text.delete(1.0, tk.END)
            self.response_text.insert(tk.END,
                                      json.dumps(response_data, ensure_ascii=False, indent=4).replace('\'', '\"'))

    def copy_headers_text(self):
        # 获取输入框内容
        input_content = self.headers_text.get('1.0', 'end-1c')
        self.clipboard_clear()
        self.clipboard_append(input_content)
        # 显示复制成功消息（可选）
        print('内容已复制到剪贴板。')

    def copy_body_text(self):
        # 获取输入框内容
        input_content = self.body_text.get('1.0', 'end-1c')
        self.clipboard_clear()
        self.clipboard_append(input_content)
        # 显示复制成功消息（可选）
        print('内容已复制到剪贴板。')

    def copy_response_text(self):
        # 获取输入框内容
        input_content = self.response_text.get('1.0', 'end-1c')
        self.clipboard_clear()
        self.clipboard_append(input_content)
        # 显示复制成功消息（可选）
        print('内容已复制到剪贴板。')

    def copy_response_key(self):
        key = 'data'
        # 获取输入框内容
        input_content = self.response_text.get('1.0', 'end-1c')
        # 将内容复制到剪贴板
        self.clipboard_clear()
        try:
            res = self.try_parse_json(input_content)  # <class 'dict'>
            print(type(res))
            self.clipboard_append(find_value_in_nested_dict(res, key))
            # 显示复制成功的消息（可选）
            print('内容已复制到剪贴板。')
        except Exception as e:
            self.clipboard_append('0000')
            # 显示复制成功的消息（可选）
            print(f'内容复制到薄贴板失败，具体原因为：{e}')

    def search_text(self, search_term):
        # 清除之前高亮
        self.response_text.tag_remove('match', '1.0', 'end')
        # 如录有裂索内容，则进行搜索
        if search_term:
            start_pos = "1.0"
            while True:
                # 搜索第一个匹配项
                pos = self.response_text.search(search_term, start_pos, stopindex='end')
                # 如果没有找到匹配项，退出循环
                if not pos:
                    break
                # 标记匹配项
                end_pos = f"{pos}+{len(search_term)}c"
                self.response_text.tag_add("match", pos, end_pos)
                # 设置新的搜索起始位置
                start_pos = end_pos


if __name__ == '__main__':
    app = SimplePostmanApp()
    app.mainloop()

"""

export PATH="/usr/local/bin/python3.7:$PATH"

Windows打包应用
pyinstaller --onefile --windowed D:\TestDev\PostmanByTk1.py

Mac打包应用
一、pyinstaller 主要用于 Python3
which -a python3.7--看python3.7安装在哪里
pip show pyinstaller--看pyinstaller安装在哪里
/usr/local/bin/python3.7 -m pip install pyinstaller--指定版本安装
/opt/homebrew/bin/python3.12 -m pip install pyinstaller--指定版本安装
pip3.7 show pyinstaller--检查是否安装成功 

#pyinstaller --onefile --windowed aboutPostman.py
<<<<<<< HEAD
python3.7 -m PyInstaller --onefile --windowed /Users/ketangchen/Documents/UiAutoOfApp20240703/UiAutoOfApp/utils/PostmanByTk1.py ##重点看python3.7这里，python3.7下安装了PyInstaller

#dist路径
/Users/ketangchen/Documents/UiAutoOfApp/utils/dist

brew install create-dmg # arch -arm64 brew install create-dmg
create-dmg \
  --volname "AboutPostman" \
  --window-size 800 600 \
  --app-drop-link 400 200 \
  /Users/ketangchen/Desktop/AboutPostman.dmg  \
  /Users/ketangchen/Documents/UiAutoOfApp/utils/dist/aboutPostman.app


未能打开磁盘映像

/usr/local/bin/python3.7 -m pip install Pillow   -i https://pypi.tuna.tsinghua.edu.cn/simple
pip3 install scholarly  -i https://pypi.tuna.tsinghua.edu.cn/simple
=======
python3.7 -m PyInstaller --onefile --windowed aboutPostman.py ##重点看python3.7这里，python3.7下安装了PyInstaller
/usr/local/bin/python3.7 pip -m install Faker  -i https://pypi.tuna.tsinghua.edu.cn/simple
>>>>>>> 3fb88284a342c0e5e0553f94f36da1024ae76b58

二、py2app
python3 -m pip uninstall setuptools==69.5.1

pip install --upgrade setuptools
/usr/local/bin/python3.7 -m pip install setuptools
pip install --upgrade setuptools
/usr/local/bin/python3.7 -m pip install setuptools
创建你的setup.py脚本。若无，则可以使用py2applet创建一个。在终端中运行
py2applet --make-setup aboutPostman.py
运行setup.py脚本来创建应用
python3.7 setup.py py2app

python3.7 setup.py py2app --debug

chmod +x /path/to/your/app/YourApp.appYourApp
chmod x /Users/ketangchen/Documents/UiAutoOfApp/utils/dist/aboutPostman.app/Contents/MacOS/aboutPostman

#dist路径
/Users/ketangchen/Documents/UiAutoOfApp/utils/dist

/Users/ketangchen/Documents/UiAutoOfApp20240703/UiAutoOfApp/utils

打包依赖库
pip freeze > requirements.txt
安装依赖库
pip install -r /Users/ketangchen/Documents/UiAutoOfApp20240703/UiAutoOfApp/utils/requirements.txt



xattr -r -d com.apple.quarantine  /Users/ketangchen/Desktop/AboutPostman.dmg
xattr -r -d com.apple.quarantine  /Users/ketangchen/Documents/UiAutoOfApp/utils/dist/aboutPostman.app
"""


