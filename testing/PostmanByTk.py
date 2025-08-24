# -*- coding: utf-8 -*-
## arch -x86_64 pip install --force-reinstall opencv-python numpy
import os
import tkinter as tk  # /usr/local/bin/python3.7 pip -m install treelib  -i https://pypi.tuna.tsinghua.edu.cn/simple
from tkinter import ttk, scrolledtext  # D:\Python\Python37\python.exe pip install pillow  -i https://pypi.tuna.tsinghua.edu.cn/simple

import objc
import requests
import datetime
from faker import Faker
import json
import string
import random
from tkinter import messagebox
import glob
from PIL import Image  # /usr/local/bin/python3.7 -m pip install Pillow -i https://pypi.tuna.tsinghua.edu.cn/simple
from os.path import isfile, splitext
from tkinter import filedialog
import threading
from xmindparser import xmind_to_dict
from treelib import Tree  # /usr/local/bin/python3.7 -m pip install treelib
from uuid import uuid4
from openpyxl import Workbook
from time import sleep
# os.environ['PADDLE_DIR']  = '/Users/ketangchen/miniconda3/envs/my_env1/lib/python3.9/site-packages/paddle/libs'  # 在代码中手动设置 PaddlePaddle 的库路径（需在导入 paddleocr 前执行）
# os.environ['DYLD_LIBRARY_PATH']  = '/Users/ketangchen/miniconda3/envs/my_env1/lib/python3.9/site-packages/paddle/libs'  #
# 指定 PaddlePaddle 的库路径（需替换为实际路径）
# os.environ['PADDLE_DIR'] = '/Users/ketangchen/miniconda3/envs/my_env1/lib/python3.9/site-packages/paddle/libs'  # 例如 Conda 环境中的库路径
from paddleocr import PaddleOCR
"""
确保安装的是这几个版本，不要随便升级，否则爆；最好用虚拟环境
(my_env1) ketangdeMacBook-Pro:UiAutoOfApp ketangchen$ pip list | grep paddle
paddleocr                2.8.0
paddlepaddle             2.5.2
paddlex                  3.1.3
"""
import openai  # /usr/local/bin/python3.7 -m pip install openai -i https://pypi.tuna.tsinghua.edu.cn/simple
from tkinter import Canvas, Toplevel
from PIL import ImageGrab, ImageTk
import time
from openpyxl.styles import Border, Side, PatternFill
import pandas as pd
from tkinter.font import Font
import subprocess
from typing import List, Optional
#import keyboard

"""
/usr/local/bin/python3.7 -m pip install yourmodel -i https://pypi.tuna.tsinghua.edu.cn/simple
/usr/local/bin/python3.7 -m pip install py2app -i https://pypi.tuna.tsinghua.edu.cn/simple
/usr/local/bin/python3.7 -m pip install --upgrade paddlepaddle -i https://pypi.tuna.tsinghua.edu.cn/simple
/usr/local/bin/python3.7 -m pip install --upgrade py2app setuptools wheel -i https://pypi.tuna.tsinghua.edu.cn/simple

"""

# 确保存储日志的目录存在
current_script_path = os.getcwd()
logFileName = 'requests_log'
logFilePath = f'{current_script_path}/{logFileName}'
if not os.path.exists(logFilePath):
    os.makedirs(logFilePath)

QALogFileName = 'qa_log'
QALogFilePath = f'{current_script_path}/{QALogFileName}'
if not os.path.exists(QALogFilePath):
    os.makedirs(QALogFilePath)

screenshotFileName = 'screenshots_log'
screenshotFilePath = f'{current_script_path}/{screenshotFileName}'
if not os.path.exists(screenshotFilePath):
    os.makedirs(screenshotFilePath)

fileName = 'file'
filePath = f'{current_script_path}/{fileName}'
if not os.path.exists(filePath):
    os.makedirs(filePath)

caseExcelFileName = 'caseExcel'
caseExcelfilePath = f'{current_script_path}/{caseExcelFileName}'
if not os.path.exists(caseExcelfilePath):
    os.makedirs(caseExcelfilePath)

encodingType = 'utf-8'

# 读取JSON文件
def read_json_file1(file_path) -> dict:
    temp = []
    res = ''
    with open(file_path, 'r', encoding=encodingType) as f:
        lines = f.readlines()
        for line in lines:
            temp.append(line.strip())
    print(temp)
    try:
        res = json.loads(''.join(temp))
        print(res)
    except Exception as e:
        res = {"error": f"{e}"}
    return res


def read_json_file(file_path) -> dict:
    with open(file_path, 'r', encoding=encodingType) as file:  # 记得把对应json文件编码转化为gbk编码
        data = json.load(file)
        print(data)
        return data

def find_value_of_key_in_nested_dict(d, target_key):
    """
    非递归版本的搜索炭套宇典中的键对应值。
    :param d: 要搜索的宁典或宁典列表。
    :param target_key:要搜索的日标键
    :return: 包含找到指定键对应值。
    """

    if isinstance(d, str):
        d = json.loads(d)  # <class 'dict'>
    if isinstance(d, dict):
        stack = [(d, ())]
    elif isinstance(d, list):
        stack = [(item, ()) for item in d]
    else:
        return []

    result = []

    while stack:
        current, path = stack.pop()

        if isinstance(current, dict):
            for key, value in current.items():
                current_path = path + (key,)
                if key == target_key:
                    result.append((current_path, value))
                if isinstance(value, (dict, list)):
                    stack.append((value, current_path))
        elif isinstance(current, list):
            for idx, item in enumerate(current):
                current_path = path + (idx,)
                if isinstance(item, (dict, list)):
                    stack.append((item, current_path))

    print(result)
    for r in result:
        if len(r) != 0 and r[1] != None:
            return r[1]
    return 'None'


def write_xmind_data_to_excel(casePathDataList, caseType):
    """
    根据不同公司标准选择不同的用例格式来将用例的路径写进Excel
    """
    print(casePathDataList)
    if caseType == '版本用例' or caseType == '回归用例':
        ###中国平安用例格式
        versionCaseLevelDict = {'0': 'P0 冒烟测试', '1': 'P1 正向流程', '2': 'P2 异常反向'}
        regressionCaseLevelDict = {'0': 'P0 重要且频繁（致命，回归必跑）', '1': 'P1 重要不频繁（严重，评审选跑）', '2': 'P2 非重要功能（一般，评审选跑）'}
        defaultList = ['M 功能测试', 'P1 正向流程', 'A 需求分析']
        if caseType == '回归用例':
            defaultList = ['P2 非重要功能（一般，评审选跑）', 'APP类', ' ', '否', '不可以', 'A 需求分析']
        if len(casePathDataList) != 0:
            for d in casePathDataList:
                temp = d
                if len(d) == 3:
                    d.extend(defaultList)
                elif len(d) == 4:
                    level = d[3]
                    d.pop(-1)
                    if caseType != '回归用例':
                        if level in versionCaseLevelDict:
                            d.extend(['M 功能测试', versionCaseLevelDict[level], 'A 需求分析'])
                        else:
                            d.extend(['M 功能测试', versionCaseLevelDict['1'], 'A 需求分析'])
                    else:
                        if level in regressionCaseLevelDict:
                            d.extend([regressionCaseLevelDict[level], 'APP类', ' ', '否', '不可以', 'A 需求分析'])
                        else:
                            d.extend([regressionCaseLevelDict['2'], 'APP类', ' ', '否', '不可以', 'A 需求分析'])

                print(casePathDataList)
        # 创建一个新的工作薄
        wb = Workbook()
        # 选择默认的工作表
        ws = wb.active
        # 要写入的数据"M 功能测试"PO 冒烟测试'A 需求分析
        rows = [["用例名称", "用例步骤", "预期结果", "用例类型", "用例级别", "用例来源"]]
        if caseType == '回归用例':
            rows = [["用例名称", "用例步骤", "预期结果", "用例级别", "用例类型", "用例描述", "是否可自动化", "不可自动化原因", "用例来源"]]
        rows.extend(casePathDataList)
        print(rows)
        # 逐行写入数据
        for row in rows:
            ws.append(row)
        # 保存工作簿
        # wb.save('result.xlsx')
        wb.save(rf'{current_script_path}\\{caseExcelFileName}\\{getCurrentTimeInfo1()}-result.xlsx')
    elif caseType == '普通用例':
        ###广交所用例格式
        print(casePathDataList)
        # 创建一个新的工作薄
        wb = Workbook()
        # 选择默认的工作表
        ws = wb.active
        rows = [
                   '测试用例ID,系统名称,功能模块,一级功能,二级功能,三级功能,设计名称,前提条件,测试数据,验证点,验证内容,测试步骤,期望结果,实际结果,是否通过,测试版本,测试人员,复核人员,测试日期,优先级'.split(
                       ',')] * 2
        rows.extend(casePathDataList)
        # 逐行写入数据
        for row in rows:
            ws.append(row)

        # 定义黄色填充
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        # 定义边框样式
        border_style = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000")
        )

        # 遍历数据行，设置边框并根据条件染色
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            # 检查是否需要染色（假设第15列是“是否通过”列）
            if "优先级：" in row[19].value:  # row[4] 是第5列的单元格
                for cell in row:
                    cell.fill = yellow_fill  # 染色为黄色
            # 为每个单元格设置边框
            for cell in row:
                cell.border = border_style

        # 删除最后一列
        # wb.max_column 返回的是最后一列的索引，例如最后一列是第10列，则返回10
        ws.delete_cols(ws.max_column)

        # print(getCurrentTimeInfo1())
        # print(f'{current_script_path}\\{caseExcelFileName}\\{getCurrentTimeInfo1()}-result.xlsx')

        # 保存工作簿
        wb.save(f'{current_script_path}\\{caseExcelFileName}\\{getCurrentTimeInfo1()}-result.xlsx')


def xmind_to_tree(xmind_file_path):
    # 将Xmind文件解析为字典
    xmind_data = xmind_to_dict(xmind_file_path)
    # 创建Tree对象
    tree = Tree()

    # 递归函数，用于构建多叉树
    def build_tree_from_topic(topic, parent):
        # 为每个主题创建一个树节点
        node_id = str(uuid4())  # 生成唯一的节点ID
        tree.create_node(topic['title'], node_id, parent=parent)

        # 遍历子主题并递归调用
        for sub_topic in topic.get('topics', []):
            build_tree_from_topic(sub_topic, node_id)

    # 从Xmind数据中找到根主题，并开始构建树
    for sheet in xmind_data:
        root_topic = sheet['topic']
        build_tree_from_topic(root_topic, None)

    return tree


def get_all_leaf_paths_data_list(tree, caseType):
    """
    根据不同公司标准选择不同的用例格式来获取用例的路径
    """
    if caseType == '版本用例' or caseType == '回归用例':
        ###中国平安用例格式
        res = []
        d = {}  # 存tag和identifier到字典做映射
        all_nodes_info = tree.all_nodes()
        for a in all_nodes_info:
            d[a.identifier] = a.tag

        all_init_paths = tree.paths_to_leaves()  # 加工初始化路径，返回标题形式
        for a in all_init_paths:

            temp = [d[i] for i in a]
            print(f'casePath:{temp}')
            if len(temp) < 2:
                continue
            case_1 = temp[-1]  # 叶子结点
            case_2 = temp[-2]  # 叶子结点的父节点

            if '名称:' not in case_1 and '名称:' not in case_2:
                continue
            if '名称:' in case_1 and '名称:' in case_2:
                continue

            if '名称:' in case_1 and '名称:' not in case_2:
                try:
                    caseName = case_1.split('名称:')[1].split('步骤:')[0].replace('\r', '').replace('\n', '')
                    caseStep = case_1.split('步骤:')[1].split('预期:')[0].replace('\r', '').replace('\n', '')
                    caseExpect = case_1.split('预期:')[1].replace('\r', '').replace('\n', '')
                    temp.pop(-1)
                    casePath = '/'.join(temp) + '/' + caseName
                    res.append([casePath, caseStep, caseExpect])
                except Exception as e:
                    print(f'错误案例为：{case_1}')
                    print(f'转化异常：{e}')

            if '名称:' not in case_1 and '名称:' in case_2:
                try:
                    caseName = case_2.split('名称:')[1].split('步骤:')[0].replace('\r', '').replace('\n', '')
                    caseStep = case_2.split('步骤:')[1].split('预期:')[0].replace('\r', '').replace('\n', '')
                    caseExpect = case_2.split('预期:')[1].replace('\r', '').replace('\n', '')
                    temp.pop(-1)
                    casePath = '/'.join(temp) + '/' + caseName
                    res.append([casePath, caseStep, caseExpect, case_1])
                except Exception as e:
                    print(f'错误案例为：{case_2}')
                    print(f'转化异常：{e}')
        return res
    elif caseType == '普通用例':
        ###广交所用例格式
        rows = '测试用例ID,系统名称,功能模块,一级功能,二级功能,三级功能,设计名称,前提条件,测试数据,验证点,验证内容,测试步骤,期望结果,实际结果,是否通过,测试版本,测试人员,复核人员,测试日期,优先级'.split(
            ',')
        d = {'测试用例ID': 0, '系统名称': 1, '功能模块': 2, '一级功能': 3, '二级功能': 4, '三级功能': 5, '设计名称': 6, '前提条件': 7, '测试数据': 8,
             '验证点': 9, '验证内容': 10, '测试步骤': 11, '期望结果': 12, '实际结果': 13, '是否通过': 14, '测试版本': 15, '测试人员': 16, '复核人员': 17,
             '测试日期': 18, '优先级': 19}
        print(d)
        res = []
        all_nodes_info = tree.all_nodes()
        for a in all_nodes_info:
            d[a.identifier] = a.tag

        all_init_paths = tree.paths_to_leaves()  # 加工初始化路径，返回标题形式
        count = 0
        characterCountLen = len(str(len(all_init_paths)))
        for a in all_init_paths:
            restemp = [' '] * len(rows)
            restemp[d['测试用例ID']] = '用例' + '0' * (characterCountLen - len(f'{count + 1}')) + f'{count + 1}'
            restemp[d['功能模块']] = '功能模块'
            count += 1
            print(restemp)
            temp = [d[i] for i in a]
            print(temp)
            for k, t in enumerate(temp):
                print(t)
                if k == 1 and ('设计名称' not in t):
                    restemp[d['系统名称']] = t
                elif k == 2 and ('设计名称' not in t) and ('：' not in t):
                    restemp[d['一级功能']] = t
                elif k == 3 and ('设计名称' not in t) and ('：' not in t):
                    restemp[d['二级功能']] = t
                elif k == 4 and ('设计名称' not in t) and ('：' not in t):
                    restemp[d['三级功能']] = t
                elif '设计名称' in t:
                    restemp[d['设计名称']] = t
                elif '验证点' in t:
                    restemp[d['验证点']] = t
                elif '前提条件' in t:
                    restemp[d['前提条件']] = t
                elif '测试数据' in t:
                    restemp[d['测试数据']] = t
                elif '验证内容' in t:
                    restemp[d['验证内容']] = t
                    restemp[d['测试步骤']] = t.replace('验证内容', '测试步骤')
                    restemp[d['期望结果']] = t.replace('验证内容', '期望结果')
                elif '测试步骤' in t:
                    restemp[d['测试步骤']] = t
                elif '期望结果' in t:
                    restemp[d['期望结果']] = t
                elif '优先级' in t:
                    restemp[d['优先级']] = t
            res.append(restemp)
            print(restemp)
        return res

def getCurrentTimeInfo():
    currentTimeStamp = time.time()
    timestamp = currentTimeStamp
    print(f'timestamp:{timestamp}')
    formatted_time = time.strftime('%Y-%m-%d-%H-%M-%S', time.localtime(timestamp))
    # print(currentTimeStamp,formatted_time)
    return (currentTimeStamp, formatted_time)

def getCurrentTimeInfo1():
    currentTimeStamp = time.time()
    timestamp = currentTimeStamp
    formatted_time = time.strftime('%Y-%m-%d-%H-%M-%S', time.localtime(timestamp))
    # print(currentTimeStamp,formatted_time)
    return formatted_time


def save_request_data(url, headers, method, request_data, response_data):
    """保存请求及其响应到文件"""
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    filename = f'requests_log/{timestamp}.json'
    data = {
        'url': url,
        'headers': headers,
        'method': method,
        'request': request_data,
        'response': response_data
    }
    try:
        with open(filename, 'w') as file:
            json.dump(data, file, ensure_ascii=False, indent=4)
    except Exception as e:
        print(e)

def save_qa_data(qtext,atext):
    """保存问答到文件"""
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    filename = f'{QALogFilePath}/{timestamp}.txt'
    data = f'{qtext}\n******\n{atext}'
    try:
        with open(filename, 'w') as file:
            file.write(data)
    except Exception as e:
        print(e)

def list_request_records():
    """列出所有保存的请求记录文件"""
    files = os.listdir('requests_log') # get请求.json
    files.sort(reverse=True)
    return files

def list_qa_records():
    """列出所有保存的请求记录文件"""
    files = os.listdir('qa_log')
    files.sort(reverse=True)
    return files

def list_llm_records(which_para):
    """列出所有保存的请求记录文件"""
    if which_para=='apiURL':
        apiURL_list = find_value_of_key_in_nested_dict((read_json_file(f'{current_script_path}/configini.json')), "apiURL")
        apiURL_list.sort(reverse=True)
        return apiURL_list
    if which_para=='apiKey':
        apiKey_list = find_value_of_key_in_nested_dict((read_json_file(f'{current_script_path}/configini.json')), "apiKey")
        apiKey_list.sort(reverse=True)
        return apiKey_list
    if which_para=='apiHeaders':
        apiHeaders_list = find_value_of_key_in_nested_dict((read_json_file(f'{current_script_path}/configini.json')), "apiHeaders")
        #apiHeaders_list.sort(reverse=True)
        return apiHeaders_list

def find_all_key(dictionary):
    """
    在复杂宁典中找到所有的键。
    :param dictionary:输入的复杂宁典。
    :return:所有的健列表。
    """
    stack = [((), dictionary)]
    res = []
    while stack:
        path, current_dict = stack.pop()
        for key, value in current_dict.items():
            res.append(key)
            new_path = path + (key,)
            if isinstance(value, dict):
                stack.append((new_path, value))
    return res


def find_key_and_update_value(dictionary, key_to_find, new_value) -> None:
    """
    在复杂字典中找到指定键的路径，并更新其值
    :paramdictionary:输入的复杂宁典
    :param key_to_find:需要查找的键。
    :param new value: 新的值.
    :return: 如果找到键，返回True;否则返回FaLse。
    """
    stack = [((), dictionary)]
    flag = True
    while stack and flag:
        path, current_dict = stack.pop()
        for key, value in current_dict.items():
            new_path = path + (key,)
            if key == key_to_find:
                current_dict[key] = new_value
                flag = False
                break
            elif isinstance(value, dict):
                stack.append((new_path, value))

def generate_vin(length=17):
    valid_chars = string.digits + string.ascii_uppercase
    valid_chars = ''.join(filter(lambda x: x not in 'QIO', valid_chars))
    return ''.join(random.choices(valid_chars, k=length))

def generate_reqular_plate():
    province_codes = '京津沪渝冀豫云辽黑湘院鲁新苏浙赣鄂桂甘晋蒙陕吉闽贵粤青藏川宁琼'
    region_codes = ''.join((c for c in string.ascii_uppercase if c not in 'I0'))
    sequence = ''.join((c for c in string.ascii_uppercase if c not in 'I0'))
    sequence += ''.join((str(i) for i in range(10)))
    province = random.choice(province_codes)
    region = random.choice(region_codes)
    sequence = ''.join(random.choices(sequence, k=5))
    return f'{province}{region}{sequence}'

def generate_new_energy_plate(type='small'):
    province_codes = '京津沪渝冀豫云辽黑湘皖鲁新苏浙赣鄂桂甘蒙陕吉闽贵粤青川宁琼'
    province = random.choice(province_codes)
    prefix_letter = 'DF' if type == 'small' else random.choice('DF')
    sequence = ''.join(random.choices(string.digits, k=4)) if type == 'small' else ''.join(
        random.choices(string.digits + 'DF', k=5))
    return f"{province}{prefix_letter}{sequence}"

def generate_police_plate():
    province_codes = '京津沪渝冀豫云辽黑湘皖鲁新苏浙赣鄂桂甘晋蒙陕吉贵粤青藏川宁琼'
    region_codes = ''.join((c for c in string.ascii_uppercase if c not in 'I0'))
    province = random.choice(province_codes)
    region = random.choice(region_codes)
    number = ''.join(random.choices(string.digits, k=4))
    return f'{province}{region}{number}警'

def generate_consulate_plate():
    province_codes = '沪粤川云桂鄂闽鲁陕蒙藏黑辽渝'
    region_codes = 'ABCD'
    province = random.choice(province_codes)
    region = random.choice(region_codes)
    number = ''.join(random.choices(string.digits, k=4))
    return f'{province}{region}{number}领'

def generate_army_plate():
    army_codes = ''.join((c for c in string.ascii_uppercase if c in 'VKHBSLJNGCE'))
    region_codes = ''.join((c for c in string.ascii_uppercase if c not in 'I0'))
    army = random.choice(army_codes)
    region = random.choice(region_codes)
    number = ''.join(random.choices(string.digits, k=5))
    return f'{army}{region}{number}'

def generate_army_head_plate():
    head = random.choice(string.ascii_uppercase)
    head_number = random.choice('ADJRTVY')
    sequence = ''.join(random.choices(string.digits, k=5))
    return f'{head}{head_number}{sequence}'

def generate_ramdon_plate():
    # 随机选择车牌类型生成
    plate_types1 = {
        'regular': generate_reqular_plate,
        'new_energy_small': lambda: generate_new_energy_plate('small'),
        'new_energy_large': lambda: generate_new_energy_plate('large'),
        'police': generate_police_plate,
        'consulate': generate_consulate_plate,
        'army': generate_army_plate,
        'army_head': generate_army_head_plate
    }
    plate_types = {
        '常规车牌号': generate_reqular_plate,
        '小型新能源车牌号': lambda: generate_new_energy_plate('small'),
        '大型新能源车牌号': lambda: generate_new_energy_plate('large'),
        '警车车牌号': generate_police_plate,
        '领事馆车牌号': generate_consulate_plate,
        '武警车牌号': generate_army_plate,
        '军用车牌号': generate_army_head_plate
    }
    # 随机生成车牌
    plate_type = random.choice(list(plate_types.keys()))
    plate_number = plate_types[plate_type]()
    print(f"车牌号: {plate_number}，车牌类型: {plate_type}")
    return (plate_number, plate_type)


def generate_id_card():
    idCardList = ['450903199003131632', '450206195206027256', '4512231997090568394', '451321196403185690']
    return random.choice(idCardList)  # id cand number

## 在嵌套字典中查找指定键对应的值。
def find_value_in_nested_dict(dict_obj, key):
    if key in dict_obj:
        # 如果键在当前字典层级中，直接返口值
        return dict_obj[key]
    for k, v in dict_obj.items():
        # 如果值是字典，则递归查找
        if isinstance(v, dict):
            result = find_value_in_nested_dict(v, key)
            if result is not None:
                return result
    # 如果键在所有嵌套字典中都未找到，返回None
    return None

# 用于保存所有插入的图片（防止被垃圾回收）
image_references = {} #python3.7后字典插入后是有序的

set_LLM={
    "apiURL": '',
    "apiKey": '',
    "apiHeaders": {},
}

class SimplePostmanApp(tk.Tk):
    def __init__(self):
        super().__init__()
        # 显式调用update方法
        self.update()

        # 设置窗口背景颜色
        self.configure(bg='lightskyblue',background='lightskyblue')

        try:
            # 设置桌面图标，这里需要提供一个ico件的路径
            self.iconbitmap(f'{current_script_path}/happy.ico')
        except:
            pass

        try:
            # 设置窗口图标
            self.wm_iconbitmap(bitmap=f'{current_script_path}/happy.ico')
        except:
            pass

        # 获取电脑分辨率
        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()
        # 设置坐标
        x = (screen_width - 1250) // 2
        y = (screen_height - 750) // 2
        # 设置应用名
        self.title("simplepostman1.0")
        # 设置展示位置
        self.geometry(f"{1250}x{750}+{x}+{y-50}")

        # 设置窗口自适应
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(5, weight=1)

        # 界面布局设置
        self.create_mainWin_widgets_for_tkText_pack()

        # 初始化加载历史记录
        self.load_records()

        # 绑定快捷键《CtrL+Z)到撒操作的函数
        self.bind("<Control-Z>", lambda event: self.undo_operation(event))
        # # 绑定快捷《CtrL+Z)到撤销操作的函数#
        # self.bind("<Control-Z>", lambda event: self.undo_operation())

        # 绑定快捷键《CtrL+L)到撒操作的函数
        self.bind_all("<Control-L>", lambda event: self.start_region_selection_only_see_img())

        self.configure(bg='lightblue')

    def goto_line(self, text_widget, line_number):
        """"""
        # 创建一个消息提示子窗口
        text_widget.mark_set('insert', f'{line_number}.0')
        text_widget.see('insert')
        text_widget.focus()



    def updateLLMSet(self, paraDict):
        try:
            set_LLM['apiURL'] = paraDict['apiURL']
            set_LLM['apiKey'] = paraDict['apiKey']
            set_LLM['apiHeaders'] = paraDict['apiHeaders']
            print(set_LLM)
            self.messageInformInWin('update is successful!', 1500)
        except Exception as e:
            self.messageInformInWin(f'update is unsuccessful!Error is {e}', 1500)

    def copy_image_to_clipboard(self):
        # 获取 Text 控件中的图片对象（假设已知图片的索引或标记）
        try:
            last_item = next(reversed(image_references.items()))
            file_path=last_item[0] # 获取图片路径
            print(file_path)
            subprocess.run(['osascript', '-e', f'set the clipboard to (read (POSIX file "{file_path}") as JPEG picture)'])
            self.messageInformInWin(f'复制图片成功', 1000)
        except Exception as e:
            self.messageInformInWin(f'复制图片失败：{e}，请先确保已经执行截图或者选图等操作',1000)

    def paste_image_to_text(self, which_text, pattern):
        try:
            last_item = next(reversed(image_references.items()))
            img_path = last_item[0]  # 从全局引用处获取图片路径

            # 1. 打开图片并获取控件宽度，图片转换为PhotoImage
            image = Image.open(img_path)
            photo = ImageTk.PhotoImage(image)
            which_text.image_create("end", image=photo)
            widget_width = which_text.winfo_width() - 20  # 预留边距

            # 2. 计算缩放比例（保持宽高比）
            if widget_width < 1:  # 避免初始宽度为0
                widget_width = 300  # 默认宽度
            ratio = widget_width / image.width
            new_height = int(image.height * ratio)

            # 3. 缩放图片
            resized_image = image.resize((widget_width, new_height), Image.LANCZOS)
            photo = ImageTk.PhotoImage(resized_image)

            # 4. 插入文字和图片并保持引用
            self.clearContent(which_text)
            # which_text.insert("end", "")  # 插入文字
            which_text.image_create("end", image=photo)  # 在末尾插入图片
            which_text.insert("end", "\n")  # 继续插入文字
            which_text.image = photo  # 防止被垃圾回收
            self.messageInformInWin(f'粘贴图片成功', 1000)

            if pattern == 'qa':
                self.select_image_for_ocr_qa2(img_path)
            elif pattern == 'noqa':
                self.select_image_for_ocr_tk1(img_path)


        except Exception as e:
            self.messageInformInWin(f'粘贴图片失败：{e}，请先确保已经执行复制图片等操作',1000)

    def get_image_from_text(self, which_text, pattern):
        try:
            last_item = next(reversed(image_references.items()))
            img_path = last_item[0]  # 从全局引用处获取图片路径

            # 1. 打开图片并获取控件宽度，图片转换为PhotoImage
            image = Image.open(img_path)
            photo = ImageTk.PhotoImage(image)
            which_text.image_create("end", image=photo)
            widget_width = which_text.winfo_width() - 20  # 预留边距

            # 2. 计算缩放比例（保持宽高比）
            if widget_width < 1:  # 避免初始宽度为0
                widget_width = 300  # 默认宽度
            ratio = widget_width / image.width
            new_height = int(image.height * ratio)

            # 3. 缩放图片
            resized_image = image.resize((widget_width, new_height), Image.LANCZOS)
            photo = ImageTk.PhotoImage(resized_image)

            # 4. 插入文字和图片并保持引用
            self.clearContent(which_text)
            # which_text.insert("end", "")  # 插入文字
            which_text.image_create("end", image=photo)  # 在末尾插入图片
            which_text.insert("end", "\n")  # 继续插入文字
            which_text.image = photo  # 防止被垃圾回收
            self.messageInformInWin(f'获取图片成功', 1000)

            if pattern == 'qa':
                self.select_image_for_ocr_qa2(img_path)
            elif pattern == 'noqa':
                self.select_image_for_ocr_tk1(img_path)

        except Exception as e:
            self.messageInformInWin(f'获取图片失败：{e}，请先确保已经在文本中插入图片对象等操作',1000)

    def messageInformInWin(self, informContent, lastingTime):
        """
        # 定义一个函数，用于创建子窗口并在2秒后关闭它
        """
        # 创建一个消息提示子窗口
        popup = tk.Toplevel(self)
        popup.title("消息提示")

        # 获取电脑分辨率
        screen_width = popup.winfo_screenwidth()
        screen_height = popup.winfo_screenheight()
        # 设置坐标
        x = (screen_width - 200) // 2
        y = (screen_height - 100) // 2
        # 设置展示位置
        popup.geometry(f"{200}x{100}+{x}+{y}")
        # 设置窗口自适应
        popup.grid_columnconfigure(1, weight=1)
        popup.grid_rowconfigure(5, weight=1)

        # 在子窗口中添加一些内容
        # label = tk.Label(popup, text=informContent)
        # label.pack(pady=20)
        popup.content_text = tk.Text(
            popup,
            background='lightblue',
            highlightbackground='lightblue',
            foreground='black',
            bg='lightskyblue'
        )
        popup.content_text.pack(fill=tk.BOTH, expand=True)

        popup.content_text.insert('insert', informContent)

        # 定义一个函数，用于关闭子窗口
        def close_window():
            popup.destroy()

        # lastingTime后调用关闭窗口的函数
        popup.after(lastingTime, close_window)

    def messageInformInTransWin(self, informContent, lastingTime):
        """
        # 定义一个函数，用于创建子窗口并在2秒后关闭它
        """
        # 创建一个消息提示子窗口
        popup = tk.Toplevel(self)
        popup.title("翻译结果")

        # 获取电脑分辨率
        screen_width = popup.winfo_screenwidth()
        screen_height = popup.winfo_screenheight()
        # 设置坐标
        x = (screen_width - 400) // 2
        y = (screen_height - 200) // 2
        # 设置展示位置
        popup.geometry(f"{400}x{200}+{x}+{y}")
        # 设置窗口自适应
        popup.grid_columnconfigure(1, weight=1)
        popup.grid_rowconfigure(5, weight=1)

        # 在子窗口中添加一些内容
        # label = tk.Label(popup, text=informContent)
        # label.pack(pady=20)
        popup.content_text = tk.Text(
            popup,
            background='lightblue',
            highlightbackground='lightblue',
            foreground='black',
            bg='lightskyblue'
        )
        popup.content_text.pack(fill=tk.BOTH, expand=True)

        popup.content_text.insert('insert', informContent)

        # 定义一个函数，用于关闭子窗口
        def close_window():
            popup.destroy()

        # lastingTime后调用关闭窗口的函数
        popup.after(lastingTime, close_window)

    def undo_operation(self,event):
        print("微销操作")
        # 在这里添加撒销操作的具体实现代码

    def translateEnglishByLLM(self, keyword, model="gpt-4o-mini"):
        """单次提问函数"""
        # 配置API
        openai.api_key = \
        find_value_of_key_in_nested_dict((read_json_file(f'{current_script_path}/configini.json')), "apiKey")[0]
        openai.base_url = \
        find_value_of_key_in_nested_dict((read_json_file(f'{current_script_path}/configini.json')), "apiURL")[0]
        openai.default_headers = \
        find_value_of_key_in_nested_dict((read_json_file(f'{current_script_path}/configini.json')), "apiHeaders")[0]

        if set_LLM['apiKey']!='':
            openai.api_key = set_LLM['apiKey']
        if set_LLM['apiURL'] != '':
            openai.base_url = set_LLM['apiURL']
        if set_LLM['apiHeaders'] != '':
            openai.default_headers = set_LLM['apiHeaders']

        #question=f"{keyword}\n这个是英文单词(如果是单词的话,拼写是否错误?拼写错误的话请转换换为最相近的单词,并标出英文音标)还是英文段落?它是什么意思?请精准地翻译成中文"
        question = f'\"{keyword}\"这个是什么意思?'
        delay = 1.0
        print(question)
        #self.messageInformInWin("提问中......", 2000)
        try:
            print(f"\n正在处理问题: {question[:30]}...")

            completion = openai.chat.completions.create(
                model=model,
                messages=[{"role": "user", "content": question}],
                timeout=20  # 增加超时限制
            )
            answer = completion.choices[0].message.content
            # 打印当前结果
            print(f"【问题】{question}\n【回答】{answer}\n")
            save_qa_data(question,answer)
            self.continuing_save_question_and_ask(question,answer)
            self.messageInformInTransWin(answer,300000) #
            # self.load_qa_records(self.qa_record_combo)
            sleep(delay)  # 避免频繁请求
        except Exception as e:
            answer = f"调用{self.translateEnglishByLLM.__name__}方法发生ERROR: {str(e)}"
            print(answer)
            save_qa_data(question, answer)
            self.messageInformInTransWin(answer, 300000)
            sleep(delay)  # 避免频繁请求
        #return answer

    def askAI(self, qtext, atext, flagLabel="1", model="gpt-4o-mini"):
        """单次提问函数"""
        # 配置API
        openai.api_key = \
        find_value_of_key_in_nested_dict((read_json_file(f'{current_script_path}/configini.json')), "apiKey")[0]
        openai.base_url = \
        find_value_of_key_in_nested_dict((read_json_file(f'{current_script_path}/configini.json')), "apiURL")[0]
        openai.default_headers = \
        find_value_of_key_in_nested_dict((read_json_file(f'{current_script_path}/configini.json')), "apiHeaders")[0]

        if set_LLM['apiKey']!='':
            openai.api_key = set_LLM['apiKey']
        if set_LLM['apiURL'] != '':
            openai.base_url = set_LLM['apiURL']
        if set_LLM['apiHeaders'] != '':
            openai.default_headers = set_LLM['apiHeaders']

        question=qtext.get("1.0", tk.END)

        try:
            # 判断qtext内是否包含ImageTk.PhotoImage图片对象
            if self.has_photoimage(qtext) and flagLabel==1:
                print(f'文本框里有图片！')
                last_item = next(reversed(image_references.items()))
                img_path = last_item[0]  # 从全局引用处获取图片路径
                # self.get_image_from_text(qtext, 'qa')
                # self.clearContent(qtext)

                # 1. 打开图片并获取控件宽度，图片转换为PhotoImage
                image = Image.open(img_path)
                photo = ImageTk.PhotoImage(image)
                qtext.image_create("end", image=photo)
                widget_width = qtext.winfo_width() - 20  # 预留边距

                # 2. 计算缩放比例（保持宽高比）
                if widget_width < 1:  # 避免初始宽度为0
                    widget_width = 300  # 默认宽度
                ratio = widget_width / image.width
                new_height = int(image.height * ratio)

                # 3. 缩放图片
                resized_image = image.resize((widget_width, new_height), Image.LANCZOS)
                photo = ImageTk.PhotoImage(resized_image)

                # 4. 插入文字和图片并保持引用
                self.clearContent(qtext)
                # which_text.insert("end", "")  # 插入文字
                # qtext.image_create("end", image=photo)  # 在末尾插入图片
                # qtext.insert("end", "\n")  # 继续插入文字
                # qtext.image = photo  # 防止被垃圾回收
                self.messageInformInWin(f'获取图片成功', 3000)

                question = self.see_img_and_get_text(img_path, 'ch', 1)
                qtext.insert("end", f"\n{question}")  # 继续插入文字
                # image_references[img_path] = photo  # 保存引用
            elif flagLabel==0:
                print(f'文本框里没有图片！')
            else:
                print(f'文本框里没有图片！')
        except Exception as e:
            pass
            #self.messageInformInWin(f"错误是:{e}",40000)

        delay = 1.0
        self.messageInformInWin(f"提问中......", 2000)
        try:
            print(f"\n正在处理问题: {question[:30]}...")

            completion = openai.chat.completions.create(
                model=model,
                messages=[{"role": "user", "content": question}],
                timeout=20  # 增加超时限制
            )
            answer = completion.choices[0].message.content
            #extra_requirement='\n要求回答中不能包含*、-、+等字符，公式除外'
            #answer=answer.replace('*','')
            # 打印当前结果
            print(f"【问题】{question}\n【回答】{answer}\n")
            save_qa_data(question,answer)
            self.continuing_save_question_and_ask(question,answer)
            self.clearContent(atext)
            atext.insert(tk.END,f'{answer}')
            self.load_qa_records(self.qa_record_combo)
            self.qa_record_combo.set(list_qa_records()[0])
            sleep(delay)  # 避免频繁请求
        except Exception as e:
            answer = f"ERROR: {str(e)}"
            print(answer)
            save_qa_data(question, answer)
            atext.insert(tk.END, f'{answer}')
            self.load_qa_records(self.qa_record_combo)
            sleep(delay)  # 避免频繁请求
        #return answer

    def batch_askAI(self, questions, qtext, atext, model="gpt-4o-mini", delay=1.5):
        """批量提问函数"""
        # 配置API
        openai.api_key = \
        find_value_of_key_in_nested_dict((read_json_file(f'{current_script_path}/configini.json')), "apiKey")[0]
        openai.base_url = \
        find_value_of_key_in_nested_dict((read_json_file(f'{current_script_path}/configini.json')), "apiURL")[0]
        openai.default_headers = \
        find_value_of_key_in_nested_dict((read_json_file(f'{current_script_path}/configini.json')), "apiHeaders")[0]

        results = []
        for idx, question in enumerate(questions, 1):
            try:
                print(f"\n正在处理问题 {idx}/{len(questions)}: {question[:30]}...")

                completion = openai.chat.completions.create(
                    model=model,
                    messages=[{"role": "user", "content": question}],
                    timeout=20  # 增加超时限制
                )

                answer = completion.choices[0].message.content
                results.append((question, answer))

                # 打印当前结果
                print(f"【问题】{question}\n【回答】{answer[:100]}...\n")
                sleep(delay)  # 避免频繁请求

            except Exception as e:
                print(f"问题 {idx} 处理失败: {str(e)}")
                results.append((question, f"ERROR: {str(e)}"))

        #return results

    def read_all_row_records_from_excel(self,which_excelPath_combo,which_sheetName_combo,which_columnName_combo,pattern):
        """
        读取Excel文件中所有工作表的每一行数据
        参数:
        file_path (str): 输入文件路径
        返回:
        dict: 以工作表名为键，包含该表所有行数据的列表为值的字典
        "*.png *.jpg *.jpeg *.bmp")
        """
        """通过选择的记录填充表单"""
        filetypes = [("JSON Files", "*.xlsx *.csv")]
        excel_file_path = filedialog.askopenfilename(title="选择excel文件", filetypes=filetypes)
        # 打开一个文件选择对话框，用户选择json文件
        if not isfile(excel_file_path):  # 检查所选文件是否存在
            print(f"文件不存在:{excel_file_path}")
            #which_entry.url_entry.delete(0, tk.END)
        else:
            pass
        try:
            # 使用ExcelFile对象提高读取效率
            excel_file = pd.ExcelFile(excel_file_path)
            sheet_data = {}
            res=[]
            sheet_name_res = []
            column_name_res = []
            # 遍历每个工作表
            for sheet_name in excel_file.sheet_names:
                # 读取整个工作表
                df = excel_file.parse(sheet_name)
                sheet_name_res.append(sheet_name)
                # 将DataFrame转换为行数据列表（每行转为字典）
                rows = df.to_dict('records')
                sheet_data[sheet_name] = rows
                res.append([sheet_name]+rows)
                print([sheet_name]+rows)
                # 打印预览信息
                print(f"\n工作表 '{sheet_name}' 共 {len(rows)} 行数据:")
                column_name_res.extend(df.columns.tolist())
                print("属性包括:", df.columns.tolist())
            if pattern==0:
                pass
                #return res
            if pattern==1:
                which_excelPath_combo['values'] = [excel_file_path]
                which_excelPath_combo.set(which_excelPath_combo['values'][0])
                which_sheetName_combo['values'] = sheet_name_res
                which_sheetName_combo.set(which_sheetName_combo['values'][0])
                which_columnName_combo['values'] = column_name_res
                which_columnName_combo.set(which_columnName_combo['values'][0])
                #return sheet_name_res

        except FileNotFoundError:
            print(f"调用{self.read_all_row_records_from_excel.__name__}方法错误: 文件未找到 - {excel_file_path}")
            self.messageInformInWin(f"调用{self.read_all_row_records_from_excel.__name__}方法错误: 文件未找到 - {excel_file_path}",6000)
            return []
        except Exception as e:
            print(f"调用{self.read_all_row_records_from_excel.__name__}方法错误: 处理文件时发生错误 - {e}")
            self.messageInformInWin(f"调用{self.read_all_row_records_from_excel.__name__}方法错误: 处理文件时发生错误 - {e}",6000)
            return []

    def modify_keywords_in_a_sheet_of_excel(self, file_path, sheetName, columnName, yourRule, keyword='',output_path=None):
        """
        修改 Excel 文件工作表中的某列数据，并保留所有未修改的工作表
        参数:
        file_path (str): 输入文件路径
        sheetName (str): 要修改的工作表名
        columnName (str): 要修改的列名
        yourRule (str): 修改规则（目前仅支持'clearEmpty'）
        output_path (str, optional): 输出文件路径，默认为None表示覆盖原文件
        """
        try:
            # 读取Excel文件
            excel_file = pd.ExcelFile(file_path)

            # 获取所有工作表名
            sheet_names = excel_file.sheet_names

            # 用于存储所有工作表数据（包括修改和未修改的）
            all_dfs = {}

            # 先读取所有工作表到字典
            for sheet in sheet_names:
                all_dfs[sheet] = excel_file.parse(sheet)

                # 处理指定工作表
            if sheetName in all_dfs:
                df = all_dfs[sheetName]
                if columnName in df.columns:
                    if yourRule == 'stripEmptyInColumn':
                        # 清除指定列前后的空格、换行符等
                        df[columnName] = df[columnName].astype(str).str.replace('\n', '', regex=False)  # 去除换行符
                        df[columnName] = df[columnName].str.strip()  # 去除首尾空格
                        print(
                            f"调用{self.modify_keywords_in_a_sheet_of_excel.__name__} 方法成功修改工作表 '{sheetName}' 中的\"{columnName}\"列")
                    elif yourRule == 'clearKeyChInColumn' and keyword!='':
                        # 清除指定列里的指定字符
                        df[columnName] = df[columnName].str.replace(keyword, '')  # 去除换行符
                        print(
                            f"调用{self.modify_keywords_in_a_sheet_of_excel.__name__} 方法成功去除工作表 '{sheetName}' 中的\"{columnName}\"列的\"{keyword}\"字符")

                    elif yourRule == 'other':
                        pass  # 其他规则可以在这里扩展
                    else:
                        pass
                else:
                    print(
                        f"调用{self.modify_keywords_in_a_sheet_of_excel.__name__} 方法发现工作表 '{sheetName}' 中不存在\"{columnName}\"列")
            else:
                print(f"调用{self.modify_keywords_in_a_sheet_of_excel.__name__} 方法发现文件中不存在工作表 '{sheetName}'")

            # 确定输出路径
            if output_path is None:
                output_path = file_path

                # 将所有工作表（包括修改和未修改的）写入新文件
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                for sheet_name, df in all_dfs.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)

            print(f"调用{self.modify_keywords_in_a_sheet_of_excel.__name__} 方法成功更新所有工作表并保存到: {output_path}")
            self.messageInformInWin(
                f"调用{self.modify_keywords_in_a_sheet_of_excel.__name__} 方法成功更新所有工作表并保存到: {output_path}",
                3000
            )

        except FileNotFoundError:
            print(f"调用{self.modify_keywords_in_a_sheet_of_excel.__name__} 方法错误: 文件未找到 - {file_path}")
            self.messageInformInWin(
                f"调用{self.modify_keywords_in_a_sheet_of_excel.__name__} 方法错误: 文件未找到 - {file_path}",
                3000
            )
        except Exception as e:
            print(f"调用{self.modify_keywords_in_a_sheet_of_excel.__name__} 方法错误: 处理文件时发生错误 - {e}")
            self.messageInformInWin(
                f"调用{self.modify_keywords_in_a_sheet_of_excel.__name__} 方法错误: 处理文件时发生错误 - {e}",
                3000
            )

    def view_data_in_a_sheet_of_excel(self, file_path, sheetName, which_display_text):
        """
        修改 Excel 文件工作表中的某列数据，并保留所有未修改的工作表
        参数:
        file_path (str): 输入文件路径
        sheetName (str): 要修改的工作表名
        columnName (str): 要修改的列名
        which_display_text (t.Text): 展示的文本框
        """
        try:
            # 使用ExcelFile对象提高读取效率
            excel_file = pd.ExcelFile(file_path)
            sheet_data = {}
            res=[]
            sheet_name_res = []
            column_name_res = []
            # 读取工作表
            df = excel_file.parse(sheetName)
            sheet_name_res.append(sheetName)
            # 将DataFrame转换为行数据列表（每行转为字典）
            rows = df.to_dict('records')
            sheet_data[sheetName] = rows
            res.append([sheetName]+rows)

            excelData=str(rows).replace("\'","\"")
            self.clearContent(which_display_text)
            which_display_text.insert(tk.END,f'{excelData}')

            print([sheetName]+rows)
            # 打印预览信息
            print(f"\n工作表 '{sheetName}' 共 {len(rows)} 行数据:")
            column_name_res.extend(df.columns.tolist())
            print("属性包括:", df.columns.tolist())

        except FileNotFoundError:
            print(f"调用{self.read_all_row_records_from_excel.__name__}方法错误: 文件未找到 - {file_path}")
            self.messageInformInWin(f"调用{self.read_all_row_records_from_excel.__name__}方法错误: 文件未找到 - {file_path}",6000)
        except Exception as e:
            print(f"调用{self.read_all_row_records_from_excel.__name__}方法错误: 处理文件时发生错误 - {e}")
            self.messageInformInWin(f"调用{self.read_all_row_records_from_excel.__name__}方法错误: 处理文件时发生错误 - {e}",6000)

    def read_all_row_records_from_excel1(self,excel_file_path):
        try:
            # 使用ExcelFile对象提高读取效率
            excel_file = pd.ExcelFile(excel_file_path)
            sheet_data = {}
            res=[]
            # 遍历每个工作表
            for sheet_name in excel_file.sheet_names:
                # 读取整个工作表
                df = excel_file.parse(sheet_name)
                # 将DataFrame转换为行数据列表（每行转为字典）
                rows = df.to_dict('records')
                sheet_data[sheet_name] = rows
                res.append([sheet_name]+rows)
                print([sheet_name]+rows)
                # 打印预览信息
                print(f"\n工作表 '{sheet_name}' 共 {len(rows)} 行数据:")
                print("属性包括:", df.columns.tolist())
            return res

        except FileNotFoundError:
            print(f"调用{self.read_all_row_records_from_excel1.__name__}方法错误: 文件未找到 - {excel_file_path}")
            self.messageInformInWin(f"调用{self.read_all_row_records_from_excel.__name__}方法错误: 文件未找到 - {excel_file_path}",6000)
            return []
        except Exception as e:
            print(f"调用{self.read_all_row_records_from_excel1.__name__}方法错误: 处理文件时发生错误 - {e}")
            self.messageInformInWin(f"调用{self.read_all_row_records_from_excel1.__name__}方法错误: 处理文件时发生错误 - {e}",6000)
            return []

    def continuing_save_question_and_ask(self,qtex,atex):
        """逐次存储问答数据"""
        question=qtex
        answer=atex
        save_qa_file="save_qa_file.xlsx"
        # 创建新的Excel工作簿
        if not os.path.exists(save_qa_file):
            pd.DataFrame(columns=["时间", "问题", "回答"]).to_excel(save_qa_file, index=False)
        try:
            # 读取现有数据
            try:
                df_existing = pd.read_excel(save_qa_file)
            except:
                df_existing = pd.DataFrame(columns=["时间", "问题", "回答"])
            # 追加新数据
            df_new = pd.DataFrame([[getCurrentTimeInfo1(), question, answer]], columns=["时间", "问题", "回答"])
            df_combined = pd.concat([df_existing, df_new], ignore_index=True)
            # 写入Excel
            df_combined.to_excel(save_qa_file, index=False)
        except Exception as e:
            print(f"ERROR: {str(e)}")
            error_msg = f"ERROR: {str(e)}"
            # 即使出错也写入Excel
            try:
                df_existing = pd.read_excel(save_qa_file)
            except:
                df_existing = pd.DataFrame(columns=["时间", "问题", "回答"])
            # 追加新数据
            df_new = pd.DataFrame([[getCurrentTimeInfo1(), question, error_msg]], columns=["时间", "问题", "回答"])
            df_combined = pd.concat([df_existing, df_new], ignore_index=True)
            # 写入Excel
            df_combined.to_excel(save_qa_file, index=False)

    def write_xmind_to_excel(self, which_combobox):
        # 设置文件类型过滤器，只显示XMind文件(.xmind)
        filetypes = [('XMind files', '*.xmind')]
        xmind_file_path = filedialog.askopenfilename(title="选择xmind文件", filetypes=filetypes)
        # 打开一个文件选择对话框，用户选择xmind文件
        if not isfile(xmind_file_path):
            # 检查所选文件是否存在
            print(f"文件不存在:{xmind_file_path}")
        else:
            pass

        caseType = which_combobox.get()

        # 将Xmind解析为多叉树
        xmind_tree = xmind_to_tree(xmind_file_path)

        # 获取所有叶子节点的路径
        all_leaf_paths_list = get_all_leaf_paths_data_list(xmind_tree, caseType)
        # print(all leaf_paths_list)

        write_xmind_data_to_excel(all_leaf_paths_list, caseType)  ###

        # 清空输入框的内容
        self.response_text.delete("1.0", "end-1c")
        # 输入框插入内容
        self.response_text.insert('insert', f'生成excel路径为: {current_script_path}/{caseExcelFileName}')

    def create_SetBodyKey_sub_window(self):
        # 创建设置入参子窗口
        sub_window = tk.Toplevel(self)
        sub_style = ttk.Style(sub_window)  # 为子窗口创建独立的样式Style
        sub_style.theme_use('clam')
        # 配置Combobox及其下拉菜单样式
        sub_style.configure('TCombobox', fieldbackground='white', background='white', foreground='black')
        sub_style.configure('TCombobox.Listbox', fieldbackground='white', background='white', foreground='black')

        try:
            # 设置子窗口图标
            sub_window.wm_iconbitmap(bitmap=f'{current_script_path}/happy.ico')
        except:
            pass
        # 设置窗口背景颜色
        sub_window.configure(bg='lightskyblue')
        # 获取电脑分辨率
        screen_width = sub_window.winfo_screenwidth()
        screen_height = sub_window.winfo_screenheight()
        # 设置坐标
        x = (screen_width - 400) // 2
        y = (screen_height - 300) // 2
        # 设置应用名
        sub_window.title("SetHeadersAndBodyKey")
        # 设置展示位置
        sub_window.geometry(f"{400}x{300}+{x + 655}+{y}")
        # 设置窗口自适应
        sub_window.grid_columnconfigure(1, weight=1)
        sub_window.grid_rowconfigure(5, weight=1)

        self.create_setBodyKey_sub_widgets(sub_window)

    def create_SetLLM_sub_window(self):
        # 创建设置入参子窗口
        sub_window = tk.Toplevel(self)
        self.create_menu(sub_window)
        sub_style = ttk.Style(sub_window)  # 为子窗口创建独立的样式Style
        sub_style.theme_use('clam')
        # 配置Combobox及其下拉菜单样式
        sub_style.configure('TCombobox', fieldbackground='white', background='white', foreground='black')
        sub_style.configure('TCombobox.Listbox', fieldbackground='white', background='white', foreground='black')

        try:
            # 设置子窗口图标
            sub_window.wm_iconbitmap(bitmap=f'{current_script_path}/happy.ico')
        except:
            pass
        # 设置窗口背景颜色
        sub_window.configure(bg='lightskyblue')
        # 获取电脑分辨率
        screen_width = sub_window.winfo_screenwidth()
        screen_height = sub_window.winfo_screenheight()
        # 设置坐标
        x = (screen_width - 400) // 2
        y = (screen_height - 300) // 2
        # 设置应用名
        sub_window.title("SetLLM")
        # 设置展示位置
        sub_window.geometry(f"{400}x{300}+{x + 655}+{y}")
        # 设置窗口自适应
        sub_window.grid_columnconfigure(1, weight=1)
        sub_window.grid_rowconfigure(5, weight=1)

        self.create_setLLM_sub_widgets(sub_window)

    def create_modify_excel_sub_window(self):
        # 创建设置入参子窗口
        sub_window = tk.Toplevel(self)
        sub_style = ttk.Style(sub_window)  # 为子窗口创建独立的样式Style
        sub_style.theme_use('clam')
        # 配置Combobox及其下拉菜单样式
        sub_style.configure('TCombobox', fieldbackground='white', background='white', foreground='black')
        sub_style.configure('TCombobox.Listbox', fieldbackground='white', background='white', foreground='black')

        try:
            # 设置子窗口图标
            sub_window.wm_iconbitmap(bitmap=f'{current_script_path}/happy.ico')
        except:
            pass
        # 设置窗口背景颜色
        sub_window.configure(bg='lightskyblue')
        # 获取电脑分辨率
        screen_width = sub_window.winfo_screenwidth()
        screen_height = sub_window.winfo_screenheight()
        # 设置坐标
        x = (screen_width - 1000) // 2
        y = (screen_height - 600) // 2
        # 设置应用名
        sub_window.title("ModifyExcel")
        # 设置展示位置
        sub_window.geometry(f"{1000}x{600}+{x}+{y-50}")
        # 设置窗口自适应
        sub_window.grid_columnconfigure(1, weight=1)
        sub_window.grid_rowconfigure(5, weight=1)

        self.create_modify_excel_sub_widgets(sub_window)

    def create_tools_sub_window(self):
        # 创建工具子窗口
        sub_window = tk.Toplevel(self)
        sub_style = ttk.Style(sub_window)  # 为子窗口创建独立的样式Style
        sub_style.theme_use('clam')
        # 配置Combobox及其下拉菜单样式
        sub_style.configure('TCombobox', fieldbackground='white', background='white', foreground='black')
        sub_style.configure('TCombobox.Listbox', fieldbackground='white', background='white', foreground='black')

        try:
            # 设置子窗口图标
            sub_window.wm_iconbitmap(bitmap=f'{current_script_path}/happy.ico')
        except:
            pass
        # 设置窗口背景颜色
        sub_window.configure(
            bg='lightskyblue',
            background='lightblue',
            highlightbackground='lightblue'
        )
        # 获取电脑分辨率
        screen_width = sub_window.winfo_screenwidth()
        screen_height = sub_window.winfo_screenheight()
        # 设置坐标
        x = (screen_width - 1100) // 2
        y = (screen_height - 500) // 2
        # 设置应用名
        sub_window.title("BusinessTools")
        # 设置展示位置
        sub_window.geometry(f"{1100}x{500}+{x}+{y-80}")
        # 设置窗口自适应
        sub_window.grid_columnconfigure(1, weight=1)
        sub_window.grid_rowconfigure(5, weight=1)
        # 在子窗口中添加标签
        # Label = tk.Label(sub_window，text="这是一个子窗口")
        # label.pack()
        self.create_business_tools_sub_widgets(sub_window)

    def create_tools_sub_window1(self):
        # 创建工具子窗口
        sub_window = tk.Toplevel(self)
        sub_style = ttk.Style(sub_window)  # 为子窗口创建独立的样式Style
        sub_style.theme_use('clam')
        # 配置Combobox及其下拉菜单样式
        sub_style.configure('TCombobox', fieldbackground='white', background='white', foreground='black')
        sub_style.configure('TCombobox.Listbox', fieldbackground='white', background='white', foreground='black')

        try:
            # 设置子窗口图标
            sub_window.wm_iconbitmap(bitmap=f'{current_script_path}/happy.ico')
        except:
            pass
        # 设置窗口背景颜色
        sub_window.configure(background='lightskyblue')
        # 获取电脑分辨率
        screen_width = sub_window.winfo_screenwidth()
        screen_height = sub_window.winfo_screenheight()
        # 设置坐标
        x = (screen_width - 1200) // 2
        y = (screen_height - 600) // 2
        # 设置应用名
        sub_window.title("ModifyImg")
        # 设置窗口展示位置和大小
        sub_window.geometry(f"{1200}x{600}+{x}+{y}")
        # 设置窗口自适应
        sub_window.grid_columnconfigure(1, weight=1)
        sub_window.grid_rowconfigure(5, weight=1)
        # 在子窗口中添加标签
        # Label = tk.Label(sub_window，text="这是一个子窗口")
        # label.pack()
        self.create_ocr_text_sub_widgets(sub_window)

    def create_tools_sub_window_1(self):
        # 创建工具子窗口
        sub_window = tk.Toplevel(self)
        sub_style = ttk.Style(sub_window)  # 为子窗口创建独立的样式Style
        sub_style.theme_use('clam')
        # 配置Combobox及其下拉菜单样式
        sub_style.configure('TCombobox', fieldbackground='white', background='white', foreground='black')
        sub_style.configure('TCombobox.Listbox', fieldbackground='white', background='white', foreground='black')

        try:
            # 设置子窗口图标
            sub_window.wm_iconbitmap(bitmap=f'{current_script_path}/happy.ico')
        except:
            pass
        # 设置窗口背景颜色
        sub_window.configure(bg='lightskyblue')
        # 获取电脑分辨率
        screen_width = sub_window.winfo_screenwidth()
        screen_height = sub_window.winfo_screenheight()
        # 设置坐标
        x = (screen_width - 800) // 2
        y = (screen_height - 400) // 2
        # 设置应用名
        sub_window.title("ModifyImg")
        # 设置展示位置
        sub_window.geometry(f"{800}x{400}+{x}+{y}")
        # 设置窗口自适应
        sub_window.grid_columnconfigure(1, weight=1)
        sub_window.grid_rowconfigure(5, weight=1)
        # 在子窗口中添加标签
        # Label = tk.Label(sub_window，text="这是一个子窗口")
        # label.pack()
        self.create_modify_img_sub_widgets(sub_window)

    def create_tools_sub_window2(self):
        # 创建工具子窗口
        sub_window = tk.Toplevel(self)
        sub_style = ttk.Style(sub_window)  # 为子窗口创建独立的样式Style
        sub_style.theme_use('clam')
        # 配置Combobox及其下拉菜单样式
        sub_style.configure('TCombobox', fieldbackground='white', background='white', foreground='black')
        sub_style.configure('TCombobox.Listbox', fieldbackground='white', background='white', foreground='black')

        try:
            # 设置子窗口图标
            sub_window.wm_iconbitmap(bitmap=f'{current_script_path}/happy.ico')
        except:
            pass
        # 设置窗口背景颜色
        sub_window.configure(background='lightskyblue')
        # 获取电脑分辨率
        screen_width = sub_window.winfo_screenwidth()
        screen_height = sub_window.winfo_screenheight()
        # 设置坐标
        x = (screen_width - 1200) // 2
        y = (screen_height - 600) // 2
        # 设置应用名
        sub_window.title("SeeImgToTxt")
        # 设置窗口展示位置和大小
        sub_window.geometry(f"{1200}x{600}+{x}+{y}")
        # 设置窗口自适应
        sub_window.grid_columnconfigure(1, weight=1)
        sub_window.grid_rowconfigure(5, weight=1)
        # 在子窗口中添加标签
        # Label = tk.Label(sub_window，text="这是一个子窗口")
        # label.pack()
        self.create_ocr_text_sub_widgets(sub_window)

    def create_tools_sub_window3(self):
        # 创建工具子窗口
        sub_window = tk.Toplevel(self)
        sub_style = ttk.Style(sub_window)  # 为子窗口创建独立的样式Style
        sub_style.theme_use('clam')
        # 配置Combobox及其下拉菜单样式
        sub_style.configure('TCombobox', fieldbackground='white', background='white', foreground='black')
        sub_style.configure('TCombobox.Listbox', fieldbackground='white', background='white', foreground='black')

        try:
            # 设置子窗口图标
            sub_window.wm_iconbitmap(bitmap=f'{current_script_path}/happy.ico')
        except:
            pass
        # 设置窗口背景颜色
        sub_window.configure(background='lightskyblue',bg='lightskyblue',highlightbackground='lightblue')
        # 获取电脑分辨率
        screen_width = sub_window.winfo_screenwidth()
        screen_height = sub_window.winfo_screenheight()
        # 设置坐标
        x = (screen_width - 1300) // 2
        y = (screen_height - 700) // 2
        # 设置应用名
        sub_window.title("QaByLLM")
        # 设置窗口展示位置和大小
        sub_window.geometry(f"{1300}x{700}+{x}+{y-80}")
        # 设置窗口自适应
        sub_window.grid_columnconfigure(1, weight=1)
        sub_window.grid_rowconfigure(5, weight=1)
        # 在子窗口中添加标签
        # Label = tk.Label(sub_window，text="这是一个子窗口")
        # label.pack()
        self.create_ocr_qa_sub_widgets(sub_window)

    # def toggle_window(self,which_win):
    #     if which_win.state() == 'normal':
    #         which_win.withdraw()  # 隐藏窗口
    #     else:
    #         which_win.deiconify()  # 显示窗口

    def create_tools_sub_window4(self):
        # 创建工具子窗口
        sub_window = tk.Toplevel(self)

        # # 绑定全局快捷键：Command+Shift+L
        # keyboard.add_hotkey('Command+Shift+L', lambda :self.toggle_window(sub_window))

        sub_style = ttk.Style(sub_window)  # 为子窗口创建独立的样式Style
        sub_style.theme_use('clam')
        # 配置Combobox及其下拉菜单样式
        sub_style.configure('TCombobox', fieldbackground='white', background='white', foreground='black')
        sub_style.configure('TCombobox.Listbox', fieldbackground='white', background='white', foreground='black')

        # try:
        #     # 设置子窗口图标
        #     sub_window.wm_iconbitmap(bitmap=f'{current_script_path}/happy.ico')
        # except:
        #     pass
        # sub_window.withdraw()
        # 设置窗口背景颜色
        sub_window.configure(bg='systemTransparent')
        # sub_window.attributes('-topmost', True)  # 强制置顶
        #self.force_topmost(sub_window)
        # 获取电脑分辨率
        screen_width = sub_window.winfo_screenwidth()
        screen_height = sub_window.winfo_screenheight()
        # 设置坐标
        x = (screen_width - 100) // 1
        y = (screen_height - 60) // 1
        # x = (screen_width - 150) // 2
        # y = (screen_height - 30) // 2
        # 设置应用名
        sub_window.title("")
        # 设置窗口展示位置和大小
        sub_window.geometry(f"{100}x{60}+{x}+{y}")
        # 设置窗口自适应
        sub_window.grid_columnconfigure(1, weight=1)
        sub_window.grid_rowconfigure(5, weight=1)
        # 在子窗口中添加标签
        # Label = tk.Label(sub_window，text="这是一个子窗口")
        # label.pack()
        self.create_screenshot_sub_widgets(sub_window)

    def force_topmost(self,window):
        """
        Mac全屏状态下打开的有三个窗口，要求实现tkinter窗口能在切换到这些全屏窗口时能继续强制置顶
        """
        # 获取 NSWindow 对象
        ns_window = objc.objc_object(c_void_p=window.winfo_id())
        # 设置窗口层级为最高（覆盖全屏应用）
        ns_window.setLevel_(3)  # 3 对应 kCGMaximumWindowLevelKey
        ns_window.setCollectionBehavior_(1 << 3)  # 1<<3 为

    def change_picture_format(self):
        # # 创建设置子窗口
        # sub_window = tk.Toplevel(self)
        # # 不显示子窗口
        # sub_window.withdraw()
        # 设置只打开图片文件
        filetypes = [('Image files', ',jpg ,jpeg .png ,gif .bmp')]
        img_path = filedialog.askopenfilename(title="选择图片文件", filetypes=filetypes)
        # 打开一个文件选择对话框，用户选择图片文件
        if not isfile(img_path):
            # 检查所选文件是否存在
            print(f"文件不存在:{img_path}")
        else:
            img = Image.open(img_path)
            # 打开图片文件
            icon_sizes = [(64, 64)]
            # 定义图标大小
            base_name = splitext(img_path)[0]
            # 提取图像文件基本名称《不包括扩展名)
            icon_filename = f'{base_name}.ico'
            img.save(icon_filename, sizes=icon_sizes)
            print(f"图标文件已保存为:{icon_filename}")

    def set_font_size(self, family, size, weight):
        # bold_font = Font(family='Helvetica', size=15, weight='bold')
        font=Font(family=family, size=size, weight=weight)
        return font

    def createComboboxByOptionMenu(self, which_window, optionsList):
        """
        创建一个下拉框窗口，并提供选择功能。
        :param title: 窗口的标题
        :param options: 下拉框的选项列表
        :return: 当前选中的选项，如果未选择则返回 None
        """
        try:
            # 检查选项列表是否为空
            if not optionsList:
                raise ValueError("选项列表不能为空")
            # 创建一个StringVar来存储当前选中的值
            selected_option = tk.StringVar(which_window)
            selected_option.set(optionsList[0])  # 设置默认值

            # 1. 创建OptionMenu
            dropdown = tk.OptionMenu(which_window, selected_option, *optionsList)
            dropdown.pack(side=tk.LEFT, fill=tk.X, expand=True, pady=20)

            # 2. 获取下拉菜单对象并设置字体
            menu = which_window.nametowidget(dropdown["menu"])  # 获取关联的菜单对象
            menu.config(font=("Song", 10 , "normal"))  # 设置下拉项的字体和大小

            # 3. 修改默认显示的按钮字体
            dropdown.config(font=("Song", 10, "normal"))  # 设置顶部显示文本的字体
            # # 添加按钮以获取当前选中的值
            # def show_selected():
            #     print("当前选中:", selected_option.get())
            #     # 返回选中的值
            #     which_window.quit()  # 关闭窗口
            # tk.Button(which_window, text="显示选中项", command=show_selected).pack()
            # 返回当前选中的选项
            # selected_option.get()
        except Exception as e:
            print(f"发生错误: {e}")
            self.messageInformInWin(f"发生错误: {e}",1500)

    # # 示例调用
    # if __name__ == "__main__":
    #     selected_value = create_dropdown_window("下拉框示例 - OptionMenu", ["选项1", "选项2", "选项3", "选项4"])
    #     print(f"最终选择的选项是: {selected_value}")


    def create_mainWin_widgets_for_tkText_pack(self):
        # 主容器布局（横向分左中右三部分）
        sub_style = ttk.Style(self)  # 为子窗口创建独立的样式Style
        sub_style.theme_use('clam')
        # 配置Combobox及其下拉菜单样式
        sub_style.configure('TCombobox', fieldbackground='white', background='white', foreground='black')
        sub_style.configure('TCombobox.Listbox', fieldbackground='white', background='white', foreground='black')

        self.create_menu(self)

        left_frame = tk.Frame(self, bg="lightblue")
        center_frame = tk.Frame(self, bg="lightblue")
        right_frame = tk.Frame(self, bg="lightblue")
        left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        center_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        right_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True) # fill=tk.Y)

        # --- URL输入区域 ---
        url_frame = tk.Frame(left_frame, bg="lightblue")
        url_frame.pack(fill=tk.X, pady=2)

        self.url_label = tk.Label(
            url_frame,
            background='white',
            font=self.set_font_size('Song', 12, 'normal'),
            foreground='black',
            text="URL:"
        )
        self.url_label.pack(side=tk.LEFT)

        self.url_entry = tk.Entry(
            url_frame,
            bg="lightskyblue",
            highlightbackground='white'
        )
        self.url_entry.pack(side=tk.LEFT, fill=tk.X, expand=True)

        # --- 方法选择区域 ---
        method_frame = tk.Frame(left_frame, bg="lightblue",highlightbackground='white')
        method_frame.pack(fill=tk.X, pady=2)

        self.method_label = tk.Label(
            method_frame,
            background='white',
            font=self.set_font_size('Song', 12, 'normal'),
            foreground='black',
            text="Method:"
        )
        self.method_label.pack(side=tk.LEFT)

        #self.method_combobox = self.createComboboxByOptionMenu(method_frame, ["GET", "POST", "PUT", "DELETE"])

        self.method_combobox = ttk.Combobox(
            method_frame,
            font=self.set_font_size('Song', 10, 'normal'),
            values=["GET", "POST", "PUT", "DELETE"]
        )
        self.method_combobox.pack(side=tk.LEFT, fill=tk.X, expand=True)
        self.method_combobox.current(0)

        # --- Headers输入区域 ---
        headers_frame = tk.Frame(left_frame, bg="lightblue")
        headers_frame.pack(fill=tk.BOTH, expand=True, pady=5)

        self.headers_label = tk.Label(
            headers_frame,
            background='white',
            font=self.set_font_size('Song', 12, 'normal'),
            foreground='black',
            text="Headers(JSON):"
        )
        self.headers_label.pack(anchor='w')

        headers_scrollbar = tk.Scrollbar(headers_frame)
        headers_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.headers_text = tk.Text(
            headers_frame,
            yscrollcommand=headers_scrollbar.set,
            background='lightblue',
            highlightbackground='white', #tk.Text背景色
            font=self.set_font_size('Song', 12, 'normal'),
            foreground='black',
            width=45,
            height=10,
            bg="lightskyblue"
        )
        self.headers_text.pack(fill=tk.BOTH, expand=True)
        self.add_function_buttons_in_request_text_for_tkText_pack(headers_frame, self.headers_text, 1)
        headers_scrollbar.config(command=self.headers_text.yview)

        # --- Body输入区域 ---
        body_frame = tk.Frame(left_frame, bg="lightblue")
        body_frame.pack(fill=tk.BOTH, expand=True, pady=5)

        self.body_label = tk.Label(
            body_frame,
            background='white',
            font=self.set_font_size('Song', 12, 'normal'),
            foreground='black',
            text="Body(JSON):"
        )
        self.body_label.pack(anchor='w')

        # self.body_text = scrolledtext.ScrolledText(
        body_scrollbar = tk.Scrollbar(body_frame)
        body_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.body_text = tk.Text(
            body_frame,
            yscrollcommand=body_scrollbar.set,
            background='lightblue',
            highlightbackground='white',  # tk.Text背景色
            font=self.set_font_size('Song', 12, 'normal'),
            foreground='black',
            width=45,
            height=15,
            bg="lightskyblue"
        )
        self.body_text.pack(fill=tk.BOTH, expand=True)
        self.add_function_buttons_in_request_text_for_tkText_pack(body_frame, self.body_text, 1)
        body_scrollbar.config(command=self.body_text.yview)

        # --- 登陆设置框架 ---
        login_seting_frame = tk.Frame(left_frame, bg="lightblue")
        login_seting_frame.pack(fill=tk.X, pady=5)

        # 选择前置登录接口记录的路径
        self.before_login_buton = tk.Button(
            login_seting_frame,
            text="BeforeLoginPath",
            font=self.set_font_size('Song', 10, 'normal'),
            anchor='center',
            highlightbackground='lightblue',  # tk.Button按钮背景色
            command=self.findLoginRecordFilePath
        )
        self.before_login_buton.pack(side=tk.LEFT)

        #设置ttk.Combobox背景色，全局生效
        style = ttk.Style()
        style.theme_use('clam')  # 必须使用支持自定义的主题
        # 配置Combobox及其下拉菜单
        style.configure('TCombobox', fieldbackground='white',background='white', foreground='black')
        style.configure('TCombobox.Listbox', fieldbackground='white',background='white', foreground='black')

        self.before_login_combobox = ttk.Combobox(
            login_seting_frame,
            font=self.set_font_size('Song', 10, 'normal'),
            values=[f'{current_script_path}/reguests_log/post数字员工请求登录demo.json']
        )
        self.before_login_combobox.pack(side=tk.LEFT)
        self.before_login_combobox.current(0)

        # --- 发送按钮框架 ---
        send_frame = tk.Frame(left_frame, bg="lightblue")
        send_frame.pack(fill=tk.X, pady=5)

        self.send_label = tk.Label(
            send_frame,
            background='white',
            font=self.set_font_size('Song', 12, 'normal'),
            foreground='black',
            text="Send Request:"
        )
        self.send_label.pack(side=tk.LEFT)

        self.send_button = tk.Button(
            send_frame,
            text="Send",
            bg='#FFB6C1',
            fg='black',
            highlightbackground='blue',#tk.Button按钮背景色
            font=self.set_font_size('Song', 12, 'normal'),
            command=lambda: self.thread_it(self.send_request_by_userInputData)
        )
        self.send_button.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)

        # --- Response输出区域 ---
        response_frame = tk.Frame(center_frame, bg="lightblue")
        response_frame.pack(fill=tk.BOTH, expand=True)

        self.response_label = tk.Label(
            response_frame,
            background='white',
            font=self.set_font_size('Song', 12, 'normal'),
            foreground='black',
            text="Response(JSON):"
        )
        self.response_label.pack(anchor='w')

        # self.response_text = scrolledtext.ScrolledText(
        response_scrollbar = tk.Scrollbar(response_frame)
        response_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.response_text = tk.Text(
            response_frame,
            yscrollcommand=response_scrollbar.set,
            background='lightblue',
            highlightbackground='white',  # tk.Text背景色
            font=self.set_font_size('Song', 12, 'normal'),
            foreground='black',
            width=45,
            height=10,
            bg="lightskyblue"
        )
        self.response_text.pack(fill=tk.BOTH, expand=True)
        self.add_function_buttons_in_request_text_for_tkText_pack(response_frame, self.response_text, 1)
        response_scrollbar.config(command=self.response_text.yview)

        # --- 右侧功能按钮区域 ---
        self.set_button = tk.Button(
            right_frame,
            text="Setkey",
            highlightbackground='lightblue',  # tk.Button按钮背景色
            command=self.create_SetBodyKey_sub_window
        )
        self.set_button.pack(fill=tk.X, pady=2)

        self.set_button_modifyExcel = tk.Button(
            right_frame,
            text="ModifyExcel",
            highlightbackground='lightblue',  # tk.Button按钮背景色
            command=self.create_modify_excel_sub_window
        )
        self.set_button_modifyExcel.pack(fill=tk.X, pady=2)

        self.set_button_businessTools = tk.Button(
            right_frame,
            text="BusinessTools",
            highlightbackground='lightblue',  # tk.Button按钮背景色
            command=self.create_tools_sub_window
        )
        self.set_button_businessTools.pack(fill=tk.X, pady=2)

        self.set_button_ChangeImg = tk.Button(
            right_frame,
            text="ModifyImg",
            highlightbackground='lightblue',  # tk.Button按钮背景色
            command=self.create_tools_sub_window1
        )
        self.set_button_ChangeImg.pack(fill=tk.X, pady=2)

        self.set_button_SeeImgToTxt = tk.Button(
            right_frame,
            text="SeeImgToTxt",
            highlightbackground='lightblue',  # tk.Button按钮背景色
            command=self.create_tools_sub_window2
        )
        self.set_button_SeeImgToTxt.pack(fill=tk.X, pady=2)

        self.set_button_QaByAI = tk.Button(
            right_frame,
            text="QaByLLM",
            highlightbackground='lightblue',  # tk.Button按钮背景色
            command=self.create_tools_sub_window3
        )
        self.set_button_QaByAI.pack(fill=tk.X, pady=2)

        self.set_button_setLLM = tk.Button(
            right_frame,
            text="SetLLM",
            highlightbackground='lightblue',  # tk.Button按钮背景色
            command=self.create_SetLLM_sub_window
        )
        self.set_button_setLLM.pack(fill=tk.X, pady=2)

        self.set_button_screenshot = tk.Button(
            right_frame,
            text="ScreenSub",
            highlightbackground='lightblue',  # tk.Button按钮背景色
            command=self.create_tools_sub_window4
        )
        self.set_button_screenshot.pack(fill=tk.X, pady=2)

        self.app_path_copy_button = tk.Button(
            right_frame,
            text="AppPathCopy",
            highlightbackground='lightblue',  # tk.Button按钮背景色
            command=self.app_path_copy
        )
        self.app_path_copy_button.pack(fill=tk.X, pady=2)

        # --- 底部记录操作区域 ---
        bottom_frame = tk.Frame(center_frame, bg="lightblue")
        bottom_frame.pack(fill=tk.X, pady=5)

        # 复制关键字对应的值
        self.select_key_btn = tk.Button(
            bottom_frame,
            text="CopyResKey",
            highlightbackground='lightblue',  # tk.Button按钮背景色
            font=self.set_font_size('Song', 10, 'normal'),
            anchor="center",
            command=lambda: self.copy_value_of_key(self, self.response_text, self.select_key_combobox.get())
        )
        self.select_key_btn.pack(side=tk.LEFT)

        # 设置关键字输入框
        self.select_key_combobox = ttk.Combobox(
            bottom_frame,
            font=self.set_font_size('Song', 10, 'normal'),
            state="NORMAL"
        )
        self.select_key_combobox.pack(side=tk.LEFT)
        self.select_key_combobox["values"] = ['reportNo']
        self.select_key_combobox["values"] = find_value_in_nested_dict(
            read_json_file(f'{current_script_path}/configini.json'), "CopyKey")
        self.select_key_combobox.current(0)  # 设置默认值为列表中的第一个元素

        self.refresh_response_text_key_button = tk.Button(
            bottom_frame,
            text="RefreshKey",
            font=self.set_font_size('Song', 10, 'normal'),
            highlightbackground='lightblue',  # tk.Button按钮背景色
            anchor="center",
            command=lambda: self.refreshKey(self.response_text, self.select_key_combobox)
        )
        self.refresh_response_text_key_button.pack(side=tk.LEFT)

        # 搜索按钮
        self.search_btn = tk.Button(
            bottom_frame,
            text="FindReskey",
            highlightbackground='lightblue',  # tk.Button按钮背景色
            font=self.set_font_size('Song', 10, 'normal'),
            anchor="center",
            command=lambda: self.search_keyword_in_text(self.search_combobox.get(), self.response_text)
        )
        self.search_btn.pack(side=tk.LEFT)

        # 搜索response关键字结果
        self.search_combobox = ttk.Combobox(
            bottom_frame,
            font=self.set_font_size('Song', 10, 'normal'),
            state='NORMAL'
        )
        self.search_combobox.pack(side=tk.LEFT)
        self.search_combobox['values'] = ['CollegeName']

        bottom_frame1 = tk.Frame(center_frame, bg="lightblue")
        bottom_frame1.pack(fill=tk.X, pady=5)

        self.select_record_button = tk.Button(
            bottom_frame1,
            text="Select Records:",
            highlightbackground='lightblue',  # tk.Button按钮背景色
            font=self.set_font_size('Song', 10, 'normal'),
            command=lambda :self.find_file_to_fill_record(self.record_combobox)
        )
        self.select_record_button.pack(side=tk.LEFT)

        self.record_combobox = ttk.Combobox(
            bottom_frame1,
            state='readonly',
            values=list_request_records(),
            font=self.set_font_size('Song', 10, 'normal'),
            width=50
        )
        self.record_combobox.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        #self.record_combobox.current(0)
        self.record_combobox.bind('<<ComboboxSelected>>', self.fill_record)
        self.record_combobox.set(list_request_records()[1])

        self.delete_record_button = tk.Button(
            bottom_frame1,
            text="Delete Records",
            highlightbackground='lightblue',  # tk.Button按钮背景色
            font=self.set_font_size('Song', 10, 'normal'),
            command=lambda: self.thread_it(self.delete_records)
        )
        self.delete_record_button.pack(side=tk.LEFT)

        # 样式配置（保留原有逻辑）
        self.response_text.tag_config("match", background="yellow")

    def app_path_copy(self):
        input_content = f'{current_script_path}/'  # 将内容复制到剪贴板
        self.clipboard_clear()
        self.clipboard_append(input_content)  # 显示复制成功的消息《可选》
        print(f"{input_content}内容已复制到剪贴板。")  # messagebox.showinfo(input_content)
        # self.messageInformInWin(input_content, 2800)
        time.sleep(1)

    def create_mainWin_widgets1(self):
        sub_style = ttk.Style(self)  # 为子窗口创建独立的样式Style
        sub_style.theme_use('clam')
        # 配置Combobox及其下拉菜单样式
        sub_style.configure('TCombobox', fieldbackground='white', background='white', foreground='black')
        sub_style.configure('TCombobox.Listbox', fieldbackground='white', background='white', foreground='black')

        # URL输入
        self.url_label = tk.Label(
            self,
            background='white',
            font=self.set_font_size('Song', 12, 'normal'),
            foreground='black',  # 字体颜色
            text="URL:"
        )
        self.url_label.grid(column=0, row=0)

        self.url_entry = tk.Entry(self, bg="lightskyblue")
        self.url_entry.grid(column=1, row=0, sticky='EW')  # , columnspan=2)

        self.set_button = tk.Button(self, text="Setkey", command=self.create_SetBodyKey_sub_window)  # 组件按钮格式化功能
        self.set_button.grid(column=2, row=0)

        # 请求方法选择
        self.method_label = tk.Label(
            self,
            background='white',
            font=self.set_font_size('Song', 12, 'normal'),
            foreground='black',  # 字体颜色
            text="Method:"
        )
        self.method_label.grid(column=0, row=1)

        self.method_combobox = ttk.Combobox(
            self,
            values=["GET", "POST", "PUT", "DELETE"]
        )
        self.method_combobox.grid(column=1, row=1, sticky='EW')
        self.method_combobox.current(0)

        # 选择前置登录接口记录的路径
        self.before_login_buton = tk.Button(
            self,
            text="BeforeLoginpath",
            font=self.set_font_size('Song', 10, 'normal'),
            command=self.findLoginRecordFilePath
        )
        self.before_login_buton.grid(column=2, row=1)

        self.before_login_combobox = ttk.Combobox(
            self,
            values=[f'{current_script_path}/reguests_log/post数字员工请求登录demo.json']
        )
        self.before_login_combobox.grid(column=3, row=1, sticky='EW')
        self.before_login_combobox.current(0)

        # Headers输入
        self.headers_label = tk.Label(
            self,
            background='white',
            font=self.set_font_size('Song', 12, 'normal'),
            foreground='black',  # 字体颜色
            text="Headers(JSON):"
        )
        self.headers_label.grid(column=0, row=2)

        self.headers_text = scrolledtext.ScrolledText(
            self,
            background='lightblue',
            font=self.set_font_size('Song', 12, 'normal'),
            foreground='black',  # 字体颜色
            width=45,
            height=10,
            bg="lightskyblue"
        )
        self.headers_text.grid(column=1, row=2, rowspan=2, sticky='NSEW')
        # scrolledtext.ScrolledText内添加功能按钮
        self.add_function_buttons_in_request_text(self, self.headers_text, 0)

        # Body输入
        self.body_label = tk.Label(
            self,
            background='white',
            font=self.set_font_size('Song', 12, 'normal'),
            foreground='black',  # 字体颜色
            text="Body(JSON):"
        )
        self.body_label.grid(column=0, row=4)

        self.body_text = scrolledtext.ScrolledText(
            self,
            background='lightblue',
            font=self.set_font_size('Song', 12, 'normal'),
            foreground='black',  # 字体颜色
            width=45,
            height=30,
            bg="lightskyblue"
        )
        self.body_text.grid(column=1, row=4, rowspan=2, sticky='NSEW')
        # scrolledtext.ScrolledText内添加功能按钮
        self.add_function_buttons_in_request_text(self, self.body_text, 0)

        # 发送按钮
        self.send_label = tk.Label(
            self,
            background='white',
            font=self.set_font_size('Song', 12, 'normal'),
            foreground='black',  # 字体颜色
            text="Send Request:"
        )
        self.send_label.grid(column=0, row=6)

        # self,send_button = ttk,Button(self,text="Send", command=self.send_request)不加线程的话会一直不回弹,导致主线程卡死
        self.send_button = tk.Button(
            self,
            text="Send",
            bg='#FFB6C1',
            fg='black',
            font=self.set_font_size('Song', 12, 'normal'),
            command=lambda: self.thread_it(self.send_request_by_userInputData)
        )
        self.send_button.grid(column=1, row=6, sticky='EW') # 加线程

        self.refresh_response_text_key_button = tk.Button(
            self,
            text="RefreshKey",
            anchor="center",
            command=lambda: self.refreshKey(self.response_text, self.select_key_combobox)
        )
        self.refresh_response_text_key_button.grid(column=2, row=7)

        # 复制关键字对应的值
        self.select_key_btn = tk.Button(
            self,
            text="CopyResKey",
            anchor="center",
            command=lambda: self.copy_value_of_key(self, self.response_text,self.select_key_combobox.get())
        )
        self.select_key_btn.grid(column=2, row=8)

        # 设置关键字输入框
        self.select_key_combobox = ttk.Combobox(self, state="NORMAL")
        self.select_key_combobox.grid(column=3, row=8, sticky='EW')
        self.select_key_combobox["values"] = ['reportNo']
        self.select_key_combobox["values"] = find_value_in_nested_dict(
            read_json_file(f'{current_script_path}/configini.json'), "CopyKey")
        self.select_key_combobox.current(0)  # 设置默认值为列表中的第一个元素

        # Response响应输出
        self.response_label = tk.Label(
            self,
            background='white',
            font=self.set_font_size('Song', 12, 'normal'),
            foreground='black',  # 字体颜色
            text="Response(JSON):"
        )
        self.response_label.grid(column=0, row=7)

        self.response_text = scrolledtext.ScrolledText(
            self,
            background='lightblue',
            font=self.set_font_size('Song', 12, 'normal'),
            foreground='black',  # 字体颜色
            width=45,
            height=30,
            bg="lightskyblue"
        )
        self.response_text.grid(column=1, row=7, rowspan=13, sticky='EW')  # ,columnspan=3
        # scrolledtext.ScrolledText内添加功能按钮
        self.add_function_buttons_in_request_text(self, self.response_text, 1)

        # 搜索按钮
        self.search_btn = tk.Button(
            self,
            text="FindReskey",
            anchor="center",
            command=lambda: self.search_text(self.search_combobox.get(), self.response_text)
        )
        self.search_btn.grid(column=2, row=10)

        # 搜索response关键字结果
        self.search_combobox = ttk.Combobox(self, state='NORMAL')
        self.search_combobox.grid(column=3, row=10, sticky='EW')
        self.search_combobox['values'] = ['CollegeName']
        self.search_combobox.current(0)  # 设置默认值为列表中的第一个元素

        self.set_button_businessTools = tk.Button(self, text="BusinessTools", command=self.create_tools_sub_window)  # 组件按钮格式化功能
        self.set_button_businessTools.grid(column=2, row=2)

        self.set_button_businessTools = tk.Button(self, text="ModifyExcel", command=self.create_modify_excel_sub_window)  # 组件按钮格式化功能
        self.set_button_businessTools.grid(column=3, row=2)

        self.set_button_ChangeImg = tk.Button(self, text="ChangeImg", command=self.create_tools_sub_window1)  # 组件按钮格式化功能
        self.set_button_ChangeImg.grid(column=2,row=3)

        self.set_button_SeeImgToTxt = tk.Button(self, text="SeeImgToTxt", command=self.create_tools_sub_window2)  # 组件按钮格式化功能
        self.set_button_SeeImgToTxt.grid(column=2,row=4)

        self.set_button_QaByAI = tk.Button(self, text="QaByLLM", command=self.create_tools_sub_window3)  # 组件按钮格式化功能
        self.set_button_QaByAI.grid(column=2,row=5)
        self.set_button_setLLM = tk.Button(self, text="SetLLM", command=self.create_SetLLM_sub_window)  # 组件按钮格式化功能
        self.set_button_setLLM.grid(column=3,row=5)

        self.set_button_screenshot = tk.Button(self, text="ScreenSub", command=self.create_tools_sub_window4)  # 组件按钮格式化功能
        self.set_button_screenshot.grid(column=3,row=4)

        self.select_record_button = tk.Button(
            self,
            text="Select Records:",
            anchor='center',
            command=lambda : self.find_file_to_fill_record(self.record_combobox)
        )
        self.select_record_button.grid(column=0, row=20)
        self.record_combobox = ttk.Combobox(self, state='readonly', width=50)
        self.record_combobox.grid(column=1, row=20, sticky='EW')
        self.record_combobox.bind('<<ComboboxSelected>>', self.fill_record)

        self.delete_record_button = tk.Button(
            self,
            text="Delete Records",
            anchor='center',
            command=lambda: self.thread_it(self.delete_records)
        )
        self.delete_record_button.grid(column=2, row=20)

        # 为匹配文本定义样式
        self.response_text.tag_config("match", background="yellow")  #

    def create_setBodyKey_sub_widgets(self, sub_win):
        sub_style = ttk.Style(sub_win)  # 为子窗口创建独立的样式Style
        sub_style.theme_use('clam')
        # 配置Combobox及其下拉菜单样式
        sub_style.configure('TCombobox', fieldbackground='white', background='white', foreground='black')
        sub_style.configure('TCombobox.Listbox', fieldbackground='white', background='white', foreground='black')

        # surveyUm标签
        sub_win.surveyUm_label = ttk.Label(sub_win, text="surveyUm:")
        sub_win.surveyUm_label.grid(column=0, row=0)
        # 选择surveyUm
        sub_win.surveyUm_combobox = ttk.Combobox(sub_win, state='NORMAL')
        sub_win.surveyUm_combobox.grid(column=1, row=0, sticky='EW')
        # sub_win.surveyUm_combobox['values'] =['CHENKETAN6096','HUJUN464']

        # 获取配置里的UM账号
        UMList = read_json_file(f'{current_script_path}/configini.json')['UM']
        sub_win.surveyUm_combobox['values'] = UMList

        sub_win.surveyUm_combobox.current(0)  # 设置默认值为列表中的第一个元素
        print(sub_win.surveyUm_combobox.get())

        # JSON文件名为configin.json
        json_file_path = f'{current_script_path}/configini.json'

        # 保存surveyUm
        sub_win.set_surveyUm_button = ttk.Button(
            sub_win,
            text="save",
            command=lambda: self.changeKeyValueInJsonFile(json_file_path,'surveyUm',sub_win.surveyUm_combobox.get())
        )  # 组件按期定格式
        sub_win.set_surveyUm_button.grid(column=2, row=0)

        # surveyUmphone标签
        sub_win.surveyUmphone_label = ttk.Label(sub_win, text="surveyUmphone:")
        sub_win.surveyUmphone_label.grid(column=0, row=1)
        # 选择surveyUmphone
        sub_win.surveyUmphone_combobox = ttk.Combobox(sub_win, state='NORMAL')
        sub_win.surveyUmphone_combobox.grid(column=1, row=1, sticky='EW')
        sub_win.surveyUmphone_combobox['values'] = ['17374899426']
        sub_win.surveyUmphone_combobox.current(0)  # 设置默认值为列表中的第一个元素
        # 保存surveyUmphone
        sub_win.set_surveyumphone_button = ttk.Button(
            sub_win,
            text="save",
            command=lambda: self.changeKeyValueInJsonFile(json_file_path,"surveyumphone",sub_win.surveyUmphone_combobox.get())
        )
        sub_win.set_surveyumphone_button.grid(column=2, row=1)

        # mobileNo标签
        sub_win.mobileNo_label = ttk.Label(sub_win, text="mobileNo:")
        sub_win.mobileNo_label.grid(column=0, row=2)
        # 选择mobileNo
        sub_win.mobileNo_combobox = ttk.Combobox(sub_win, state='NORMAL')
        sub_win.mobileNo_combobox.grid(column=1, row=2, sticky='EW')
        sub_win.mobileNo_combobox['values'] = ['17374899426']
        sub_win.mobileNo_combobox.current(0)  # 设置默认值为列表中的第一个元素
        # 保存mobileNo
        sub_win.set_mobileNo_button = ttk.Button(
            sub_win,
            text="save",
            command=lambda: self.changeKeyValueInJsonFile(json_file_path,'mobileNo',sub_win.mobileNo_combobox.get())
        )  # 组件按钮绑定格式化函数功能
        sub_win.set_mobileNo_button.grid(column=2, row=2)

        # 选取修改headers里的键值
        sub_win.headers_text_key_label = ttk.Label(sub_win, text="selectHeadersKey:")
        sub_win.headers_text_key_label.grid(column=0, row=10)
        # 选择headers修改的键
        sub_win.headers_text_key_combobox = ttk.Combobox(sub_win, state='NORMAL')
        sub_win.headers_text_key_combobox.grid(column=1, row=10, sticky='EW')
        sub_win.headers_text_key_combobox["values"] = ["Authorization", "Content-Type", "X-Portal-Token", "Cookie"]
        sub_win.headers_text_key_combobox.current(0)  # 设置默认值为列表中的第一个元素
        # 刷新键值
        sub_win.refresh_headers_text_keyValue_button = ttk.Button(
            sub_win,
            text="refreshKey",
            command=lambda: self.refreshKey(self.headers_text,sub_win.headers_text_key_combobox)
        )
        sub_win.refresh_headers_text_keyValue_button.grid(column=2, row=10)

        # 修改headers里的键值
        sub_win.headers_text_keyValue_label = ttk.Label(sub_win, text="updateHeadersKey:")
        sub_win.headers_text_keyValue_label.grid(column=0, row=11)
        # 选择headers修改
        sub_win.headers_text_keyValue_combobox = ttk.Combobox(sub_win, state='NORMAL')
        sub_win.headers_text_keyValue_combobox.grid(column=1, row=11, sticky='EW')
        sub_win.headers_text_keyValue_combobox['values'] = ['456']
        sub_win.headers_text_keyValue_combobox.current(0)  # 设置默认值为列表中的第一个元素
        # 更改键值
        sub_win.set_headers_text_keyValue_button = ttk.Button(
            sub_win,
            text="change",
            command=lambda: self.changeKeyValueInJsonstr(
                self.headers_text,
                sub_win.headers_text_key_combobox.get(),
                sub_win.headers_text_keyValue_combobox.get()
            )
        )
        sub_win.set_headers_text_keyValue_button.grid(column=2, row=11)

        # 选取修改body里的键值
        sub_win.body_text_key_label = ttk.Label(sub_win, text="selectBodyKey:")
        sub_win.body_text_key_label.grid(column=0, row=12)
        # 选择body修改的键
        sub_win.body_text_key_combobox = ttk.Combobox(sub_win, state="NORMAL")
        sub_win.body_text_key_combobox.grid(column=1, row=12, sticky='EW')
        sub_win.body_text_key_combobox['values'] = ["reportNo", "businessKey"]
        sub_win.body_text_key_combobox.current(0)  # 设置默认值为列表中的第一个元素
        # 刷新键值
        sub_win.refresh_body_text_keyValue_button = ttk.Button(
            sub_win,
            text="refreshkey",
            command=lambda: self.refreshKey(self.body_text,sub_win.body_text_key_combobox)
        )
        sub_win.refresh_body_text_keyValue_button.grid(column=2, row=12)

        # 修改body里的键值
        sub_win.body_text_keyValue_label = ttk.Label(sub_win, text="updateBodyKey:")
        sub_win.body_text_keyValue_label.grid(column=0, row=13)
        # 选择body修改的键
        sub_win.body_text_keyValue_combobox = ttk.Combobox(sub_win, state='NORMAL')
        sub_win.body_text_keyValue_combobox.grid(column=1, row=13, sticky='EW')
        sub_win.body_text_keyValue_combobox['values'] = ['456']
        sub_win.body_text_keyValue_combobox.current(0)  # 设置默认值为列表中的第一个元素
        # 更改键值
        sub_win.set_body_text_keyValue_button = ttk.Button(
            sub_win,
            text="change",
            command=lambda: self.changeKeyValueInJsonstr(
                self.body_text,
                sub_win.body_text_key_combobox.get(),
                sub_win.body_text_keyValue_combobox.get()
            )
        )
        sub_win.set_body_text_keyValue_button.grid(column=2, row=13)

    def create_modify_excel_sub_widgets(self, sub_win):
        sub_style = ttk.Style(sub_win)  # 为子窗口创建独立的样式Style
        sub_style.theme_use('clam')
        # 配置Combobox及其下拉菜单样式
        sub_style.configure('TCombobox', fieldbackground='white', background='white', foreground='black')
        sub_style.configure('TCombobox.Listbox', fieldbackground='white', background='white', foreground='black')

        left_frame = tk.Frame(sub_win, bg="lightblue")
        # center_frame = tk.Frame(sub_win, bg="lightblue")
        # right_frame = tk.Frame(sub_win, bg="lightblue")
        left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        # center_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        # right_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)  # fill=tk.Y)

        # --- sheetname选择区域 ---
        sheetname_frame0 = tk.Frame(left_frame, bg="lightblue")
        sheetname_frame0.pack(fill=tk.X, pady=2)

        # 手动选择excel文件标签
        sub_win.choiceExcel_label = tk.Label(
            sheetname_frame0,
            text="ChoiceFile:",
            font=self.set_font_size('Song', 10, 'normal'),
            foreground='black',
            background='white',
            highlightbackground='white'
        )
        sub_win.choiceExcel_label.pack(side=tk.LEFT)

        # 手动选择excel文件按钮
        sub_win.choiceExcel_button = tk.Button(
            sheetname_frame0,
            font=self.set_font_size('Song', 10, 'normal'),
            text="ChoiceExcel",
            bg="white",
            highlightbackground='lightblue',
            command=lambda: self.read_all_row_records_from_excel(
                sub_win.ExcelPath_combobox,
                sub_win.sheetName_combobox,
                sub_win.columnName_combobox,
                1
            )
        )  # 组件按期定格式
        sub_win.choiceExcel_button.pack(side=tk.LEFT)

        # --- sheetname选择区域 ---
        sheetname_frame1 = tk.Frame(left_frame, bg="lightblue")
        sheetname_frame1.pack(fill=tk.X, pady=2)

        # ExcelPath标签
        sub_win.ExcelPath_label = tk.Label(
            sheetname_frame1,
            text="ExcelPath:",
            font=self.set_font_size('Song', 10, 'normal'),
            foreground='black',
            background='white',
            highlightbackground='white'
        )
        sub_win.ExcelPath_label.pack(side=tk.LEFT)

        # 下拉框选择ExcelPath
        sub_win.ExcelPath_combobox = ttk.Combobox(
            sheetname_frame1,
            state='NORMAL',
            font=self.set_font_size('Song', 10, 'normal'),
            values=['']
        )
        sub_win.ExcelPath_combobox.pack(side=tk.LEFT, fill=tk.X, expand=True)
        sub_win.ExcelPath_combobox.current(0)

        sheetname_frame2 = tk.Frame(left_frame, bg="lightblue")
        sheetname_frame2.pack(fill=tk.X, pady=2)

        # sheetname标签
        sub_win.sheetname_label = tk.Label(
            sheetname_frame2,
            text="SheetName:",
            font=self.set_font_size('Song', 10, 'normal'),
            foreground='black',
            background='white',
            highlightbackground='white'
        )
        sub_win.sheetname_label.pack(side=tk.LEFT)

        # 下拉框选择sheetname
        sub_win.sheetName_combobox = ttk.Combobox(
            sheetname_frame2,
            state='NORMAL',
            font=self.set_font_size('Song', 10, 'normal'),
            values=['']
        )
        sub_win.sheetName_combobox.pack(side=tk.LEFT, fill=tk.X, expand=True)
        sub_win.sheetName_combobox.current(0)

        sheetname_frame3 = tk.Frame(left_frame, bg="lightblue")
        sheetname_frame3.pack(fill=tk.X, pady=2)

        # columnName标签
        sub_win.columnName_label = tk.Label(
            sheetname_frame3,
            text="columnName:",
            font=self.set_font_size('Song', 10, 'normal'),
            foreground='black',
            background='white',
            highlightbackground='white'
        )
        sub_win.columnName_label.pack(side=tk.LEFT)

        # 下拉框选择columnName
        sub_win.columnName_combobox = ttk.Combobox(
            sheetname_frame3,
            state='NORMAL',
            font=self.set_font_size('Song', 10, 'normal'),
            values=['']
        )
        sub_win.columnName_combobox.pack(side=tk.LEFT, fill=tk.X, expand=True)
        sub_win.columnName_combobox.current(0)

        # --- sheetname选择区域 ---
        sheetname_frame4 = tk.Frame(left_frame, bg="lightblue")
        sheetname_frame4.pack(fill=tk.X, pady=2)

        # 查找控件（嵌套Frame实现水平排列）
        sub_win.excel_operation_combobox = ttk.Combobox(
            sheetname_frame4,
            font=self.set_font_size('Song', 10, 'normal'),
            background='white',
            values=['stripEmptyInColumn', 'clearKeyChInColumn', ''],
            state='NORMAL'
        )
        sub_win.excel_operation_combobox.pack(side=tk.LEFT)  # ,  expand=True, fill="both")
        sub_win.excel_operation_combobox.current(0)

        # 修改excel文件
        sub_win.choiceExcel_button = tk.Button(
            sheetname_frame4,
            font=self.set_font_size('Song', 10, 'normal'),
            text="ModifyExcel",
            bg="white",
            highlightbackground='lightblue',
            command=lambda: self.excelCommonOperation(
                sub_win,
                sub_win.excel_operation_combobox,
                sub_win.ExcelPath_combobox,
                sub_win.sheetName_combobox,
                sub_win.columnName_combobox
            )
        )  # 组件按期定格式
        sub_win.choiceExcel_button.pack(side=tk.LEFT)



        # 查看excel文件
        sub_win.viewExcel_button = tk.Button(
            sheetname_frame4,
            font=self.set_font_size('Song', 10, 'normal'),
            text="ViewExcel",
            bg="white",
            highlightbackground='lightblue',
            command=lambda: self.thread_it(
                self.view_data_in_a_sheet_of_excel(
                    sub_win.ExcelPath_combobox.get(),
                    sub_win.sheetName_combobox.get(),
                    sub_win.excel_display_text
                )
            )
        )  # 组件按期定格式
        sub_win.viewExcel_button.pack(side=tk.LEFT)

        # 查找控件（嵌套Frame实现水平排列）
        sub_win.find_btn_combobox = ttk.Combobox(
            sheetname_frame4,
            font=self.set_font_size('Song', 10, 'normal'),
            background='white',
            values=['*', ',', '、', '\n', ' ', '\t'],
            state='NORMAL'
        )
        sub_win.find_btn_combobox.pack(side=tk.LEFT)  # ,  expand=True, fill="both")
        sub_win.find_btn_combobox.current(0)

        sub_win.find_replace_btn = tk.Button(
            sheetname_frame4,
            font=self.set_font_size('Song', 10, 'normal'),
            background='white',
            highlightbackground='lightblue',
            foreground='black',
            text="Find",
            command=lambda: self.find_or_replace(
                sub_win.excel_display_text,
                sub_win.find_btn_combobox,
                sub_win.replace_btn_combobox,
                'find'
            )
        )
        sub_win.find_replace_btn.pack(side=tk.LEFT)

        # 替换控件（嵌套Frame实现水平排列）
        sub_win.replace_btn_combobox = ttk.Combobox(
            sheetname_frame4,
            font=self.set_font_size('Song', 10, 'normal'),
            background='white',
            values=['*', ',', '、', '\n', ' ', '\t'],
            state='NORMAL'
        )
        sub_win.replace_btn_combobox.pack(side="left", padx=2)  # ,  expand=True, fill="both")
        sub_win.replace_btn_combobox.current(0)

        sub_win.replace_btn = tk.Button(
            sheetname_frame4,
            font=self.set_font_size('Song', 10, 'normal'),
            background='white',
            highlightbackground='lightblue',
            foreground='black',
            text="Replace",
            command=lambda: self.find_or_replace(
                sub_win.excel_display_text,
                sub_win.find_btn_combobox,
                sub_win.replace_btn_combobox,
                'replace'
            )
        )
        sub_win.replace_btn.pack(side="left")

        # --- Excel展示区域 ---
        excel_frame = tk.Frame(left_frame, bg="lightblue")
        excel_frame.pack(fill=tk.BOTH, expand=True, pady=5)

        sub_win.excel_display_label = tk.Label(
            excel_frame,
            background='white',
            font=self.set_font_size('Song', 10, 'normal'),
            foreground='black',
            text="ExcelDataDisplay:"
        )
        sub_win.excel_display_label.pack(anchor='w')

        excel_display_scrollbar = tk.Scrollbar(excel_frame)
        excel_display_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        sub_win.excel_display_text = tk.Text(
            excel_frame,
            yscrollcommand=excel_display_scrollbar.set,
            background='lightblue',
            highlightbackground='white',  # tk.Text背景色
            font=self.set_font_size('Song', 12, 'normal'),
            foreground='black',
            width=45,
            height=10,
            bg="lightskyblue"
        )
        sub_win.excel_display_text.pack(fill=tk.BOTH, expand=True)
        self.add_function_buttons_in_request_text_for_tkText_pack(excel_frame, sub_win.excel_display_text, 1)
        excel_display_scrollbar.config(command=sub_win.excel_display_text.yview)


    def create_setLLM_sub_widgets(self, sub_win):
        sub_style = ttk.Style(sub_win)  # 为子窗口创建独立的样式Style
        sub_style.theme_use('clam')
        # 配置Combobox及其下拉菜单样式
        sub_style.configure('TCombobox', fieldbackground='white', background='white', foreground='black')
        sub_style.configure('TCombobox.Listbox', fieldbackground='white', background='white', foreground='black')

        # apiURL标签
        sub_win.apiURL_label = tk.Label(
            sub_win,
            background='white',
            font=self.set_font_size('Song', 12, 'normal'),
            foreground='black',  # 字体颜色
            text="apiURL:"
        )
        sub_win.apiURL_label.grid(column=0, row=0)
        # 选择apiURL
        sub_win.apiURL_combobox = ttk.Combobox(sub_win, state='NORMAL')
        sub_win.apiURL_combobox.grid(column=1, row=0, sticky='Ew')
        # 获取配置里的apiURL
        apiURLList = read_json_file(f'{current_script_path}/configini.json')['apiURL']
        sub_win.apiURL_combobox['values'] = apiURLList
        sub_win.apiURL_combobox.current(0)  # 设置默认值为列表中的第一个元素
        print(sub_win.apiURL_combobox.get())
        # 保存apiURL
        # JSON文件名为configin.json
        json_file_path = f'{current_script_path}/configini.json'
        sub_win.set_apiURL_button = tk.Button(
            sub_win,
            text="add",
            background="lightblue",
            highlightbackground="lightblue",
            command=lambda: self.changeKeyValueInJsonFile_forLLMSet(
                json_file_path,
                'apiURL',
                sub_win.apiURL_combobox.get(),
                sub_win.apiURL_combobox,
                0
            )
        )  # 组件按期定格式
        sub_win.set_apiURL_button.grid(column=2, row=0)
        self.load_llm_records(sub_win.apiURL_combobox,'apiURL')

        # apiKey标签
        sub_win.apiKey_label = tk.Label(
            sub_win,
            background='white',
            font=self.set_font_size('Song', 12, 'normal'),
            foreground='black',  # 字体颜色
            text="apiKey:"
        )
        sub_win.apiKey_label.grid(column=0, row=1)
        # 选择apiKey
        sub_win.apiKey_combobox = ttk.Combobox(sub_win, state='NORMAL')
        sub_win.apiKey_combobox.grid(column=1, row=1, sticky='EW')
        # 获取配置里的apiKey
        apiKeyList = read_json_file(f'{current_script_path}/configini.json')['apiKey']
        sub_win.apiKey_combobox['values'] = apiKeyList
        sub_win.apiKey_combobox.current(0)  # 设置默认值为列表中的第一个元素
        print(sub_win.apiKey_combobox.get())
        # 保存apiKey
        # JSON文件名为configin.json
        json_file_path = f'{current_script_path}/configini.json'
        sub_win.set_apiKey_button = tk.Button(
            sub_win,
            text="add",
            background="lightblue",
            highlightbackground="lightblue",
            command=lambda: self.changeKeyValueInJsonFile_forLLMSet(
                json_file_path,
                'apiKey',
                sub_win.apiKey_combobox.get(),
                sub_win.apiKey_combobox,
                0
            )
        )  # 组件按期定格式
        sub_win.set_apiKey_button.grid(column=2, row=1)
        self.load_llm_records(sub_win.apiKey_combobox, 'apiKey')

        # apiHeaders标签
        sub_win.apiHeaders_label = tk.Label(
            sub_win,
            background='white',
            font=self.set_font_size('Song', 12, 'normal'),
            foreground='black',  # 字体颜色
            text="apiHeaders:"
        )
        sub_win.apiHeaders_label.grid(column=0, row=2)
        # 选择apiHeaders
        sub_win.apiHeaders_combobox = ttk.Combobox(sub_win, state='NORMAL')
        sub_win.apiHeaders_combobox.grid(column=1, row=2, sticky='EW')
        # 获取配置里的apiHeaders
        apiHeadersList = read_json_file(f'{current_script_path}/configini.json')['apiHeaders']
        sub_win.apiHeaders_combobox['values'] = apiHeadersList
        sub_win.apiHeaders_combobox.current(0)  # 设置默认值为列表中的第一个元素
        print(f'列表中的第一个元素:{sub_win.apiHeaders_combobox.get()}')
        # 保存apiHeaders
        sub_win.set_apiHeaders_button = tk.Button(
            sub_win,
            text="add",
            background="lightblue",
            highlightbackground="lightblue",
            command=lambda: self.changeKeyValueInJsonFile_forLLMSet(
                json_file_path,
                'apiHeaders',
                sub_win.apiHeaders_combobox.get(),
                sub_win.apiHeaders_combobox,
                1
            )
        )  # 组件按钮绑定格式化函数功能
        sub_win.set_apiHeaders_button.grid(column=2, row=2)
        self.load_llm_records(sub_win.apiHeaders_combobox, 'apiHeaders')

        sub_win.set_update_button = tk.Button(
            sub_win,
            text="update",
            bg="lightskyblue",
            highlightbackground='lightblue',
            background='lightblue',
            foreground='black',
            font=self.set_font_size('Song', 12, 'normal'),
            command=lambda: self.updateLLMSet(
                {
                    'apiURL':sub_win.apiURL_combobox.get(),
                    'apiKey':sub_win.apiKey_combobox.get(),
                    'apiHeaders':json.loads(sub_win.apiHeaders_combobox.get().replace('\'','\"'))
                }
            )
        )  # 组件按钮绑定格式化函数功能
        sub_win.set_update_button.grid(column=0, row=3)

        sub_win.set_screen_button = tk.Button(
            sub_win,
            text="screen",
            bg="lightskyblue",
            highlightbackground='lightblue',
            background='lightblue',
            foreground='black',
            font=self.set_font_size('Song', 12, 'normal'),
            command=lambda: self.start_region_selection_only_see_img()
        )  # 组件按钮绑定格式化函数功能
        sub_win.set_screen_button.grid(column=1, row=3)

    def create_business_tools_sub_widgets(self, sub_win):
        """创建图片OCR识别界面布局（左图右文）- 支持图片自适应填充"""
        # --- 主框架---
        main_frame = tk.Frame(sub_win, background='lightblue', highlightbackground='white')
        main_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        sub_style = ttk.Style(sub_win)  # 为子窗口创建独立的Style
        sub_style.theme_use('clam')
        # 配置Combobox及其下拉菜单
        sub_style.configure('TCombobox', fieldbackground='white', foreground='black')
        sub_style.configure('TCombobox.Listbox', background='white', foreground='black')

        # ===== 左侧内容展示区域 =====
        left_img_frame = tk.Frame(
            main_frame,
            bd=2,
            relief=tk.SUNKEN,
            highlightbackground='white',
            background='lightblue'
        )
        left_img_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5, pady=5)
        scrollbar = tk.Scrollbar(left_img_frame)
        self.left_img_output_text = tk.Text(
            left_img_frame,
            wrap=tk.WORD,
            yscrollcommand=scrollbar.set,
            background='lightblue',
            highlightbackground='white',
            font=self.set_font_size('Song', 12, 'normal'),
            foreground='blue',  # 字体颜色（蓝色）
            padx=5,
            pady=5
        )
        # tk.Text内添加功能按钮
        self.add_function_buttons_in_request_text_for_tkText_pack(sub_win, self.left_img_output_text, 1)
        scrollbar.config(command=self.left_img_output_text.yview)

        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.left_img_output_text.pack(fill=tk.BOTH, expand=True)
        self.left_img_output_text.insert(tk.END, "")

        # ===== 右侧按钮区域 =====
        right_button_frame = tk.Frame(
            main_frame,
            bd=2,
            relief=tk.SUNKEN,
            highlightbackground='white',
            background='lightblue'
        )
        right_button_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=5, pady=5)

        sub_win.getHtmlData_button = tk.Button(
            right_button_frame,
            font=self.set_font_size('Song', 12, 'normal'),
            highlightbackground='lightblue',
            text="GetHtmlData",
            command=lambda: self.thread_it(
                self.getHtmlData(
                    sub_win.left_img_output_text,
                    sub_win.left_img_output_text.get("1.0", "end-1c")
                )
            )
        )
        sub_win.getHtmlData_button.pack(fill="x", pady=2)

        sub_win.getPythonCode_button = tk.Button(
            right_button_frame,
            font=self.set_font_size('Song', 12, 'normal'),
            highlightbackground='lightblue',
            text="GetPyCode",
            command=lambda: self.thread_it(
                self.restore_python_code_by_AI(
                    sub_win.left_img_output_text,
                    sub_win.left_img_output_text.get("1.0", "end-1c")
                )
            )
        )
        sub_win.getPythonCode_button.pack(fill="x", pady=2)

        # 图片格式转换按钮
        sub_win.change_picture_format_btn = tk.Button(
            right_button_frame,
            font=self.set_font_size('Song', 12, 'normal'),
            highlightbackground='lightblue',
            text="ChangePictureFormat",
            command=self.change_picture_format
        )
        sub_win.change_picture_format_btn.pack(fill="x", pady=5)

        sub_win.getTestData_button = tk.Button(
            right_button_frame,
            font=self.set_font_size('Song', 12, 'normal'),
            highlightbackground='lightblue',
            text="GetTestData",
            command=lambda: self.thread_it(
                self.getTestData(self.left_img_output_text, sub_win.encrypt_combobox)
            )
        )
        sub_win.getTestData_button.pack(fill="x", pady=2)



        # 加密相关控件（使用Frame嵌套实现水平排列）
        encrypt_frame = tk.Frame(
            right_button_frame,
            background='lightblue',
            highlightbackground='lightblue'
        )
        encrypt_frame.pack(fill="x", pady=2)

        sub_win.encrypt_btn = tk.Label(
            encrypt_frame,
            font=self.set_font_size('Song', 12, 'normal'),
            background='white',
            highlightbackground='white',
            foreground='black',
            text="EncryptChar"
        )
        sub_win.encrypt_btn.pack(side="left", padx=2,  expand=True, fill="both")

        sub_win.encrypt_combobox = ttk.Combobox(
            encrypt_frame,
            state='NORMAL'
        )
        sub_win.encrypt_combobox.pack(side="left", fill="x", expand=True)
        sub_win.encrypt_combobox['values'] = ['测', '测试']
        sub_win.encrypt_combobox.current(0)

        # 关键字复制控件
        key_frame = tk.Frame(
            right_button_frame,
            background='lightblue',
            highlightbackground='lightblue'
        )
        key_frame.pack(fill="x", pady=2)
        sub_win.select_key_btn1 = tk.Button(
            key_frame,
            background='lightblue',
            text='GetTxtKeyCopy',
            font=self.set_font_size('Song', 12, 'normal'),
            highlightbackground='lightblue',
            command=lambda: self.copy_value_of_key(sub_win, self.left_img_output_text, sub_win.select_key_combobox1.get())
        )
        sub_win.select_key_btn1.pack(side="left", padx=2,  expand=True, fill="both")
        sub_win.select_key_combobox1 = ttk.Combobox(
            key_frame,
            background='lightblue',
            state='NORMAL'
        )
        sub_win.select_key_combobox1.pack(side="left", fill="x")
        sub_win.select_key_combobox1['values'] = ['123']
        sub_win.select_key_combobox1.current(0)

        # XMind转Excel控件（嵌套Frame实现水平排列）
        xmind_frame = tk.Frame(
            right_button_frame,
            background='lightblue',
            highlightbackground='lightblue'
        )
        xmind_frame.pack(fill="x", pady=2)

        sub_win.choose_case_combobox = ttk.Combobox(
            xmind_frame,
            background='white',
            font=self.set_font_size('Song', 10, 'normal'),
            values=['普通用例', '版本用例', '回归用例'],
            state='readonly'
        )
        sub_win.choose_case_combobox.pack(side="left", fill="x", expand=True, padx=2)
        sub_win.choose_case_combobox.current(0)

        sub_win.WriteXmindToExcel_btn = tk.Button(
            xmind_frame,
            text="WriteXmindToExcel",
            font=self.set_font_size('Song', 12, 'normal'),
            background='lightblue',
            highlightbackground='lightblue',
            command=lambda: self.write_xmind_to_excel(sub_win.choose_case_combobox)
        )
        sub_win.WriteXmindToExcel_btn.pack(side="left", padx=2)


        # 查找控件（嵌套Frame实现水平排列）
        find_frame = tk.Frame(
            right_button_frame,
            background='lightblue',
            highlightbackground='lightblue'
        )
        find_frame.pack(fill="x", pady=2)

        sub_win.find_btn_combobox = ttk.Combobox(
            find_frame,
            font=self.set_font_size('Song', 10, 'normal'),
            background='white',
            values=['*', ',','、','\n',' ','\t'],
            state='NORMAL'
        )
        sub_win.find_btn_combobox.pack(side="left", padx=2)  # ,  expand=True, fill="both")
        sub_win.find_btn_combobox.current(0)

        sub_win.find_replace_btn = tk.Button(
            find_frame,
            font=self.set_font_size('Song', 12, 'normal'),
            background='white',
            highlightbackground='lightblue',
            foreground='black',
            text="Find",
            command=lambda: self.find_or_replace(
                self.left_img_output_text,
                sub_win.find_btn_combobox,
                sub_win.replace_btn_combobox,
                'find'
            )
        )
        sub_win.find_replace_btn.pack(side="left", padx=2,  expand=True, fill="both")



        # 替换控件（嵌套Frame实现水平排列）
        replace_frame = tk.Frame(
            right_button_frame,
            background='lightblue',
            highlightbackground='lightblue'
        )
        replace_frame.pack(fill="x", pady=2)

        sub_win.replace_btn_combobox = ttk.Combobox(
            replace_frame,
            font=self.set_font_size('Song', 10, 'normal'),
            background='white',
            values=['*', ',','、','\n',' ','\t'],
            state='NORMAL'
        )
        sub_win.replace_btn_combobox.pack(side="left", padx=2)  # ,  expand=True, fill="both")
        sub_win.replace_btn_combobox.current(0)

        sub_win.replace_btn = tk.Button(
            replace_frame,
            font=self.set_font_size('Song', 12, 'normal'),
            background='white',
            highlightbackground='lightblue',
            foreground='black',
            text="Replace",
            command=lambda: self.find_or_replace(
                self.left_img_output_text,
                sub_win.find_btn_combobox,
                sub_win.replace_btn_combobox,
                'replace'
            )
        )
        sub_win.replace_btn.pack(side="left", padx=2,  expand=True, fill="both")

    def search_keyword_in_text(self, keyword, which_text):
        # 检查输入有效性
        if not keyword.strip() or not which_text.get("1.0", "end-1c"):
            return

        # 初始化存储变量（首次调用时）
        if not hasattr(self, "_search_state"):
            self._search_state = {
                "widget": None,  # 当前搜索的文本框对象
                "positions": [],  # 所有匹配位置
                "index": -1,  # 当前跳转位置索引
                "last_keyword": ""  # 上一次搜索的关键词（用于缓存判断）
            }

        # 如果切换了文本框或关键词，重置搜索状态
        if (self._search_state["widget"] != which_text or
                self._search_state["last_keyword"] != keyword):
            self._search_state = {
                "widget": which_text,
                "positions": [],
                "index": -1,
                "last_keyword": keyword
            }
            # 清空旧高亮并重新搜索
            which_text.tag_remove("highlight", "1.0", "end")
            self._find_all_matches(keyword, which_text)

        # 无匹配项时返回
        if not self._search_state["positions"]:
            return

            # 循环跳转逻辑
        self._search_state["index"] = (self._search_state["index"] + 1) % len(self._search_state["positions"])
        next_pos = self._search_state["positions"][self._search_state["index"]]

        # 执行跳转并高亮当前项
        which_text.see(next_pos)
        which_text.tag_remove("current_highlight", "1.0", "end")
        which_text.tag_add("current_highlight", next_pos, f"{next_pos}+{len(keyword)}c")
        which_text.tag_config("current_highlight", background="orange")  # 当前项特殊高亮

    def has_photoimage(self,which_text):
        content = which_text.dump("1.0", "end", tag=True, text=True, image=True)
        for index in range(len(content)-1,-1,-1):
            item=content[index]
            print(item) # ('image', 'pyimage2', '1.0'), ('text', '\n', '1.1')
            if item[0] == "image" and "pyimage" in item[1]:#:isinstance(item[1], ImageTk.PhotoImage):
                # 生成带时间戳的文件名
                timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                default_filename = f"图片对象保存_{timestamp}.png"
                # 保存截图
                img_path = f'{current_script_path}/{screenshotFileName}/{default_filename}'
                self.screenshot.save(img_path)
                image_references[img_path] = item[1]  # 保存引用
                return True
        return False

    def _find_all_matches(self, keyword, widget):
        """辅助方法：查找所有匹配位置并高亮"""
        start = "1.0"
        while True:
            pos = widget.search(keyword, start, stopindex="end", nocase=True)
            if not pos:
                break
            end = f"{pos}+{len(keyword)}c"
            widget.tag_add("highlight", pos, end)
            self._search_state["positions"].append(pos)
            start = end
            # 配置普通高亮样式
        widget.tag_config("highlight", foreground="yellow",  background="blue",font=("Song", 12, "normal"))

    def find_or_replace(self,which_text,find_combo,replace_combo,pattern):
        if pattern=='find':
            self.search_keyword_in_text(find_combo.get(), which_text)
        elif pattern=='replace':
            textContent=which_text.get('1.0','end')
            old_str=find_combo.get()
            new_str=replace_combo.get()
            splitList=textContent.split(old_str)
            self.clearContent(which_text)
            which_text.insert('end',new_str.join(splitList))
        else:
            pass

    def create_business_tools_sub_widgets1(self, sub_win):
        # 主容器分为左右两部分：左侧文本区域 + 右侧按钮区域
        left_frame = tk.Frame(sub_win,highlightbackground='white')
        right_frame = tk.Frame(sub_win,highlightbackground='white')
        left_frame.pack(side="left", fill="both", expand=True, padx=5, pady=5)
        right_frame.pack(side="right", fill="y", padx=5, pady=5)

        # 左侧文本区域（ScrolledText）
        #sub_win.output_text = scrolledtext.ScrolledText(
        sub_win.output_text = tk.Text(
            left_frame,
            width=30,
            height=20,
            bg="lightskyblue",
            highlightbackground='white',
            background='lightblue',
            foreground='black',
            font=self.set_font_size('Song',12,'normal')
        )
        sub_win.output_text.pack(fill="both", expand=True)
        self.add_function_buttons_in_request_text_for_tkText_pack(sub_win, sub_win.output_text, 1)

        # 右侧按钮区域（垂直排列）
        sub_win.getHtmlData_button = tk.Button(
            right_frame,
            font=self.set_font_size('Song', 12, 'normal'),
            highlightbackground='lightblue',
            text="GetHtmlData",
            command=lambda: self.thread_it(
                self.getHtmlData(sub_win.output_text, sub_win.output_text.get("1.0", "end-1c"))
            )
        )
        sub_win.getHtmlData_button.pack(fill="x", pady=2)

        sub_win.getPythonCode_button = tk.Button(
            right_frame,
            font=self.set_font_size('Song', 12, 'normal'),
            highlightbackground='lightblue',
            text="GetPyCode",
            command=lambda: self.thread_it(
                self.restore_python_code_by_AI(sub_win.output_text, sub_win.output_text.get("1.0", "end-1c"))
            )
        )
        sub_win.getPythonCode_button.pack(fill="x", pady=2)

        sub_win.getTestData_button = tk.Button(
            right_frame,
            font=self.set_font_size('Song', 12, 'normal'),
            highlightbackground='lightblue',
            text="GetTestData",
            command=lambda: self.thread_it(
                self.getTestData(sub_win.output_text, sub_win.encrypt_combobox)
            )
        )
        sub_win.getTestData_button.pack(fill="x", pady=2)

        # 加密相关控件（使用Frame嵌套实现水平排列）
        encrypt_frame = tk.Frame(right_frame,highlightbackground='lightblue')
        encrypt_frame.pack(fill="x", pady=2)
        sub_win.encrypt_btn = tk.Label(
            encrypt_frame,
            font=self.set_font_size('Song', 12, 'normal'),
            background='white',
            foreground='black',
            text="EncryptChar"
        )
        sub_win.encrypt_btn.pack(side="left", padx=2)
        sub_win.encrypt_combobox = ttk.Combobox(encrypt_frame, state='NORMAL')
        sub_win.encrypt_combobox.pack(side="left", fill="x", expand=True)
        sub_win.encrypt_combobox['values'] = ['测', '测试']
        sub_win.encrypt_combobox.current(0)

        # 关键字复制控件
        key_frame = tk.Frame(right_frame,highlightbackground='lightblue')
        key_frame.pack(fill="x", pady=2)
        sub_win.select_key_btn1 = tk.Button(
            key_frame,
            text='GetTxtKeyCopy',
            font=self.set_font_size('Song', 12, 'normal'),
            highlightbackground='lightblue',
            command=lambda: self.copy_value_of_key(sub_win, sub_win.output_text, sub_win.select_key_combobox1.get())
        )
        sub_win.select_key_btn1.pack(side="left", padx=2)
        sub_win.select_key_combobox1 = ttk.Combobox(key_frame, state='NORMAL')
        sub_win.select_key_combobox1.pack(side="left", fill="x", expand=True)
        sub_win.select_key_combobox1['values'] = ['123']
        sub_win.select_key_combobox1.current(0)

        # 图片格式转换按钮
        sub_win.change_picture_format_btn = tk.Button(
            right_frame,
            font=self.set_font_size('Song', 12, 'normal'),
            text="ChangePictureFormat",
            command=self.change_picture_format
        )
        sub_win.change_picture_format_btn.pack(fill="x", pady=5)

        # XMind转Excel控件（嵌套Frame实现水平排列）
        xmind_frame = tk.Frame(right_frame,highlightbackground='lightblue')
        xmind_frame.pack(fill="x", pady=2)
        sub_win.choose_case_combobox = ttk.Combobox(
            xmind_frame,
            values=['普通用例', '版本用例', '回归用例'],
            state='readonly'
        )
        sub_win.choose_case_combobox.pack(side="left", fill="x", expand=True, padx=2)
        sub_win.choose_case_combobox.current(0)
        sub_win.WriteXmindToExcel_btn = tk.Button(
            xmind_frame,
            text="WriteXmindToExcel",
            font=self.set_font_size('Song', 12, 'normal'),
            highlightbackground='lightblue',
            command=lambda: self.write_xmind_to_excel(sub_win.choose_case_combobox)
        )
        sub_win.WriteXmindToExcel_btn.pack(side="left", padx=2)

    def create_modify_img_sub_widgets(self, sub_win):
        sub_style = ttk.Style(sub_win)  # 为子窗口创建独立的样式Style
        sub_style.theme_use('clam')
        # 配置Combobox及其下拉菜单样式
        sub_style.configure('TCombobox', fieldbackground='white', background='white', foreground='black')
        sub_style.configure('TCombobox.Listbox', fieldbackground='white', background='white', foreground='black')

        """创建图片OCR识别界面布局（左图右文）- 支持图片自适应填充"""
        # --- 主框架---
        main_frame = tk.Frame(sub_win,background='lightblue',highlightbackground='white')
        main_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        sub_style = ttk.Style(sub_win)  # 为子窗口创建独立的样式Style
        sub_style.theme_use('clam')
        # 配置Combobox及其下拉菜单样式
        sub_style.configure('TCombobox', fieldbackground='white', background='white',foreground='black')
        sub_style.configure('TCombobox.Listbox', fieldbackground='white', background='white', foreground='black')

        # ===== 左侧图片展示区域 =====
        left_img_frame = tk.Frame(
            main_frame,
            bd=2,
            relief=tk.SUNKEN,
            highlightbackground='white',
            background='lightblue'
        )
        left_img_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5, pady=5)
        scrollbar = tk.Scrollbar(left_img_frame)
        self.left_img_output_text = tk.Text(
            left_img_frame,
            wrap=tk.WORD,
            yscrollcommand=scrollbar.set,
            background='lightblue',
            highlightbackground='white',
            font=self.set_font_size('Song', 12, 'normal'),
            foreground='blue',# 字体颜色（蓝色）
            padx=5,
            pady=5
        )
        # tk.Text内添加功能按钮
        self.add_function_buttons_in_img_text(sub_win, self.left_img_output_text, 1)
        scrollbar.config(command=self.left_img_output_text.yview)

        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.left_img_output_text.pack(fill=tk.BOTH, expand=True)
        self.left_img_output_text.insert(tk.END, "")

        # ===== 右侧OCR结果区域 =====
        right_text_frame = tk.Frame(main_frame, bd=2, relief=tk.SUNKEN,highlightbackground='white',background='lightblue')
        right_text_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=5, pady=5)

        scrollbar1 = tk.Scrollbar(right_text_frame)
        self.text_output = tk.Text(
            right_text_frame,
            wrap=tk.WORD,
            yscrollcommand=scrollbar1.set,
            background='lightblue',
            highlightbackground='white',
            font=self.set_font_size('Song', 12, 'normal'),
            foreground='blue',# 字体颜色（蓝色）
            padx=5,
            pady=5
        )
        # tk.Text内添加功能按钮
        self.add_function_buttons_in_img_text(sub_win, self.text_output, 2)
        scrollbar1.config(command=self.text_output.yview)

        scrollbar1.pack(side=tk.RIGHT, fill=tk.Y)
        self.text_output.pack(fill=tk.BOTH, expand=True)
        self.text_output.insert(tk.END, "")

        # --- 第二行主框架 ---
        main_frame1 = tk.Frame(sub_win, background='lightblue')
        main_frame1.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        # ===== 底部按钮区域 =====
        btn_frame = tk.Frame(main_frame1, background='lightblue')
        btn_frame.pack(fill=tk.X, padx=5, pady=5)

        self.see_lang_label_btn1=tk.Label(
            btn_frame,
            background='white',
            font=self.set_font_size('Song', 12, 'normal'),
            foreground='black',  # 字体颜色
            text="识别语言:"
        )
        self.see_lang_label_btn1.pack(side=tk.LEFT, padx=(0, 5))

        # 创建样式对象
        style = ttk.Style()
        # 配置Combobox的样式
        # 注意：不同平台可能对样式的支持有所差异
        style.configure(
            'TCombobox',
            fieldbackground='lightblue',  # 输入框的背景色
            background='lightgreen' # 下拉列表的背景色
        )
        self.lang_combo_see = ttk.Combobox(
            btn_frame,
            font=self.set_font_size('Song', 12, 'normal'),
            foreground='black',  # 字体颜色
            values=["中文", "英文", "日文", "韩文", "阿拉伯文"],
            style = 'TCombobox'
        )

        self.lang_combo_see.pack(side=tk.LEFT, padx=(0, 10))
        self.lang_combo_see.current(0)

        self.see_content_label_btn1=tk.Label(
            btn_frame,
            background='white',
            font=self.set_font_size('Song', 12, 'normal'),
            foreground='black',  # 字体颜色
            text="识别内容:"
        )
        self.see_content_label_btn1.pack(side=tk.LEFT, padx=(0, 5))
        self.lang_combo_pattern = ttk.Combobox(btn_frame, values=["文本", "定位"])
        self.lang_combo_pattern.pack(side=tk.LEFT, padx=(0, 10))
        self.lang_combo_pattern.current(0)

        self.copy_app_path_button = tk.Button(
            btn_frame,
            text="AppPathCopy",
            highlightbackground='lightblue',  # tk.Button按钮背景色
            command=self.app_path_copy
        )
        self.copy_app_path_button.pack(side=tk.LEFT, padx=(0, 10))

    def create_modify_img_sub_widgets1(self, sub_win):
        # 准备一些图片路径作为示例
        image_paths = [
            "test"
        ]

        # 创建一个 Combobox 组件用于选择图片路径
        combobox = ttk.Combobox(sub_win, values=image_paths, state="readonly")
        combobox.grid(column=2, row=8)
        combobox.current(0)  # 默认选择第一项

        # 创建一个 Label 组件用于显示图片
        image_label = tk.Label(sub_win, width=1200, height=800)
        image_label.grid(column=0, row=0, columnspan=2, rowspan=8, sticky='NSEW')

        # 测试选择标签
        sub_win.test_label = ttk.Button(sub_win, text="test:")
        sub_win.test_label.grid(column=2, row=0)
        sub_win.test_combobox = ttk.Combobox(sub_win, state="NORMAL")
        sub_win.test_combobox.grid(column=2, row=1, sticky='EW')
        sub_win.test_combobox['values'] = ['test', '123']
        sub_win.test_combobox.current(0)

        # 用于存储当前显示的图片的 PhotoImage 对象
        current_photo = None

        # 加载图片并调整大小以适应 Label
        def load_and_resize_image(path, label):
            try:
                # 使用 PIL 库加载图片
                image = Image.open(path)
                # 调整图片大小以适应 Label
                image = image.resize((label.winfo_width(), label.winfo_height()), Image.ANTIALIAS)
                # 创建 PhotoImage 对象
                photo = ImageTk.PhotoImage(image)
                # 更新 Label 显示图片
                label.configure(image=photo)
                global current_photo
                current_photo = photo
            except IOError:
                print(f"无法加载图片：{path}")

        # 根据下拉框的选择更新图片
        def update_image(event):
            # 获取下拉框的值，即图片路径
            selected_path = f'{current_script_path}/{fileName}/{combobox.get()}.png'

            new_selected_path = selected_path

            # 如果路径有效，则加载并显示图片
            load_and_resize_image(new_selected_path, image_label)

        # 绑定下拉框的事件
        combobox.bind("<<ComboboxSelected>>", update_image)

    def resize_image_to_fit(self, image_path, max_width, max_height):
        """将图片等比缩放至适应Label尺寸"""
        img = Image.open(image_path)
        img.thumbnail((max_width, max_height), Image.LANCZOS)  # 保持比例缩放
        return ImageTk.PhotoImage(img)

    def dada(self,which_text,text_content,color_style,family,size,weight):
        # which_text.tag_configure("red_text", foreground="red", font=("Times New Roman", 14))
        # which_text.insert("end", "这是红色文本", "red_text")
        font=self.set_font_size(family, size, weight)
        which_text.tag_configure("your_text", foreground=color_style, font=font)
        which_text.insert("end", text_content, "your_text")

    def create_ocr_text_sub_widgets(self, sub_win):
        sub_style = ttk.Style(sub_win)  # 为子窗口创建独立的样式Style
        sub_style.theme_use('clam')
        # 配置Combobox及其下拉菜单样式
        sub_style.configure('TCombobox', fieldbackground='white', background='white', foreground='black')
        sub_style.configure('TCombobox.Listbox', fieldbackground='white', background='white', foreground='black')

        """创建图片OCR识别界面布局（左图右文）- 支持图片自适应填充"""
        # --- 主框架---
        main_frame = tk.Frame(sub_win,background='lightblue',highlightbackground='white')
        main_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        sub_style = ttk.Style(sub_win)  # 为子窗口创建独立的样式Style
        sub_style.theme_use('clam')
        # 配置Combobox及其下拉菜单样式
        sub_style.configure('TCombobox', fieldbackground='white', background='white',foreground='black')
        sub_style.configure('TCombobox.Listbox', fieldbackground='white', background='white', foreground='black')

        # ===== 左侧图片展示区域 =====
        left_img_frame = tk.Frame(
            main_frame,
            bd=2,
            relief=tk.SUNKEN,
            highlightbackground='white',
            background='lightblue'
        )
        left_img_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5, pady=5)
        scrollbar = tk.Scrollbar(left_img_frame)
        self.left_img_output_text = tk.Text(
            left_img_frame,
            wrap=tk.WORD,
            yscrollcommand=scrollbar.set,
            background='lightblue',
            highlightbackground='white',
            font=self.set_font_size('Song', 12, 'normal'),
            foreground='blue',# 字体颜色（蓝色）
            padx=5,
            pady=5
        )
        # tk.Text内添加功能按钮
        self.add_function_buttons_in_img_text(sub_win, self.left_img_output_text, 1)
        scrollbar.config(command=self.left_img_output_text.yview)

        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.left_img_output_text.pack(fill=tk.BOTH, expand=True)
        self.left_img_output_text.insert(tk.END, "")

        # ===== 右侧OCR结果区域 =====
        right_text_frame = tk.Frame(main_frame, bd=2, relief=tk.SUNKEN,highlightbackground='white',background='lightblue')
        right_text_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=5, pady=5)

        scrollbar1 = tk.Scrollbar(right_text_frame)
        self.text_output = tk.Text(
            right_text_frame,
            wrap=tk.WORD,
            yscrollcommand=scrollbar1.set,
            background='lightblue',
            highlightbackground='white',
            font=self.set_font_size('Song', 12, 'normal'),
            foreground='blue',# 字体颜色（蓝色）
            padx=5,
            pady=5
        )
        # tk.Text内添加功能按钮
        self.add_function_buttons_in_img_text(sub_win, self.text_output, 2)
        scrollbar1.config(command=self.text_output.yview)

        scrollbar1.pack(side=tk.RIGHT, fill=tk.Y)
        self.text_output.pack(fill=tk.BOTH, expand=True)
        self.text_output.insert(tk.END, "")

        # --- 第二行主框架 ---
        main_frame1 = tk.Frame(sub_win, background='lightblue')
        main_frame1.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        # ===== 底部按钮区域 =====
        btn_frame = tk.Frame(main_frame1, background='lightblue')
        btn_frame.pack(fill=tk.X, padx=5, pady=5)

        self.see_lang_label_btn1=tk.Label(
            btn_frame,
            background='white',
            font=self.set_font_size('Song', 12, 'normal'),
            foreground='black',  # 字体颜色
            text="识别语言:"
        )
        self.see_lang_label_btn1.pack(side=tk.LEFT, padx=(0, 5))

        # 创建样式对象
        style = ttk.Style()
        # 配置Combobox的样式
        # 注意：不同平台可能对样式的支持有所差异
        style.configure(
            'TCombobox',
            fieldbackground='lightblue',  # 输入框的背景色
            background='lightgreen' # 下拉列表的背景色
        )
        self.lang_combo_see = ttk.Combobox(
            btn_frame,
            font=self.set_font_size('Song', 12, 'normal'),
            foreground='black',  # 字体颜色
            values=["中文", "英文", "日文", "韩文", "阿拉伯文"],
            style = 'TCombobox'
        )

        self.lang_combo_see.pack(side=tk.LEFT, padx=(0, 10))
        self.lang_combo_see.current(0)

        self.see_content_label_btn1=tk.Label(
            btn_frame,
            background='white',
            font=self.set_font_size('Song', 12, 'normal'),
            foreground='black',  # 字体颜色
            text="识别内容:"
        )
        self.see_content_label_btn1.pack(side=tk.LEFT, padx=(0, 5))
        self.lang_combo_pattern = ttk.Combobox(btn_frame, values=["文本", "定位"])
        self.lang_combo_pattern.pack(side=tk.LEFT, padx=(0, 10))
        self.lang_combo_pattern.current(0)

        self.copy_app_path_button = tk.Button(
            btn_frame,
            text="AppPathCopy",
            highlightbackground='lightblue',  # tk.Button按钮背景色
            command=self.app_path_copy
        )
        self.copy_app_path_button.pack(side=tk.LEFT, padx=(0, 10))

    def create_ocr_qa_sub_widgets(self, sub_win):
        sub_style = ttk.Style(sub_win)  # 为子窗口创建独立的样式Style
        sub_style.theme_use('clam')
        # 配置Combobox及其下拉菜单样式
        sub_style.configure('TCombobox', fieldbackground='white', background='white', foreground='black')
        sub_style.configure('TCombobox.Listbox', fieldbackground='white', background='white', foreground='black')

        """创建图片OCR识别界面布局（左图右文）- 支持图片自适应填充"""
        # --- 第一行主框架 ---
        main_frame = tk.Frame(sub_win,background='lightblue',highlightbackground='white')
        main_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # ===== 左侧输入问题区域 =====
        left_text_frame = tk.Frame(main_frame,background='lightblue',highlightbackground='white')
        left_text_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5, pady=5)

        scrollbar = tk.Scrollbar(left_text_frame)
        # self.left_text_output = scrolledtext.ScrolledText(left_text_frame, wrap=tk.WORD, padx=5, pady=5, bg="lightskyblue")
        self.left_output_text = tk.Text(
            left_text_frame,
            wrap=tk.WORD,
            yscrollcommand=scrollbar.set,
            font=self.set_font_size('Song', 14, 'normal'),
            foreground='blue',
            bg='lightskyblue',
            highlightbackground='white',
            padx=5,
            pady=5
        )
        # tk.Text内添加功能按钮
        self.add_function_buttons_in_qa_text(sub_win, self.left_output_text, 1)
        scrollbar.config(command=self.left_output_text.yview)

        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.left_output_text.pack(fill=tk.BOTH, expand=True)
        self.left_output_text.insert(tk.END, " ") #请输入问题.......

        # ===== 右侧RIGHT输出结果区域 =====
        right_text_frame = tk.Frame(main_frame,background='lightblue',highlightbackground='white')
        right_text_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=5, pady=5)

        scrollbar = tk.Scrollbar(right_text_frame,background='lightblue')
        self.right_output_text = tk.Text(
            right_text_frame,
            wrap=tk.WORD,
            yscrollcommand=scrollbar.set,
            font=self.set_font_size('Song', 14, 'normal'),
            foreground='blue',
            bg='lightskyblue',
            highlightbackground='white',
            padx=5,
            pady=5
        )
        # tk.Text内添加功能按钮
        self.add_function_buttons_in_qa_text(sub_win, self.right_output_text, 2)
        scrollbar.config(command=self.right_output_text.yview)

        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.right_output_text.pack(fill=tk.BOTH, expand=True)
        self.right_output_text.insert(tk.END, "")

        # --- 第二行主框架 ---
        main_frame1 = tk.Frame(sub_win,background = 'lightblue',highlightbackground='white')
        main_frame1.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # ===== 第二行左侧功能区域 =====
        left_frame = tk.Frame(main_frame1,background = 'lightblue',highlightbackground='white')
        left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # --- 按钮 ---
        btn_group1 = tk.Frame(left_frame, bd=1, relief=tk.RAISED,background='lightblue',highlightbackground='white')
        btn_group1.pack(fill=tk.X, pady=5)

        self.select_qa_record_button=tk.Button(
            btn_group1,
            text="Select QARecords:",
            font=self.set_font_size('Song', 10, 'normal'),
            background='lightblue',
            highlightbackground='lightblue',
            command=lambda: self.find_file_to_fill_QA_record(self.qa_record_combo, self.left_output_text, self.right_output_text)
        )
        self.select_qa_record_button.pack(side=tk.LEFT, padx=5)

        self.qa_record_combo = ttk.Combobox(
            btn_group1,
            state='readonly',
            values=list_qa_records(),
            font=self.set_font_size('Song', 10, 'normal'),
            background = 'lightblue',
            width=30
        )
        self.qa_record_combo.pack(side=tk.LEFT, padx=5)
        ##绑定回调函数不传递额外参数##
        self.qa_record_combo.bind('<<ComboboxSelected>>', self.fill_QA_record)
        self.load_qa_records(self.qa_record_combo)
        self.qa_record_combo.set(list_qa_records()[0])

        self.delete_qa_record_button = tk.Button(
            btn_group1,
            text="Delete QARecords",
            font=self.set_font_size('Song', 10, 'normal'),
            background='lightblue',
            highlightbackground='lightblue',
            command=lambda: self.thread_it(self.delete_QA_records(self.qa_record_combo))
        )
        self.delete_qa_record_button.pack(side=tk.LEFT, padx=5)

        # 截屏翻译英文
        only_screen_trans_btn = tk.Button(
            btn_group1,
            font=self.set_font_size('Song', 10, 'normal'),
            text="ScreenTrans",
            highlightbackground='lightblue',  # tk.Text背景色
            command=lambda: self.start_region_selection_see_English_img_and_translate()
        )
        only_screen_trans_btn.pack(side=tk.LEFT, padx=5)

        # 搜索按钮
        trans_English_btn = tk.Button(
            btn_group1,
            text="Translate",
            highlightbackground='lightblue',  # tk.Button按钮背景色
            font=self.set_font_size('Song', 10, 'normal'),
            anchor="center",
            command=lambda: self.translateEnglishByLLM(left_English_entry.get())
        )
        trans_English_btn.pack(side=tk.LEFT, padx=5)

        # 搜索which_text关键字结果
        left_English_entry = ttk.Entry(
            btn_group1,
            font=self.set_font_size('Song', 10, 'normal'),
            state='NORMAL'
        )
        left_English_entry.pack(side=tk.LEFT, padx=5)

        # 搜索按钮
        trans_English_paste_btn = tk.Button(
            btn_group1,
            text="Paste",
            highlightbackground='lightblue',  # tk.Button按钮背景色
            font=self.set_font_size('Song', 10, 'normal'),
            anchor="center",
            command=lambda: self.paste1(sub_win,left_English_entry)
        )
        trans_English_paste_btn.pack(side=tk.LEFT, padx=5)

        # 搜索按钮
        trans_English_clear_btn = tk.Button(
            btn_group1,
            text="Clear",
            highlightbackground='lightblue',  # tk.Button按钮背景色
            font=self.set_font_size('Song', 10, 'normal'),
            anchor="center",
            command=lambda: self.clearContent1(left_English_entry)
        )
        trans_English_clear_btn.pack(side=tk.LEFT, padx=5)

        # --- 第三行主框架 ---
        main_frame2 = tk.Frame(sub_win, background='lightblue',highlightbackground='white')
        main_frame2.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # ===== 底部按钮第一行 =====
        btn_frame = tk.Frame(main_frame2,background='lightblue',highlightbackground='white')
        btn_frame.pack(fill=tk.X, padx=5, pady=5)

        self.see_lang_label_btn=tk.Label(
            btn_frame,
            background='white',
            font=self.set_font_size('Song', 10, 'normal'),
            foreground='black',  # 字体颜色
            highlightbackground='lightblue',
            text="识别语言:",
        )
        self.see_lang_label_btn.pack(side=tk.LEFT, padx=(0, 5))
        self.choice_lang_combo = ttk.Combobox(
            btn_frame,
            values=["中文", "英文", "日文", "韩文", "阿拉伯文"],
            background='lightblue',
            font=self.set_font_size('Song', 8, 'normal')
        )
        self.choice_lang_combo.pack(side=tk.LEFT, padx=(0, 10)) #, expand=True, fill="both")
        self.choice_lang_combo.current(0)

        # ===== 底部按钮第二行 =====
        btn_frame1 = tk.Frame(main_frame2, background='lightblue', highlightbackground='white')
        btn_frame1.pack(fill=tk.X, padx=5, pady=5)

        self.common_question_label_btn=tk.Label(
            btn_frame1,
            background='white',
            font=self.set_font_size('Song', 10, 'normal'),
            foreground='black',  # 字体颜色
            highlightbackground='lightblue',
            text="常用提问话术:"
        )
        self.common_question_label_btn.pack(side=tk.LEFT, padx=(0, 5))
        conversationsList=find_value_of_key_in_nested_dict(
            (read_json_file(f'{current_script_path}/configini.json')),
            "conversation"
        )

        self.conversation_combo = ttk.Combobox(
            btn_frame1,
            values=conversationsList,
            background='lightblue',
            font=self.set_font_size('Song', 8, 'normal'),
            width=150 #通过 width 参数直接设置 Combobox 的显示宽度。
        )
        self.conversation_combo.pack(side=tk.LEFT, padx=(0, 10)) #, expand=True, fill="both")
        self.conversation_combo.current(0)

        self.insert_conversation_btn= tk.Button(
            btn_frame1,
            text="插入话术",
            background = 'lightblue',
            font=self.set_font_size('Song', 10, 'normal'),
            foreground='black',  # 字体颜色
            highlightbackground='lightblue',
            command=lambda: self.thread_it(self.insert_content_in_text(self.left_output_text, self.conversation_combo.get()))
        )
        self.insert_conversation_btn.pack(side=tk.LEFT)

        self.copy_appPath_button = tk.Button(
            btn_frame1,
            text="AppPathCopy",
            font=self.set_font_size('Song', 10, 'normal'),
            highlightbackground='lightblue',  # tk.Button按钮背景色
            command=self.app_path_copy
        )
        self.copy_appPath_button.pack(side=tk.LEFT)

    def create_screenshot_sub_widgets(self, sub_win):
        sub_win.screenshot_btn = tk.Button(
            sub_win,
            font=self.set_font_size('Song', 10, 'normal'),
            text="ScreenSub",
            background='lightblue',
            command=lambda: self.start_region_selection_only_see_img()
        )  # 带参数
        sub_win.screenshot_btn.grid(column=0, row=0)

        # 截屏翻译英文
        sub_win.screen_trans_btn = tk.Button(
            sub_win,
            font=self.set_font_size('Song', 10, 'normal'),
            text="ScreenTrans",
            highlightbackground='lightblue',  # tk.Text背景色
            command=lambda: self.start_region_selection_see_English_img_and_translate()
        )
        sub_win.screen_trans_btn.grid(column=0, row=1)

    def create_menu(self, which_window):
        """
        通用菜单工厂函数
        :param root: 父窗口对象
        :param menu_config: 字典格式的菜单配置
        :return: 构建好的菜单对象
        """
        menu_config = {
            "文件": [
                {"label": "新建", "command": lambda: print("新建")},
                {"label": "打开", "command": lambda: print("打开")},
                "---",
                {"label": "退出", "command": which_window.quit}
            ],
            "编辑": [
                {"label": "复制", "accelerator": "Ctrl+C"},
                {"label": "粘贴", "state": "disabled"}  # 禁用状态示例
            ]
        }
        menubar = tk.Menu(which_window)
        for menu_label, items in menu_config.items():
            submenu = tk.Menu(menubar, tearoff=0)
            menubar.add_cascade(label=menu_label, menu=submenu)
            for item in items:
                if item == "---":  # 分隔线
                    submenu.add_separator()
                else:
                    submenu.add_command(**item)
        which_window.config(menu=menubar)
        return menubar

    def select_image_for_ocr_tk(self):
        """Tkinter版本的选择图片方法"""
        img_path = filedialog.askopenfilename(
            title="选择图片",
            filetypes=[("图片文件", "*.png *.jpg *.jpeg *.bmp")]
        )

        if os.path.exists(img_path):
            print(f'img_path is:{img_path}')
            try:
                # 调用OCR识别  param lang must in dict_keys(['ch', 'en', 'korean', 'japan', 'chinese_cht', 'ta', 'te', 'ka', 'latin', 'arabic', 'cyrillic', 'devanagari']), but got chi_sim
                lang_map = {"中文": "ch", "英文": "en",
                            "日文": "japan", "韩文": "korean", "阿拉伯文": "arabic"}
                selected_lang = self.lang_combo_see.get()
                lang = lang_map[selected_lang]
                pattern = self.lang_combo_pattern.get()

                # 1. 打开图片并获取控件宽度
                image = Image.open(img_path)
                photo = ImageTk.PhotoImage(image)
                image_references[img_path] = photo  # 保存引用
                widget_width = self.left_img_output_text.winfo_width() - 20  # 预留边距

                # 2. 计算缩放比例（保持宽高比）
                if widget_width < 1:  # 避免初始宽度为0
                    widget_width = 300  # 默认宽度
                ratio = widget_width / image.width
                new_height = int(image.height * ratio)

                # 3. 缩放图片
                resized_image = image.resize((widget_width, new_height), Image.LANCZOS)
                photo = ImageTk.PhotoImage(resized_image)

                # 4. 插入文字和图片并保持引用
                self.left_img_output_text.insert("end", "")  # 插入文字
                self.left_img_output_text.image_create("end", image=photo)  # 在末尾插入图片
                self.left_img_output_text.insert("end", "\n")  # 继续插入文字
                self.left_img_output_text.image = photo  # 防止被垃圾回收

                # 调用OCR方法
                self.seeImgToTxtByPaddleOcr(self.text_output, img_path, lang, pattern)

            except Exception as e:
                self.left_img_output_text.insert("end", f"图片处理失败：{e}")  # 插入异常描述
                #messagebox.showerror(" 错误", f"图片处理失败: {str(e)}")

    def select_image_for_ocr_qa(self):
        """Tkinter版本的选择图片方法"""
        img_path = filedialog.askopenfilename(
            title="选择图片",
            filetypes=[("图片文件", "*.png *.jpg *.jpeg *.bmp")]
        )

        if img_path:
            try:
                # 调用OCR识别  param lang must in dict_keys(['ch', 'en', 'korean', 'japan', 'chinese_cht', 'ta', 'te', 'ka', 'latin', 'arabic', 'cyrillic', 'devanagari']), but got chi_sim
                lang_map = {"中文": "ch", "英文": "en",
                            "日文": "japan", "韩文": "korean", "阿拉伯文": "arabic"}
                selected_lang = self.choice_lang_combo.get()
                lang = lang_map[selected_lang]

                # 1. 打开图片并获取控件宽度
                image = Image.open(img_path)
                photo = ImageTk.PhotoImage(image)
                image_references[img_path] = photo  # 保存引用
                widget_width = self.left_output_text.winfo_width() - 20  # 预留边距

                # 2. 计算缩放比例（保持宽高比）
                if widget_width < 1:  # 避免初始宽度为0
                    widget_width = 300  # 默认宽度
                ratio = widget_width / image.width
                new_height = int(image.height * ratio)

                # 3. 缩放图片
                resized_image = image.resize((widget_width, new_height), Image.LANCZOS)
                photo = ImageTk.PhotoImage(resized_image)

                # 4. 插入文字和图片并保持引用
                self.clearContent(self.left_output_text)
                self.left_output_text.insert("end", "")# 插入文字
                self.left_output_text.image_create("end", image=photo)  # 在末尾插入图片
                self.left_output_text.insert("end", "\n")  # 继续插入文字
                self.left_output_text.image = photo  # 防止被垃圾回收

                # 调用OCR方法，先插入图片到文本框里了，再用图片路径识别文字，插入文字到文本框里
                self.see_img_and_insert_text(self.left_output_text, img_path, lang, 1)

                # 从插入图片和文字的文本框里进行问答操作
                self.thread_it(self.askAI(self.left_output_text, self.right_output_text, "0", "gpt-4o-mini"))

                #save_qa_data(self.left_text_output.get("1.0", "end-1c"), self.right_text_output.get("1.0", "end-1c"))


            except Exception as e:
                self.left_output_text.insert(tk.END, f'调用{self.select_image_for_ocr_qa.__name__}方法进行图片处理失败: {e}')
                #messagebox.showerror(" 错误", f"图片处理失败: {str(e)}")

    def image_path_for_ocr_qa(self, img_path):
        if img_path:
            try:
                # 调用OCR识别  param lang must in dict_keys(['ch', 'en', 'korean', 'japan', 'chinese_cht', 'ta', 'te', 'ka', 'latin', 'arabic', 'cyrillic', 'devanagari']), but got chi_sim
                lang_map = {"中文": "ch", "英文": "en",
                            "日文": "japan", "韩文": "korean", "阿拉伯文": "arabic"}
                selected_lang = self.choice_lang_combo.get()
                lang = lang_map[selected_lang]

                # 调用OCR方法
                self.see_img_and_insert_text(self.left_output_text, img_path, lang, 1)

                #self.askAI(self.left_img_output, self.text_output,"gpt-4o-mini")

                #save_qa_data(self.left_text_output.get("1.0", "end-1c"), self.right_text_output.get("1.0", "end-1c"))

            except Exception as e:
                self.left_output_text.insert(tk.END, f'图片处理失败: {e}')
                #messagebox.showerror(" 错误", f"图片处理失败: {str(e)}")

    def select_image_for_ocr_qa2(self,img_path):
        if img_path:
            try:
                # 调用OCR识别  param lang must in dict_keys(['ch', 'en', 'korean', 'japan', 'chinese_cht', 'ta', 'te', 'ka', 'latin', 'arabic', 'cyrillic', 'devanagari']), but got chi_sim
                lang_map = {"中文": "ch", "英文": "en",
                            "日文": "japan", "韩文": "korean", "阿拉伯文": "arabic"}
                selected_lang = self.choice_lang_combo.get()
                lang = lang_map[selected_lang]

                # # 1. 打开图片并获取控件宽度
                # image = Image.open(img_path)
                # widget_width = self.left_output_text.winfo_width() - 20  # 预留边距
                #
                # # 2. 计算缩放比例（保持宽高比）
                # if widget_width < 1:  # 避免初始宽度为0
                #     widget_width = 300  # 默认宽度
                # ratio = widget_width / image.width
                # new_height = int(image.height * ratio)
                #
                # # 3. 缩放图片
                # resized_image = image.resize((widget_width, new_height), Image.LANCZOS)
                # photo = ImageTk.PhotoImage(resized_image)
                #
                # # 4. 插入文字和图片并保持引用
                # self.left_output_text.insert("end", "")  # 插入文字
                # self.left_output_text.image_create("end", image=photo)  # 在末尾插入图片
                # self.left_output_text.insert("end", "\n")  # 继续插入文字
                # self.left_output_text.image = photo  # 防止被垃圾回收

                # 调用OCR方法
                self.see_img_and_insert_text(self.left_output_text, img_path, lang, 1)

                self.askAI(self.left_output_text, self.right_output_text, "gpt-4o-mini")

                save_qa_data(self.left_output_text.get("1.0", "end-1c"), self.right_output_text.get("1.0", "end-1c"))

            except Exception as e:
                self.left_output_text.insert(tk.END, f'图片处理失败: {e}')
                #messagebox.showerror(" 错误", f"图片处理失败: {str(e)}")

    def select_image_for_ocr_tk1(self,img_path):
        if img_path:
            try:
                # 调用OCR识别  param lang must in dict_keys(['ch', 'en', 'korean', 'japan', 'chinese_cht', 'ta', 'te', 'ka', 'latin', 'arabic', 'cyrillic', 'devanagari']), but got chi_sim
                lang_map = {"中文": "ch", "英文": "en",
                            "日文": "japan", "韩文": "korean", "阿拉伯文": "arabic"}
                selected_lang = self.lang_combo_see.get()
                lang = lang_map[selected_lang]

                # 1. 打开图片并获取控件宽度
                image = Image.open(img_path)
                widget_width = self.left_img_output_text.winfo_width() - 20  # 预留边距

                # 2. 计算缩放比例（保持宽高比）
                if widget_width < 1:  # 避免初始宽度为0
                    widget_width = 300  # 默认宽度
                ratio = widget_width / image.width
                new_height = int(image.height * ratio)

                # 3. 缩放图片
                resized_image = image.resize((widget_width, new_height), Image.LANCZOS)
                photo = ImageTk.PhotoImage(resized_image)

                # 4. 插入文字和图片并保持引用
                #self.left_img_output_text.insert("end", "")  # 插入文字
                self.left_img_output_text.image_create("end", image=photo)  # 在末尾插入图片
                self.left_img_output_text.insert("end", "\n")  # 继续插入文字
                self.left_img_output_text.image = photo  # 防止被垃圾回收

                # 调用OCR方法
                self.see_img_and_insert_text(self.text_output, img_path, lang, 1)

                #self.askAI(self.left_img_output, self.text_output,"gpt-4o-mini")

                #save_qa_data(self.left_text_output.get("1.0", "end-1c"), self.right_text_output.get("1.0", "end-1c"))

            except Exception as e:
                self.left_img_output_text.insert(tk.END, f'图片处理失败: {e}')
                #messagebox.showerror(" 错误", f"图片处理失败: {str(e)}")

    def seeOnly_imageText_via_ocr(self, img_path, pattern):
        if img_path:
            try:
                # 调用OCR识别  param lang must in dict_keys(['ch', 'en', 'korean', 'japan', 'chinese_cht', 'ta', 'te', 'ka', 'latin', 'arabic', 'cyrillic', 'devanagari']), but got chi_sim
                lang_map = {"中文": "ch", "英文": "en",
                            "日文": "japan", "韩文": "korean", "阿拉伯文": "arabic"}
                #selected_lang = self.lang_combo.get()
                lang = lang_map["中文"]
                # 调用OCR方法
                self.see_img_to_copy_or_translate_text(img_path, lang, pattern)
            except Exception as e:
                messagebox.showerror(" 错误", f"图片处理失败: {str(e)}")
    """
    上述代码不符合要求，选择插入的图片未能自适应填充满img_frame区域，要求不改变代码的基本功能进行修改
    """

    def adjust_combobox_width(which_combobox):
        # 获取最长条目的宽度
        max_len = max(len(x) for x in which_combobox['values'])
        which_combobox.configure(width=max_len + 2)  # 增加额外边距

    def add_function_buttons_in_qa_text(self, which_win, which_text, pattern):
        # 创建功能按钮的Frame（放在Text控件底部）
        button_frame1 = tk.Frame(which_text.master,highlightbackground='lightblue',background='lightblue')
        button_frame1.pack(side=tk.BOTTOM, fill=tk.X)

        if pattern==1:
            question_btn = tk.Button(
                button_frame1,
                font=self.set_font_size('Song',10,'normal'),
                text="Question",
                background='blue',
                activebackground='red',
                highlightbackground='blue',  # tk.Text背景色
                command=lambda: self.thread_it(self.askAI(self.left_output_text, self.right_output_text, "gpt-4o-mini"))
            )  # 提问
            question_btn.pack(side=tk.LEFT, padx=2, pady=2)

            screenshot_btn = tk.Button(
                button_frame1,
                font=self.set_font_size('Song', 10, 'normal'),
                text="ScreenQA",
                highlightbackground='lightblue',  # tk.Text背景色
                command=lambda: self.thread_it(self.start_region_selection_own_qa_inputTextWin(self))
            )
            screenshot_btn.pack(side=tk.LEFT, padx=2, pady=2)

            choice_btn = tk.Button(
                button_frame1,
                font=self.set_font_size('Song', 10, 'normal'),
                text="Choice",
                highlightbackground='lightblue',  # tk.Text背景色
                command=lambda: self.thread_it(self.select_image_for_ocr_qa())
            )
            choice_btn.pack(side=tk.LEFT, padx=2, pady=2)

            only_screenshot_btn = tk.Button(
                button_frame1,
                font=self.set_font_size('Song', 10, 'normal'),
                text="ScreenOnly",
                highlightbackground='lightblue',  # tk.Text背景色
                command=lambda: self.start_region_selection_only_see_img()
            )
            only_screenshot_btn.pack(side=tk.LEFT, padx=2, pady=2)

            # #截屏翻译英文
            # only_screen_trans_btn = tk.Button(
            #     button_frame1,
            #     font=self.set_font_size('Song', 10, 'normal'),
            #     text="ScreenTrans",
            #     highlightbackground='lightblue',  # tk.Text背景色
            #     command=lambda: self.start_region_selection_see_English_img_and_translate()
            # )
            # only_screen_trans_btn.pack(side=tk.LEFT, padx=2, pady=2)

            # 搜索按钮
            search_qa_btn = tk.Button(
                button_frame1,
                text="Find",
                highlightbackground='lightblue',  # tk.Button按钮背景色
                font=self.set_font_size('Song', 10, 'normal'),
                anchor="center",
                command=lambda: self.search_keyword_in_text(search_qa_combobox.get(), which_text)
            )
            search_qa_btn.pack(side=tk.LEFT, padx=2, pady=2)

            # 搜索which_text关键字结果
            search_qa_combobox = ttk.Combobox(
                button_frame1,
                font=self.set_font_size('Song', 10, 'normal'),
                state='NORMAL'
            )
            search_qa_combobox.pack(side=tk.LEFT)
            search_qa_combobox['values'] = ['']

        if pattern==2:
            # 自定义功能
            only_screenshot_btn = tk.Button(
                button_frame1,
                font=self.set_font_size('Song', 10, 'normal'),
                text="ScreenOnly",
                highlightbackground='lightblue',  # tk.Text背景色
                command=lambda: self.start_region_selection_only_see_img()
            )
            only_screenshot_btn.pack(side=tk.LEFT, padx=2, pady=2)

            # # 截屏翻译英文
            # only_screen_trans_btn = tk.Button(
            #     button_frame1,
            #     font=self.set_font_size('Song', 10, 'normal'),
            #     text="ScreenTrans",
            #     highlightbackground='lightblue',  # tk.Text背景色
            #     command=lambda: self.start_region_selection_see_English_img_and_translate()
            # )
            # only_screen_trans_btn.pack(side=tk.LEFT, padx=2, pady=2)

            # 搜索按钮
            search_qa_btn = tk.Button(
                button_frame1,
                text="Find",
                highlightbackground='lightblue',  # tk.Button按钮背景色
                font=self.set_font_size('Song', 10, 'normal'),
                anchor="center",
                command=lambda: self.search_keyword_in_text(search_qa_combobox.get(), which_text)
            )
            search_qa_btn.pack(side=tk.LEFT, padx=2, pady=2)

            # 搜索which_text关键字结果
            search_qa_combobox = ttk.Combobox(
                button_frame1,
                font=self.set_font_size('Song', 10, 'normal'),
                state='NORMAL'
            )
            search_qa_combobox.pack(side=tk.LEFT)
            search_qa_combobox['values'] = ['']

        # 创建功能按钮的Frame（放在Text控件底部）
        button_frame2 = tk.Frame(which_text.master,background='lightblue')
        button_frame2.pack(side=tk.BOTTOM, fill=tk.X)

        # 复制按钮
        copy_btn = tk.Button(
            button_frame2,
            font=self.set_font_size('Song', 10, 'normal'),
            text="Copy",
            highlightbackground='lightblue',  # tk.Text背景色
            command=lambda: self.copy_content(which_text)
        )
        copy_btn.pack(side=tk.LEFT, padx=2, pady=2)

        # 粘贴按钮
        paste_btn = tk.Button(
            button_frame2,
            font=self.set_font_size('Song', 10, 'normal'),
            highlightbackground='lightblue',  # tk.Text背景色
            text="Paste",
            command=lambda: self.paste(which_win, which_text)
        )
        paste_btn.pack(side=tk.LEFT, padx=2, pady=2)

        # 剪切按钮
        cut_btn = tk.Button(
            button_frame2,
            font=self.set_font_size('Song', 10, 'normal'),
            highlightbackground='lightblue',  # tk.Text背景色
            text="Cut",
            command=lambda: self.cut_content(which_text)
        )
        cut_btn.pack(side=tk.LEFT, padx=2, pady=2)

        # 清空按钮
        clear_btn = tk.Button(
            button_frame2,
            font=self.set_font_size('Song', 10, 'normal'),
            highlightbackground='lightblue',  # tk.Text背景色
            text="Clear",
            command=lambda: self.clearContent(which_text)
        )
        clear_btn.pack(side=tk.LEFT, padx=2, pady=2)

        # 复制图片
        copy_image_btn = tk.Button(
            button_frame2,
            font=self.set_font_size('Song', 10, 'normal'),
            text="CopyImage",
            highlightbackground='lightblue',  # tk.Text背景色
            command=lambda: self.copy_image_to_clipboard())
        copy_image_btn.pack(side=tk.LEFT, padx=2, pady=2)

        # 粘贴图片
        paste_image_btn = tk.Button(
            button_frame2,
            font=self.set_font_size('Song', 10, 'normal'),
            text="PasteImage",
            highlightbackground='lightblue',  # tk.Text背景色
            command=lambda: self.paste_image_to_text(which_text,'qa')
        )
        paste_image_btn.pack(side=tk.LEFT, padx=2, pady=2)

    def add_function_buttons_in_request_text_for_tkText_pack(self, which_win, which_tkText, pattern):
        # 创建功能按钮的Frame
        # ["raised", "sunken", "flat", "ridge", "solid", "groove"]
        button_frame1 = tk.Frame(
            which_tkText.master,
            background='lightblue',
            highlightbackground='lightblue',
            relief="raised"
        )
        button_frame1.pack(side=tk.BOTTOM, fill=tk.X)
        # 格式化json
        question_btn = tk.Button(
            button_frame1,
            anchor='center',
            font=self.set_font_size('Song', 10, 'normal'),
            bg='yellow',
            highlightbackground='lightblue',  # tk.Button按钮背景色
            relief="flat",  # 去除默认边框
            text="Format",
            command=lambda: self.thread_it(self.format_content(which_tkText))
        )
        question_btn.pack(side=tk.LEFT, padx=2, pady=2,  expand=True, fill="both")

        # 复制按钮
        copy_btn = tk.Button(
            button_frame1,
            anchor='center',
            font=self.set_font_size('Song', 10, 'normal'),
            highlightbackground='lightblue',  # tk.Button按钮背景色
            text="Copy",
            command=lambda: self.copy_content(which_tkText)
        )
        copy_btn.pack(side=tk.LEFT, padx=2, pady=2,  expand=True, fill="both")

        # 粘贴按钮
        paste_btn = tk.Button(
            button_frame1,
            font=self.set_font_size('Song', 10, 'normal'),
            highlightbackground='lightblue',  # tk.Button按钮背景色
            text="Paste",
            command=lambda: self.paste(which_win,which_tkText)
        )
        paste_btn.pack(side=tk.LEFT, padx=2, pady=2,  expand=True, fill="both")

        button_frame4 = tk.Frame(which_tkText.master, background='lightblue')
        button_frame4.pack(side=tk.TOP, fill=tk.X)
        # 剪切按钮
        cut_btn = tk.Button(
            button_frame1,
            font=self.set_font_size('Song', 10, 'normal'),
            highlightbackground='lightblue',  # tk.Button按钮背景色
            text="Cut",
            command=lambda: self.cut_content(which_tkText)
        )
        cut_btn.pack(side=tk.LEFT, padx=2, pady=2,  expand=True, fill="both")

        # 清空按钮
        clear_btn = tk.Button(
            button_frame1,
            font=self.set_font_size('Song', 10, 'normal'),
            highlightbackground='lightblue',  # tk.Button按钮背景色
            text="Clear",
            command=lambda: self.clearContent(which_tkText)
        )
        clear_btn.pack(side=tk.LEFT, padx=2, pady=2,  expand=True, fill="both")

        # 截屏识字,复制到粘贴板
        if pattern == 1:
            # 格式化json
            screenshot_btn = tk.Button(
                button_frame1,
                font=self.set_font_size('Song', 10, 'normal'),
                highlightbackground='lightblue',  # tk.Button按钮背景色
                text="Screen",
                command=lambda: self.thread_it(self.start_region_selection_only_see_img())
            )
            screenshot_btn.pack(side=tk.LEFT, padx=2, pady=2,  expand=True, fill="both")

    def add_function_buttons_in_img_text(self, which_win, which_text, pattern):
        # 创建功能按钮的Frame（放在Text控件底部）
        button_frame1 = tk.Frame(
            which_text.master,
            highlightbackground='lightblue',
            background='lightblue'
        )
        button_frame1.pack(side=tk.BOTTOM, fill=tk.X)

        if pattern==1:
            # 自定义功能
            select_btn1 = tk.Button(
                button_frame1,
                highlightbackground='lightblue',  # tk.Button按钮背景色
                text="ScreenSee",
                font=self.set_font_size('Song',10,'normal'),
                command=lambda: self.thread_it(self.start_region_selection_has_img_inputTextWin(self))
            )
            select_btn1.pack(side=tk.LEFT, padx=2, pady=2)

            select_btn2 = tk.Button(
                button_frame1,
                highlightbackground='lightblue',  # tk.Button按钮背景色
                text="Choice",
                font=self.set_font_size('Song', 10, 'normal'),
                command=lambda: self.thread_it(self.select_image_for_ocr_tk())
            )
            select_btn2.pack(side=tk.LEFT, padx=2, pady=2)

            only_screenshot_btn = tk.Button(
                button_frame1,
                highlightbackground='lightblue',  # tk.Button按钮背景色
                text="ScreenOnly",
                font=self.set_font_size('Song', 10, 'normal'),
                command=lambda: self.start_region_selection_only_see_img()
            )
            only_screenshot_btn.pack(side=tk.LEFT, padx=2, pady=2)

            # 搜索按钮
            search_img_btn = tk.Button(
                button_frame1,
                text="Find",
                highlightbackground='lightblue',  # tk.Button按钮背景色
                font=self.set_font_size('Song', 10, 'normal'),
                anchor="center",
                command=lambda: self.search_keyword_in_text(search_qa_combobox.get(), which_text)
            )
            search_img_btn.pack(side=tk.LEFT, padx=2, pady=2)

            # 搜索which_text关键字结果
            search_img_combobox = ttk.Combobox(
                button_frame1,
                font=self.set_font_size('Song', 10, 'normal'),
                state='NORMAL'
            )
            search_img_combobox.pack(side=tk.LEFT)
            search_img_combobox['values'] = ['']

        if pattern==2:
            # 自定义功能
            only_screenshot_btn = tk.Button(
                button_frame1,
                highlightbackground='lightblue',  # tk.Button按钮背景色
                text="ScreenOnly",
                font=self.set_font_size('Song', 10, 'normal'),
                command=lambda: self.start_region_selection_only_see_img()
            )
            only_screenshot_btn.pack(side=tk.LEFT, padx=2, pady=2)

            # 搜索按钮
            search_img_btn = tk.Button(
                button_frame1,
                text="Find",
                highlightbackground='lightblue',  # tk.Button按钮背景色
                font=self.set_font_size('Song', 10, 'normal'),
                anchor="center",
                command=lambda: self.search_keyword_in_text(search_img_combobox.get(), which_text)
            )
            search_img_btn.pack(side=tk.LEFT, padx=2, pady=2)

            # 搜索which_text关键字结果
            search_img_combobox = ttk.Combobox(
                button_frame1,
                font=self.set_font_size('Song', 10, 'normal'),
                state='NORMAL'
            )
            search_img_combobox.pack(side=tk.LEFT)
            search_img_combobox['values'] = ['']

        # 创建功能按钮的Frame（放在Text控件底部）
        button_frame2 = tk.Frame(which_text.master,background='lightblue')
        button_frame2.pack(side=tk.BOTTOM, fill=tk.X)

        # 复制按钮
        copy_btn = tk.Button(
            button_frame2,
            highlightbackground='lightblue',  # tk.Button按钮背景色
            text="Copy",
            font=self.set_font_size('Song', 10, 'normal'),
            command=lambda: self.copy_content(which_text)
        )
        copy_btn.pack(side=tk.LEFT, padx=2, pady=2)

        # 粘贴按钮
        paste_btn = tk.Button(
            button_frame2,
            highlightbackground='lightblue',  # tk.Button按钮背景色
            text="Paste",
            font=self.set_font_size('Song', 10, 'normal'),
            command=lambda: self.paste(which_win, which_text)
        )
        paste_btn.pack(side=tk.LEFT, padx=2, pady=2)

        # 剪切按钮
        cut_btn = tk.Button(
            button_frame2,
            highlightbackground='lightblue',  # tk.Button按钮背景色
            text="Cut",
            font=self.set_font_size('Song', 10, 'normal'),
            command=lambda: self.cut_content(which_text)
        )
        cut_btn.pack(side=tk.LEFT, padx=2, pady=2)

        # 清空按钮
        clear_btn = tk.Button(
            button_frame2,
            highlightbackground='lightblue',  # tk.Button按钮背景色
            text="Clear",
            font=self.set_font_size('Song', 10, 'normal'),
            command=lambda: self.clearContent(which_text)
        )
        clear_btn.pack(side=tk.LEFT, padx=2, pady=2)

        # 复制图片
        copy_image_btn = tk.Button(
            button_frame2,
            highlightbackground='lightblue',  # tk.Button按钮背景色
            text="CopyImage",
            font=self.set_font_size('Song', 10, 'normal'),
            command=lambda: self.copy_image_to_clipboard())
        copy_image_btn.pack(side=tk.LEFT, padx=2, pady=2)

        # 粘贴图片
        paste_image_btn = tk.Button(
            button_frame2,
            highlightbackground='lightblue',  # tk.Button按钮背景色
            text="PasteImage",
            font=self.set_font_size('Song', 10, 'normal'),
            command=lambda: self.paste_image_to_text(which_text,'noqa'))
        paste_image_btn.pack(side=tk.LEFT, padx=2, pady=2)

    def add_function_buttons_in_request_text(self, which_win, which_text, pattern):
        # 创建功能按钮的Frame
        #["raised", "sunken", "flat", "ridge", "solid", "groove"]
        button_frame1 = tk.Frame(
            which_text.master,
            background='lightblue',
            highlightbackground='lightblue',
            relief="raised"
        )
        button_frame1.pack(side=tk.TOP, fill=tk.X)
        # 格式化json
        question_btn = tk.Button(
            button_frame1,
            highlightbackground='lightblue',  # tk.Button按钮背景色
            anchor='center',
            font=self.set_font_size('Song', 10, 'normal'),
            bg='yellow',
            relief="flat",  # 去除默认边框
            text="Format",
            command=lambda: self.thread_it(self.format_content(which_text))
        )
        question_btn.pack(side=tk.LEFT, padx=2, pady=2)

        button_frame2 = tk.Frame(which_text.master,background='lightblue')
        button_frame2.pack(side=tk.TOP, fill=tk.X)
        # 复制按钮
        copy_btn = tk.Button(
            button_frame2,
            highlightbackground='lightblue',  # tk.Button按钮背景色
            anchor='center',
            font=self.set_font_size('Song', 10, 'normal'),
            text="Copy",
            command=lambda: self.copy_content(which_text)
        )
        copy_btn.pack(side=tk.LEFT, padx=2, pady=2)

        button_frame3 = tk.Frame(which_text.master,background='lightblue')
        button_frame3.pack(side=tk.TOP, fill=tk.X)
        # 粘贴按钮
        paste_btn = tk.Button(
            button_frame3,
            highlightbackground='lightblue',  # tk.Button按钮背景色
            font=self.set_font_size('Song', 10, 'normal'),
            text="Paste",
            command=lambda: self.paste(which_win, which_text)
        )
        paste_btn.pack(side=tk.LEFT, padx=2, pady=2)

        button_frame4 = tk.Frame(which_text.master,background='lightblue')
        button_frame4.pack(side=tk.TOP, fill=tk.X)
        # 剪切按钮
        cut_btn = tk.Button(
            button_frame4,
            highlightbackground='lightblue',  # tk.Button按钮背景色
            font=self.set_font_size('Song', 10, 'normal'),
            text="Cut",
            command=lambda: self.cut_content(which_text)
        )
        cut_btn.pack(side=tk.LEFT, padx=2, pady=2)

        button_frame5 = tk.Frame(which_text.master,background='lightblue')
        button_frame5.pack(side=tk.TOP, fill=tk.X)
        # 清空按钮
        clear_btn = tk.Button(
            button_frame5,
            highlightbackground='lightblue',  # tk.Button按钮背景色
            font=self.set_font_size('Song', 10, 'normal'),
            text="Clear",
            command=lambda: self.clearContent(which_text)
        )
        clear_btn.pack(side=tk.LEFT, padx=2, pady=2)

        # 截屏识字,复制到粘贴板
        if pattern == 1:
            button_frame6 = tk.Frame(which_text.master, width=100,background='lightblue')
            button_frame6.pack(side=tk.TOP, fill=tk.X)
            # 格式化json
            screenshot_btn = tk.Button(
                button_frame6,
                highlightbackground='lightblue',  # tk.Button按钮背景色
                font=self.set_font_size('Song', 10, 'normal'),
                text="Screen",
                command=lambda: self.thread_it(self.start_region_selection_only_see_img())
            )
            screenshot_btn.pack(side=tk.LEFT, padx=2, pady=2)

    def insert_content_in_text(self,which_text,content):
        which_text.insert(tk.END,f'\n{content}')

    def start_region_selection_own_qa_inputTextWin(self, whichWin):
        """启动区域选择模式"""
        # whichWin.withdraw()   # 隐藏窗口

        # 创建窗口
        self.region_window = Toplevel()
        self.region_window.overrideredirect(True)  # 移除窗口装饰

        # macOS专用透明设置
        self.region_window.configure(bg='systemTransparent')  # 设置背景为系统级透明
        self.region_window.attributes('-transparent', True)  # 启用透明属性
        self.region_window.attributes('-alpha', 0.8) # 边框整体透明度（0-1）
        self.region_window.attributes('-topmost', True)
        # 设置全屏尺寸
        screen_width = self.region_window.winfo_screenwidth()
        screen_height = self.region_window.winfo_screenheight()
        self.region_window.geometry(f"{screen_width}x{screen_height}+0+0")

        # 创建画布（使用标准颜色格式）
        self.canvas = Canvas(
            self.region_window,
            cursor="cross",
            bg='black',  # 改为标准颜色格式
            highlightthickness=0
        )
        self.canvas.pack(fill=tk.BOTH, expand=True)

        # 通过设置画布透明度来实现效果
        self.canvas.config(bg='systemTransparent')  # 系统透明色
        self.canvas['highlightbackground'] = 'systemTransparent'  # 系统透明色

        # 绑定事件
        self.canvas.bind("<ButtonPress-1>", self.on_press)
        self.canvas.bind("<B1-Motion>", self.on_drag)
        self.canvas.bind("<ButtonRelease-1>", self.on_release)
        self.region_window.bind("<Escape>", self.cancel_selection)

        # 提示文本
        self.canvas.create_text(
            screen_width // 2,
            30,
            text="拖动鼠标选择区域 (ESC取消)",
            fill="lightblue",
            font=('PingFang SC', 26)
        )

        # 强制刷新
        self.region_window.update_idletasks()

    def start_region_selection_has_img_inputTextWin(self, whichWin):
        """启动区域选择模式"""
        # whichWin.withdraw()   # 隐藏窗口

        # 创建窗口
        self.region_window = Toplevel()
        self.region_window.overrideredirect(True)  # 移除窗口装饰

        # macOS专用透明设置
        self.region_window.configure(bg='systemTransparent')
        self.region_window.attributes('-transparent', True)
        self.region_window.attributes('-alpha', 0.7)
        self.region_window.attributes('-topmost', True)

        # 设置全屏尺寸
        screen_width = self.region_window.winfo_screenwidth()
        screen_height = self.region_window.winfo_screenheight()
        self.region_window.geometry(f"{screen_width}x{screen_height}+0+0")

        # 创建画布（使用标准颜色格式）
        self.canvas = Canvas(
            self.region_window,
            cursor="cross",
            bg='black',  # 改为标准颜色格式
            highlightthickness=0
        )
        self.canvas.pack(fill=tk.BOTH, expand=True)

        # 通过设置画布透明度来实现效果
        self.canvas.config(bg='systemTransparent')  # 系统透明色
        self.canvas['highlightbackground'] = 'systemTransparent'  # 系统透明色

        # 绑定事件
        self.canvas.bind("<ButtonPress-1>", self.on_press)
        self.canvas.bind("<B1-Motion>", self.on_drag)
        self.canvas.bind("<ButtonRelease-1>", self.on_release1)
        self.region_window.bind("<Escape>", self.cancel_selection)

        # 提示文本
        self.canvas.create_text(
            screen_width // 2,
            30,
            text="拖动鼠标选择区域 (ESC取消)",
            fill="lightblue",
            font=('PingFang SC', 26)
        )

        # 强制刷新
        self.region_window.update_idletasks()

    def start_region_selection_only_see_img(self):
        """启动区域选择模式"""
        # whichWin.withdraw()   # 隐藏窗口

        # 创建窗口
        self.region_window = Toplevel()
        self.region_window.overrideredirect(True)  # 移除窗口装饰

        # macOS专用透明设置
        self.region_window.configure(bg='systemTransparent')
        self.region_window.attributes('-transparent', True)
        self.region_window.attributes('-alpha', 0.7)
        self.region_window.attributes('-topmost', True)

        # 设置全屏尺寸
        screen_width = self.region_window.winfo_screenwidth()
        screen_height = self.region_window.winfo_screenheight()
        self.region_window.geometry(f"{screen_width}x{screen_height}+0+0")

        # 创建画布（使用标准颜色格式）
        self.canvas = Canvas(
            self.region_window,
            cursor="cross",
            bg='black',  # 改为标准颜色格式
            highlightthickness=0
        )
        self.canvas.pack(fill=tk.BOTH, expand=True)

        # 通过设置画布透明度来实现效果
        self.canvas.config(bg='systemTransparent')  # 系统透明色
        self.canvas['highlightbackground'] = 'systemTransparent'  # 系统透明色

        # 绑定事件
        self.canvas.bind("<ButtonPress-1>", self.on_press)
        self.canvas.bind("<B1-Motion>", self.on_drag)
        self.canvas.bind("<ButtonRelease-1>", self.on_release_only_see_img)
        self.region_window.bind("<Escape>", self.cancel_selection)

        # 提示文本
        self.canvas.create_text(
            screen_width // 2,
            30,
            text="拖动鼠标选择区域 (ESC取消)",
            fill="lightblue",
            font=('PingFang SC', 26)
        )

        # 强制刷新
        self.region_window.update_idletasks()

    def start_region_selection_see_English_img_and_translate(self):
        """启动区域选择模式"""
        # whichWin.withdraw()   # 隐藏窗口

        # 创建窗口
        self.region_window = Toplevel()
        self.region_window.overrideredirect(True)  # 移除窗口装饰

        # macOS专用透明设置
        self.region_window.configure(bg='systemTransparent')
        self.region_window.attributes('-transparent', True)
        self.region_window.attributes('-alpha', 0.7)
        self.region_window.attributes('-topmost', True)

        # 设置全屏尺寸
        screen_width = self.region_window.winfo_screenwidth()
        screen_height = self.region_window.winfo_screenheight()
        self.region_window.geometry(f"{screen_width}x{screen_height}+0+0")

        # 创建画布（使用标准颜色格式）
        self.canvas = Canvas(
            self.region_window,
            cursor="cross",
            bg='black',  # 改为标准颜色格式
            highlightthickness=0
        )
        self.canvas.pack(fill=tk.BOTH, expand=True)

        # 通过设置画布透明度来实现效果
        self.canvas.config(bg='systemTransparent')  # 系统透明色
        self.canvas['highlightbackground'] = 'systemTransparent'  # 系统透明色

        # 绑定事件
        self.canvas.bind("<ButtonPress-1>", self.on_press)
        self.canvas.bind("<B1-Motion>", self.on_drag)
        self.canvas.bind("<ButtonRelease-1>", self.on_release_see_English_img_and_translate)
        self.region_window.bind("<Escape>", self.cancel_selection)

        # 提示文本
        self.canvas.create_text(
            screen_width // 2,
            30,
            text="拖动鼠标选择区域 (ESC取消)",
            fill="lightblue",
            font=('PingFang SC', 26)
        )

        # 强制刷新
        self.region_window.update_idletasks()

    def on_press(self, event):
        """鼠标按下事件"""
        self.start_x = event.x
        self.start_y = event.y

        # 创建选择矩形
        self.rect = self.canvas.create_rectangle(
            self.start_x, self.start_y,
            self.start_x, self.start_y,
            outline='red', width=2
        )

    def on_drag(self, event):
        """鼠标拖动事件"""
        if self.rect:
            self.canvas.coords(
                self.rect,
                self.start_x, self.start_y,
                event.x, event.y
            )

    def on_release(self, event):
        """鼠标释放事件"""
        if not self.rect:
            return

        # 获取选择的区域坐标
        x1, y1, x2, y2 = self.canvas.coords(self.rect)
        x1, y1, x2, y2 = int(x1), int(y1), int(x2), int(y2)

        # 确保坐标有效
        if abs(x2 - x1) < 10 or abs(y2 - y1) < 10:
            self.cancel_selection()
            return

            # 关闭选择窗口
        self.region_window.destroy()
        self.region_window = None

        # 短暂延迟确保窗口关闭
        self.after(100, lambda: self.capture_region(x1, y1, x2, y2))

    def on_release1(self, event):
        """鼠标释放事件"""
        if not self.rect:
            return

        # 获取选择的区域坐标
        x1, y1, x2, y2 = self.canvas.coords(self.rect)
        x1, y1, x2, y2 = int(x1), int(y1), int(x2), int(y2)

        # 确保坐标有效
        if abs(x2 - x1) < 10 or abs(y2 - y1) < 10:
            self.cancel_selection()
            return

            # 关闭选择窗口
        self.region_window.destroy()
        self.region_window = None

        # 短暂延迟确保窗口关闭
        self.after(100, lambda: self.capture_region1(x1, y1, x2, y2))

    def on_release_only_see_img(self, event):
        """鼠标释放事件"""
        if not self.rect:
            return

        # 获取选择的区域坐标
        x1, y1, x2, y2 = self.canvas.coords(self.rect)
        x1, y1, x2, y2 = int(x1), int(y1), int(x2), int(y2)

        # 确保坐标有效
        if abs(x2 - x1) < 10 or abs(y2 - y1) < 10:
            self.cancel_selection()
            return

            # 关闭选择窗口
        self.region_window.destroy()
        self.region_window = None

        # 短暂延迟确保窗口关闭  capture_region_see_English_img_and_translate
        self.after(100, lambda: self.capture_region_only_see_img(x1, y1, x2, y2))

    def on_release_see_English_img_and_translate(self, event):
        """鼠标释放事件"""
        if not self.rect:
            return

        # 获取选择的区域坐标
        x1, y1, x2, y2 = self.canvas.coords(self.rect)
        x1, y1, x2, y2 = int(x1), int(y1), int(x2), int(y2)

        # 确保坐标有效
        if abs(x2 - x1) < 10 or abs(y2 - y1) < 10:
            self.cancel_selection()
            return

            # 关闭选择窗口
        self.region_window.destroy()
        self.region_window = None

        # 短暂延迟确保窗口关闭
        self.after(100, lambda: self.capture_region_see_English_img_and_translate(x1, y1, x2, y2))

    def cancel_selection(self, event=None):
        """取消选择"""
        if self.region_window:
            self.region_window.destroy()
            self.region_window = None
        self.deiconify()

    def capture_region(self, x1, y1, x2, y2):
        """捕获选定区域"""
        try:
            # 调整坐标确保左上角到右下角
            if x1 > x2:
                x1, x2 = x2, x1
            if y1 > y2:
                y1, y2 = y2, y1

            # 截图
            self.screenshot = ImageGrab.grab(bbox=(x1, y1, x2, y2))

            # # 显示预览窗口
            # self.show_preview()

            # 生成带时间戳的文件名
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            default_filename = f"区域截图_{timestamp}.png"
            # 保存截图
            file_path = f'{current_script_path}/{screenshotFileName}/{default_filename}'
            self.screenshot.save(file_path)


            if os.path.exists(file_path):
                print(file_path)

                # 1. 打开图片并获取控件宽度
                image = Image.open(file_path)
                photo = ImageTk.PhotoImage(image)
                image_references[file_path] = photo  # 保存引用
                widget_width = self.left_output_text.winfo_width() - 20  # 预留边距

                # 2. 计算缩放比例（保持宽高比）
                if widget_width < 1:  # 避免初始宽度为0
                    widget_width = 300  # 默认宽度
                ratio = widget_width / image.width
                new_height = int(image.height * ratio)

                # 3. 缩放图片
                resized_image = image.resize((widget_width, new_height), Image.LANCZOS)
                photo = ImageTk.PhotoImage(resized_image)

                # 4. 插入文字和图片并保持引用
                self.clearContent(self.left_output_text)
                #self.left_text_output.insert("end", " ")  # 插入文字
                self.left_output_text.image_create("end", image=photo)  # 在末尾插入图片，"end-1c"删除前导换行符
                self.left_output_text.insert("end", "\n")  # 继续插入文字
                self.left_output_text.image = photo  # 防止被垃圾回收

            #截图识别
            self.image_path_for_ocr_qa(file_path)

            self.thread_it(self.askAI(self.left_output_text, self.right_output_text, "gpt-4o-mini"))

        except Exception as e:
            tk.messagebox.showerror(" 错误", f"调用{self.capture_region.__name__}方法截图失败: {str(e)}")
            self.deiconify()

    def capture_region1(self, x1, y1, x2, y2):
        """捕获选定区域"""
        try:
            # 调整坐标确保左上角到右下角
            if x1 > x2:
                x1, x2 = x2, x1
            if y1 > y2:
                y1, y2 = y2, y1

            # 截图
            self.screenshot = ImageGrab.grab(bbox=(x1, y1, x2, y2))

            # # 显示预览窗口
            # self.show_preview()

            # 生成带时间戳的文件名
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            default_filename = f"区域截图_{timestamp}.png"
            # 保存截图
            file_path = f'{current_script_path}/{screenshotFileName}/{default_filename}'
            self.screenshot.save(file_path)

            if os.path.exists(file_path):
                print(file_path)

                # 1. 打开图片并获取控件宽度
                image = Image.open(file_path)
                photo = ImageTk.PhotoImage(image)
                image_references[file_path] = photo  # 保存引用
                widget_width = self.left_img_output_text.winfo_width() - 20  # 预留边距

                # 2. 计算缩放比例（保持宽高比）
                if widget_width < 1:  # 避免初始宽度为0
                    widget_width = 300  # 默认宽度
                ratio = widget_width / image.width
                new_height = int(image.height * ratio)

                # 3. 缩放图片
                resized_image = image.resize((widget_width, new_height), Image.LANCZOS)
                photo = ImageTk.PhotoImage(resized_image)

                # 4. 插入文字和图片并保持引用
                #self.clearContent(self.left_text_output)
                #self.left_text_output.insert("end", " ")  # 插入文字
                # self.left_img_output_text.image_create("end-1c", image=photo)  # 在末尾插入图片，"end-1c"删除前导换行符
                # self.left_img_output_text.insert("end", "\n")  # 继续插入文字
                # self.left_img_output_text.image = photo  # 防止被垃圾回收

                print(f'输入框里的内容为：\n{self.left_img_output_text.get("1.0","end")}')

            #截图识别
            self.select_image_for_ocr_tk1(file_path)

            # self.askAI(self.left_text_output, self.right_text_output,"gpt-4o-mini")

        except Exception as e:
            tk.messagebox.showerror(" 错误", f"调用{self.capture_region1.__name__}方法截图失败: {str(e)}")
            self.deiconify()

    def capture_region_only_see_img(self, x1, y1, x2, y2):
        """捕获选定区域"""
        try:
            # 调整坐标确保左上角到右下角
            if x1 > x2:
                x1, x2 = x2, x1
            if y1 > y2:
                y1, y2 = y2, y1

            # 截图
            self.screenshot = ImageGrab.grab(bbox=(x1, y1, x2, y2))

            # # 显示预览窗口
            # self.show_preview()

            # 生成带时间戳的文件名
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            default_filename = f"区域截图_{timestamp}.png"
            # 保存截图
            file_path = f'{current_script_path}/{screenshotFileName}/{default_filename}'
            self.screenshot.save(file_path)

            #截图识别
            self.seeOnly_imageText_via_ocr(file_path,1)

            # self.askAI(self.left_text_output, self.right_text_output,"gpt-4o-mini")

        except Exception as e:
            tk.messagebox.showerror(" 错误",f"调用{self.capture_region_only_see_img.__name__}方法截图失败: {str(e)}")
            self.deiconify()

    def capture_region_see_English_img_and_translate(self, x1, y1, x2, y2):
        """捕获选定区域"""
        try:
            # 调整坐标确保左上角到右下角
            if x1 > x2:
                x1, x2 = x2, x1
            if y1 > y2:
                y1, y2 = y2, y1

            # 截图
            self.screenshot = ImageGrab.grab(bbox=(x1, y1, x2, y2))

            # # 显示预览窗口
            # self.show_preview()

            # 生成带时间戳的文件名
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            default_filename = f"区域截图_{timestamp}.png"
            # 保存截图
            file_path = f'{current_script_path}/{screenshotFileName}/{default_filename}'
            self.screenshot.save(file_path)

            #截图识别 self.clipboard
            self.seeOnly_imageText_via_ocr(file_path,0)

        except Exception as e:
            tk.messagebox.showerror(" 错误", f"调用{self.capture_region_see_English_img_and_translate.__name__}方法截图失败: {str(e)}")
            self.deiconify()

    def show_preview(self):
        """显示截图预览"""
        preview_window = Toplevel(self)
        preview_window.title(" 截图预览 - 2025年8月")
        preview_window.geometry("800x600")

        # 转换图像为Tkinter格式
        img_tk = ImageTk.PhotoImage(self.screenshot)

        # 显示图像
        label = tk.Label(preview_window, image=img_tk)
        label.image = img_tk  # 保持引用
        label.pack(pady=10)

        # 添加保存按钮
        btn_frame = tk.Frame(preview_window)
        btn_frame.pack(pady=10)

        tk.Button(
            btn_frame,
            text="保存截图",
            command=lambda: self.save_screenshot(preview_window),
            bg='#2196F3', fg='white',
            font=('微软雅黑', 12),
            padx=15
        ).pack(side=tk.LEFT, padx=10)

        tk.Button(
            btn_frame,
            text="取消",
            command=preview_window.destroy,
            bg='#9E9E9E', fg='white',
            font=('微软雅黑', 12),
            padx=15
        ).pack(side=tk.LEFT)

        # 窗口关闭时显示主窗口
        preview_window.protocol("WM_DELETE_WINDOW", preview_window.destroy)

    def save_screenshot(self, preview_window):
        """保存截图到文件"""
        try:
            # 生成带时间戳的文件名
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            default_filename = f"区域截图_{timestamp}.png"

            # 弹出保存对话框
            file_path = tk.filedialog.asksaveasfilename(
                defaultextension=".png",
                filetypes=[
                    ("PNG文件", "*.png"),
                    ("JPEG文件", "*.jpg"),
                    ("BMP文件", "*.bmp"),
                    ("所有文件", "*.*")
                ],
                initialfile=default_filename,
                title="保存区域截图"
            )

            if file_path:
                self.screenshot.save(file_path)
                tk.messagebox.showinfo(" 保存成功", f"截图已保存到:\n{file_path}")
                preview_window.destroy()
                self.deiconify()
        except Exception as e:
            tk.messagebox.showerror(" 保存失败", f"调用{self.save_screenshot.__name__}方法无法保存文件: {str(e)}")


    def nested_to_string(self, nested_list):
        """
        嵌套列表转字符串
        nested_list = [[1, [2, 3]], ['a', ['b', 'c']]]
        result = nested_to_string(nested_list)  # 输出: "1 2 3 a b c"
        """
        if isinstance(nested_list, list):
            return ' '.join(self.nested_to_string(item) for item in nested_list)
        else:
            return str(nested_list)

    def seeImgToTxtByPaddleOcr(self, which_text, img_path, lang, pattern):
        # 创建PaddleOCR对象，指定语言模型，默认为中文英文模型
        ocr = PaddleOCR(
            use_textline_orientation=True,  # 是否使用方向分类
            lang=lang  # 'ch'表示中文，'en'表示中文
        )
        # 使用OCR进行文字识别
        result = ocr.ocr(img_path, cls=True)
        print(f"识别结果是：{result}")
        print(f"识别结果是：{result[0]}")
        print(type(result))
        d = {}
        res = []
        # 结果展示
        for line in result[0]:
            coords = line[0]  # 文本坐标
            # 获取四个角坐标
            x1, y1 = coords[0]
            x2, y2 = coords[1]
            x3, y3 = coords[2]
            x4, y4 = coords[3]

            # 计算中心点坐标
            center_x = (x1 + x2 + x3 + x4) / 4
            center_y = (y1 + y2 + y3 + y4) / 4
            textResult = " "
            # print(f"检测到文本：{line[1][0]}，置信度：{line[1][1]}，该行文本的中心坐标是: ({center_x}, {center_y})")
            # print(f"{modifyTxt(line[1][0])}")
            if pattern == '文本':
                textResult = f"{line[1][0]}"
            elif pattern == '定位':
                textResult = f"检测到文本：{line[1][0]}，置信度：{line[1][1]}，该行文本的中心坐标是: ({center_x}, {center_y})"
            # 修正方案1：确保result为字符串类型
            if not isinstance(textResult, str):
                textResult = str(result)
            res.append(textResult)
            # print(f"{line[1][0]}")
            # d[line[1][0]] = (center_x, center_y)
        # return d
        resText='\n'.join(res)
        self.clearContent(which_text)
        self.copy_content_no_win(resText)
        which_text.insert('insert', resText)

    def see_img_and_insert_text(self, which_text, img_path, lang, pattern):
        # 创建PaddleOCR对象，指定语言模型，默认为中文英文模型
        ocr = PaddleOCR(
            use_textline_orientation=True,  # 是否使用方向分类
            lang=lang  # 'ch'表示中文，'en'表示中文
        )
        # 使用OCR进行文字识别
        result = ocr.ocr(img_path, cls=True)
        print(f"识别结果是：{result}")
        print(f"识别结果是：{result[0]}")
        print(type(result))
        d = {}
        res = []
        # 结果展示
        for line in result[0]:
            textResult = f"{line[1][0]}"
            # 确保result为字符串类型
            if not isinstance(textResult, str):
                textResult = str(result)
            res.append(textResult)
        resText='\n'.join(res)
        #self.clearContent(which_text)
        self.copy_content_no_win(resText)
        which_text.insert('insert', resText)

    def see_img_to_copy_or_translate_text(self, img_path, lang, pattern):
        # 创建PaddleOCR对象，指定语言模型，默认为中文英文模型
        ocr = PaddleOCR(
            use_textline_orientation=True,  # 是否使用方向分类
            lang=lang  # 'ch'表示中文，'en'表示中文
        )
        # 使用OCR进行文字识别
        result = ocr.ocr(img_path, cls=True)
        print(f"识别结果是：{result}")
        print(f"识别结果是：{result[0]}")
        print(type(result))
        d = {}
        res = []
        # 结果展示
        for line in result[0]:
            textResult = f"{line[1][0]}"
            # 确保result为字符串类型
            if not isinstance(textResult, str):
                textResult = str(result)
            res.append(textResult)
        resText='\n'.join(res)
        if pattern == 1:
            self.copy_content_no_win(resText)
        elif pattern == 0:
            print(f'resText:{resText}')
            self.translateEnglishByLLM(resText)

    def see_img_and_get_text(self, img_path, lang, pattern):
        # 创建PaddleOCR对象，指定语言模型，默认为中文英文模型
        ocr = PaddleOCR(
            use_textline_orientation=True,  # 是否使用方向分类
            lang=lang  # 'ch'表示中文，'en'表示中文
        )
        # 使用OCR进行文字识别
        result = ocr.ocr(img_path, cls=True)
        print(f"识别结果是：{result}")
        print(f"识别结果是：{result[0]}")
        print(type(result))
        d = {}
        res = []
        # 结果展示
        for line in result[0]:
            textResult = f"{line[1][0]}"
            # 确保result为字符串类型
            if not isinstance(textResult, str):
                textResult = str(result)
            res.append(textResult)
        resText='\n'.join(res)
        return resText

    # 删除日志
    def delete_records(self):
        # 设置目录路径
        directory = f'{current_script_path}/requests_log'
        # 获取所有以"20"开头的文件路径
        for filename in glob.glob(os.path.join(directory, '20*')):
            os.remove(filename)  # 移除"20"开头的文件
        self.messageInformInWin('delete is successful!',1500)
        self.load_records()

    # 删除问答日志
    def delete_QA_records(self,which_combo):
        # 设置目录路径
        directory = f'{current_script_path}/qa_log'
        # 获取所有以"20"开头的文件路径
        for filename in glob.glob(os.path.join(directory, '20*')):
            os.remove(filename)  # 移除"20"开头的文件
        self.messageInformInWin('delete is successful!', 1500)
        self.load_qa_records(which_combo)

    # 一定程度上缩进复制的网页python代码字符串
    def restore_python_code(self, which_text, code_str):  # 定义一个栈来跟踪代码块的缩进
        indent_stack = [0]
        restored_lines = []
        current_line = ''
        for line in code_str.splitlines():
            # 移除行首和行尾的空白字符
            stripped_line = line.strip()
            if stripped_line:
                # 计算当前行的缩进级别
                current_indent = len(line) - len(stripped_line)
                # 调整缩进级别
                while current_indent < indent_stack[-1]:
                    indent_stack.pop()
                if current_indent > indent_stack[-1]:
                    indent_stack.append(current_indent)
                # 添加缩进空格
                indent_spaces = ' ' * indent_stack[-1]
                # 将处理后的行添加到当前行
                current_line += '' + stripped_line
            else:
                # 如果行是空的，输出当前行并重置
                restored_lines.append(current_line)
                current_line = ''
                indent_stack = [0]
        # 添加最后一行
        if current_line:
            restored_lines.append(current_line)

        # 将恢复后的代码行合并为单一字符串
        print(restored_lines)
        restored_code = '\n'.join(restored_lines).strip()

        if 'if __name__' in restored_code:
            restored_code = restored_code.replace('if __name__', '\nif __name__')
        if 'def' in restored_code:
            restored_code = restored_code.replace('def', '\ndef')
        if 'class' in restored_code:
            restored_code = restored_code.replace('class', '\nclass')
        if 'from' in restored_code:
            restored_code = restored_code.replace('from', '\nfrom')
        if 'import' in restored_code:
            restored_code = restored_code.replace('import', '\nimport')
        if ',\n' in restored_code or ', \n' in restored_code:
            restored_code = restored_code.replace(',\n', ',')
        if '\n ' in restored_code:
            restored_code = restored_code.replace('\n ', '\n  ')
        if '(\n' in restored_code:
            restored_code = restored_code.replace('(\n', '(')
        if '=\n' in restored_code:
            restored_code = restored_code.replace('=\n', '=')
        self.clearContent(which_text)
        which_text.insert('insert', restored_code)

    def restore_python_code_by_AI(self, which_text, text_code_str):  # 用chatgpt4mini解决
        question = f"{text_code_str}这段代码格式和语法正确吗？分析包括程序的语法分析和格式分析两个步骤，回答要求如下：" \
                   f"第一步：语法不正确的话简单输出错误解释，不需要进行第二步的处理；第二步：格式不正确的话请将其修改为正确的格式输出来。要求只输出代码"
        try:
            completion = openai.chat.completions.create(
                model='gpt-4o-mini',
                messages=[{"role": "user", "content": question}],
                timeout=20  # 增加超时限制
            )
            answer = completion.choices[0].message.content
            formatted = answer
            sleep(1.5)  # 避免频繁请求
            self.clearContent(which_text)
            which_text.insert('insert', formatted.replace('```python', '').replace('```', ''))
        except Exception as e:
            print(f"问题 {question} 处理失败: {str(e)}")
            error_msg = f"ERROR: {str(e)}"
            formatted = error_msg
            self.clearContent(which_text)
            which_text.insert('insert', formatted)
        # return formatted

    def refreshKey(self, which_text, which_combobox):
        try:
            jsonStr = which_text.get("1.0", "end-1c")
            jsonDict = json.loads(jsonStr)
            print(type(jsonDict), jsonDict)
            keyList = find_all_key(jsonDict)
            print(type(keyList), keyList)
            which_combobox['values'] = keyList
        except:
            which_combobox['values'] = []

    def changeKeyValueInJsonstr(self, which_text, key_to_modify, new_value):
        try:
            jsonStr = which_text.get("1.0", "end-1c")
            jsonDict = json.loads(jsonStr)
            # 更新键对应的值
            find_key_and_update_value(jsonDict, key_to_modify, new_value)
            which_text.delete('1.0', tk.END)
            # 展示修改后的值
            display = json.dumps(jsonDict, ensure_ascii=False, indent=4)
            which_text.delete('1.0', tk.END)
            which_text.insert(tk.END, display)
            print(display)
        except Exception as e:
            which_text.delete('1.0', tk.END)
            which_text.insert(tk.END, f'更新异常: {e}')
            pass

    def changeKeyValueInJsonFile(self, json_file_path,key_to_modify, new_value):
        try:
            # 读取JSON文件
            data = read_json_file(json_file_path)

            print(type(data))
            # print(data)

            if not isinstance(data, dict):
                data = json.loads(data)

            # print(data)
            # print(type(data))

            # 修改键值，假设要修改"key_to_modify"的值，注意还没写进去文件里
            if key_to_modify in data['lobalVariable']:
                data['GlobalVariable'][key_to_modify] = new_value
            # 写回JSON文件
            with open(json_file_path, 'w', encoding=encodingType) as file:
                json.dump(data, file, ensure_ascii=False, indent=4)
            # 每调用一次函数就更新入参值
            bodyDict = json.loads(self.body_text.get("1.0", "end-1c"))
            # 修改后获取配置的全局变量值
            setKeyList = read_json_file(f'{current_script_path}/configini.json')['GlobalVariable']
            # print('setKeyList',setKeyList)#<class 'dict'>

            parameterKeyList = ["reporterCaliNo", "reporterRegisterTel", "taskHandleUM",
                                "taskHandleUlMohile""surveyorMobile", "surveyorName", "surveyorUm"]
            # 获取配置文件里的键值对
            for key in parameterKeyList:
                if 'reporterCallNo' == key or 'reporterRegisterTel' == key:
                    find_key_and_update_value(bodyDict, key, setKeyList['mobileNo'])
                if 'taskHandleUMMobile' == key or 'surveyorMobile' == key:
                    find_key_and_update_value(bodyDict, key, setKeyList['surveyumphone'])
                if 'supveyorName' == key or 'surveyorUm' == key or 'taskHandleUM' == key:
                    find_key_and_update_value(bodyDict, key, setKeyList['surveyUm'])
                    # bodyDict["survey"][key] = setKeyList['surveyUm']
            self.body_text.delete("1,0", tk.END)
            # 展示修改后的值
            display = json.dumps(bodyDict, ensure_ascii=False, indent=4)
            self.body_text.insert(tk.END, display)
            print(display)
        except Exception as e:
            self.body_text.insert(tk.END, f'更新异常: {e}')
            pass

    def changeKeyValueInJsonFile_forLLMSet(self, json_file_path, key_to_modify, new_value, which_combo, pattern):
        try:
            # 读取JSON文件
            data = read_json_file(json_file_path)
            if not isinstance(data, dict):
                data = json.loads(data)

            # 修改键值，假设要修改"key_to_modify"的值，注意还没写进去文件里
            if pattern==0:
                if key_to_modify in data and new_value not in data[key_to_modify]:
                    data[key_to_modify].append(new_value)
                    # 写回JSON文件
                    with open(json_file_path, 'w', encoding=encodingType) as file:
                        json.dump(data, file, ensure_ascii=False, indent=4)
                    self.messageInformInWin(f'modification is successful!', 2000)
                else:
                    self.messageInformInWin(f'modification exists repeatation!', 2000)
            elif pattern==1:
                try:
                    new_value = new_value.replace('\'', '\"')
                    new_value = json.loads(new_value)
                except:
                    pass
                if key_to_modify in data and new_value not in data[key_to_modify]:
                    data[key_to_modify].append(new_value)
                    print(new_value, data[key_to_modify])
                    # 写回JSON文件
                    with open(json_file_path, 'w', encoding=encodingType) as file:
                        json.dump(data, file, ensure_ascii=False, indent=4)
                    self.messageInformInWin(f'modification is successful!', 2000)
                else:
                    self.messageInformInWin(f'modification exists repeatation!', 2000)
            self.load_llm_records(which_combo,key_to_modify)
        except Exception as e:
            self.messageInformInWin(f'modification is unsuccessful!',2000)
            pass

    def changeKeyValueInJsonFile2(self, json_file_path,key_to_modify, new_value,pattern):
        try:
            # 读取JSON文件
            data = read_json_file(json_file_path)
            if not isinstance(data, dict):
                data = json.loads(data)

            # 修改键值，假设要修改"key_to_modify"的值，注意还没写进去文件里
            if pattern==0:
                if key_to_modify in data and new_value not in data[key_to_modify]:
                    data[key_to_modify].append(new_value)
                    # 写回JSON文件
                    with open(json_file_path, 'w', encoding=encodingType) as file:
                        json.dump(data, file, ensure_ascii=False, indent=4)
                    self.messageInformInWin(f'modification is successful!', 2000)
                else:
                    self.messageInformInWin(f'modification exists repeatation!', 2000)
            elif pattern==1:
                try:
                    new_value = new_value.replace('\'', '\"')
                    new_value = json.loads(new_value)
                except:
                    pass
                if key_to_modify in data and new_value not in data[key_to_modify]:
                    data[key_to_modify].append(new_value)
                    print(new_value, data[key_to_modify])
                    # 写回JSON文件
                    with open(json_file_path, 'w', encoding=encodingType) as file:
                        json.dump(data, file, ensure_ascii=False, indent=4)
                    self.messageInformInWin(f'modification is successful!', 2000)
                else:
                    self.messageInformInWin(f'modification exists repeatation!', 2000)
        except Exception as e:
            self.messageInformInWin(f'modification is unsuccessful!',2000)
            pass

    # 设置线程，避免窗口回弹崩溃
    def thread_it(self, func, *args):
        """将函数打包进线程"""
        self.myThread = threading.Thread(target=func, args=args)
        self.myThread.setDaemon(True)  # 主线程退出就直接让子线程跟随退出 ，不论是否运行完成
        self.myThread.start()

    def send_request_by_userInputData(self):  # 获取用户输入
        url = self.url_entry.get()
        method = self.method_combobox.get()
        headers_input = self.headers_text.get("1.0", tk.END)
        body_input = self.body_text.get("1.0", tk.END)
        if method == 'GET' and '&' in body_input:
            td = {}
            for i in body_input.split('&'):
                j = i.split('=')
                td[j[0]] = j[1]
            body_input = td
            #print(td)

        # 尝试将输入转换为字典
        headers = self.try_parse_json(headers_input)  # <class 'dict'>
        body = self.try_parse_json(body_input)  # <class 'dict!>

        # 前置依赖鉴权接口
        if 'Cookie' in headers:
            loginResponseInfo = self.send_request_by_recordFile(self.before_login_combobox.get())
            # print(type(loginResponseInfo), loginResponseInfo)
            if loginResponseInfo != '':
                newestAuthCookie = find_value_of_key_in_nested_dict(loginResponseInfo, "Cookie")
                # print('newestAuthCookie',newestAuthCookie)
                headers['Cookie'] = newestAuthCookie
                # print(headers)
                # 更新Cookie到headers_text的输入框里
                self.headers_text.delete('1.0', tk.END)
                self.headers_text.insert(tk.END, self.format_json(json.dumps(headers)))
        # 发送请求
        try:
            self.messageInformInWin('Sending...', 700)
            if method == 'GET':
                response = requests.request(method, url=url, params=body, headers=headers)
                # response = requests.get(unl=unl,params=body,headers=headers)
                # 展示响应
                self.response_text.delete("1.0", tk.END)
                self.response_text.insert(tk.END, self.format_json(response.text))
                save_request_data(url, headers, method, body, self.try_parse_json(response.text))
                self.load_records()
                self.record_combobox.set(list_request_records()[1])
            elif method == 'POST':
                response = requests.post(url=url, data=json.dumps(body), headers=headers)
                # print(response.headens.get( set-Cookie!))
                Cookie = response.headers.get('Set-Cookie')
                responseDict = json.loads(response.text)
                if not isinstance(responseDict, dict):
                    responseDict = json.loads(responseDict)
                    print(type(responseDict))
                if 'data' in responseDict:
                    if isinstance(responseDict['data'], list):  # 判断数据类型是否为列表
                        responseDict['data'].append({'Cookie': Cookie})  # <class 'list'>
                    elif isinstance(responseDict['data'], dict):  # 判断数据类型是否为字典
                        responseDict['data']['Cookie'] = Cookie  # <class 'dict'>'
                    elif responseDict['data'] is None:  # 判断数据类型是否为字典
                        responseDict['data'] = dict()
                        responseDict['data']['data'] = None
                        responseDict['data']['Cookie'] = Cookie  # <class 'dict'>
                else:
                    responseDict['Cookie'] = Cookie  # <class 'dict'>pass
                responseJson = json.dumps(responseDict)
                # 展示响应
                self.response_text.delete('1.0', tk.END)
                self.response_text.insert(tk.END, self.format_json(responseJson))
                save_request_data(url, headers, method, body, self.try_parse_json(responseJson))
                self.load_records()
                self.record_combobox.set(list_request_records()[1])
            else:
                response = requests.request(method, url=url, params=json.dumps(body), headers=headers)  # 展示响应
                self.response_text.delete('1.0', tk.END)
                self.response_text.insert(tk.END, self.format_json(response.text))
                save_request_data(url, headers, method, body, self.try_parse_json(response.text))
                self.load_records()
                self.record_combobox.set(list_request_records()[1])
            # messagebox.showinfo(Successful!!)

        except Exception as e:
            eDict = {}
            eDict['e'] = f'Error: {e}'
            eJson = json.dumps(eDict)
            self.response_text.delete('1.0', tk.END)
            self.response_text.insert(tk.END, self.format_json(eJson))
            save_request_data(url, headers, method, body, self.try_parse_json(eJson))  # 异常结果也保存self.load_records()

    def send_request_by_recordFile(self, jsonRecordFilePath):
        if jsonRecordFilePath == '':
            return ''
        jsonRecordDict = read_json_file(jsonRecordFilePath)  # dict
        print('jsonRecordDict', jsonRecordDict)
        # 获取用户输入
        url = jsonRecordDict['ur']
        method = jsonRecordDict['method']
        headers_request = jsonRecordDict['headers']
        body_request = jsonRecordDict['request']
        # 尝试将输入转换为字典
        headers = headers_request  # <class 'str'>
        body = json.dumps(body_request)  # <class 'str'>
        # 发送请求
        try:
            self.messageInformInWin('Sending . .', 400)
            if method == 'GET':
                response = requests.request(method, url=url, params=body,
                                            headers=headers)  # response = reguests.get(unl=unl,params=body, headers=headers)
                return json.loads(response.text)
            elif method == 'POST':
                # response = neguests.request(method, unlzunl , .params=json .dumps(body), headers=headens)
                response = requests.post(url=url, data=body, headers=headers)
                print(response)
                Cookie = response.headers.get('Set-Cookie')
                responseDict = json.loads(response.text)  # <class idict'>
                if isinstance(responseDict['data'], list):  # 判断数据类型是否为列表
                    responseDict['data'].append({'Cookie': Cookie})  # <class"list'>
                if isinstance(responseDict['data'], dict):  # 判断数据类型是否为字典
                    responseDict['data']['Cookie'] = Cookie  # <class 'dict'>
                return responseDict
            else:
                response = requests.request(method, url=url, params=json.dumps(body), headers=headers)
                return json.loads(response.text)
        except Exception as e:
            self.response_text.delete('1.0', tk.END)
            self.response_text.insert(tk.END, f"Error: {e}")
            # messagebox.showinfo( Unsuccessful!

    # 尝试解析JSON字符串
    def try_parse_json(self, json_input):
        try:
            return json.loads(json_input)  #
        except Exception as e:
            return {'error': f'Error: {e}'}

    def combinateCommonOperation(self, which_win, which_text, which_combobox):
        operationName = which_combobox.get()
        print(operationName)
        if operationName == 'Format':
            self.thread_it(self.format_content(which_text))
        elif operationName == 'Copy':
            self.thread_it(self.copy_content(which_text))
        elif operationName == 'Paste':
            self.thread_it(self.paste(which_win, which_text))
        elif operationName == 'Clear':
            self.thread_it(self.clearContent(which_text))
        pass

    def excelCommonOperation(self, which_win, which_operation_combo,which_excelPath_combo,which_sheetName_combo,which_columnName_combo):
        operationName = which_operation_combo.get()
        print(operationName)
        if operationName == 'stripEmptyInColumn':
            self.modify_keywords_in_a_sheet_of_excel(
                which_excelPath_combo.get(),
                which_sheetName_combo.get(),
                which_columnName_combo.get(),
                operationName
            )
        elif operationName == 'clearKeyChInColumn':
            self.modify_keywords_in_a_sheet_of_excel(
                which_excelPath_combo.get(),
                which_sheetName_combo.get(),
                which_columnName_combo.get(),
                operationName,
                '*'
            )
        elif operationName == 'Paste':
            pass
        elif operationName == 'Clear':
            pass
        pass

    def format_content(self, which_text):
        formatted = self.format_json(which_text.get("1.0", tk.END))
        which_text.delete('1.0', tk.END)
        which_text.insert(tk.END, formatted)

    # 格式化JSON字符串
    def format_json(self, json_input) -> str:
        try:
            parsed = json.loads(json_input)
            formatted = json.dumps(parsed, ensure_ascii=False, indent=4)
            return formatted  # jsonStr
        except json.JSONDecodeError:
            return ""  # 返回一个空的JSON对象，如果输入无法解析

    def findLoginRecordFilePath(self):
        filetypes = [("JSON Files", "*,json")]
        path = filedialog.askopenfilename(title="选择登录接口文件", filetypes=filetypes)
        # 打开一个文件选择对话框，用户选择json文件
        if not isfile(path):  # 检查所选文件是否存在
            print(f"文件不存在:{path}")
        else:
            pass
        try:
            self.before_login_combobox.insert(tk.END, path)
        except:
            defaultPath = f'{current_script_path}/requests_log/post数字员工请求登录demo.json'
            self.before_login_combobox.insert(tk.END, defaultPath)

    def find_file_to_fill_record(self,which_combo):
        """通过选择的记录填充表单"""
        #record_file = self.record_combobox.get()
        #record_file = which_combo.get()
        # path = f'{curnent_script-pathl/(logFileNamel/irecord_filel
        filetypes = [("JSON Files", "*.json")]
        path = filedialog.askopenfilename(title="选择json文件", filetypes=filetypes)
        # 打开一个文件选择对话框，用户选择json文件
        if not isfile(path):  # 检查所选文件是否存在
            print(f"调用{self.find_file_to_fill_record.__name__}方法查找的文件不存在！")
            self.messageInformInWin(f"调用{self.find_file_to_fill_record.__name__}方法查找的文件不存在！",4000)
        else:
            pass
        try:
            with open(path, 'r') as file:
                data = json.load(file)
                request_data = data['request']
                response_data = data['response']

                self.url_entry.delete(0, tk.END)
                self.url_entry.insert(0, data['url'])
                self.method_combobox.set(data['method'])
                self.headers_text.delete(1.0, tk.END)
                self.headers_text.insert(
                    tk.END,
                    json.dumps(data["headers"], ensure_ascii=False, indent=4).replace('\'', '\"')
                )
                self.body_text.delete(1.0, tk.END)
                self.body_text.insert(
                    tk.END,
                    json.dumps(request_data, ensure_ascii=False, indent=4).replace('\'', '\"')
                )
                self.response_text.delete(1.0, tk.END)
                self.response_text.insert(
                    tk.END,
                    json.dumps(response_data, ensure_ascii=False, indent=4).replace('\'', '\"')
                )
        except:
            pass

    def find_file_to_fill_QA_record(self,which_combo,qtext,atext):
        """通过选择的记录填充表单"""
        #record_file = which_combo #self.record_combobox.get()
        # path = f'{curnent_script-pathl/(logFileNamel/irecord_filel
        filetypes = [("Some Files", "*.txt")]
        path = filedialog.askopenfilename(title="选择文件", filetypes=filetypes)
        # 打开一个文件选择对话框，用户选择json文件
        if not isfile(path):  # 检查所选文件是否存在
            print(f"文件不存在:{path}")
        else:
            pass
        try:
            res=[]
            with open(path, 'r') as file:
                lines=file.readlines()
                for line in lines:
                    res.append(line)

            qa=''.join(res)
            q = qa.split('******')[0].strip()
            a = qa.split('******')[1].strip()

            qtext.delete(1.0, tk.END)
            qtext.insert(tk.END,q)
            atext.delete(1.0, tk.END)
            atext.insert(tk.END,a)

        except:
            pass

    def copy_image(self, which_text):  # 获取输入框的内容
        input_content = which_text.get("1.0", "end-1c")  # 将内容复制到剪贴板
        self.clipboard_clear()
        self.clipboard_append(input_content)  # 显示复制成功的消息《可选》
        print("内容已复制到剪贴板。")  # messagebox.showinfo(input_content)
        self.messageInformInWin(input_content, 2800)
        time.sleep(2)

    def copy_content(self, which_text):  # 获取输入框的内容
        input_content = which_text.get("1.0", "end-1c")  # 将内容复制到剪贴板
        self.clipboard_clear()
        self.clipboard_append(input_content)  # 显示复制成功的消息《可选》
        print("内容已复制到剪贴板。")  # messagebox.showinfo(input_content)
        self.messageInformInWin(input_content, 2800)
        time.sleep(2)

    def copy_content_no_win(self, text):  # 获取输入框的内容
        input_content = text  # 将内容复制到剪贴板
        self.clipboard_clear()
        self.clipboard_append(input_content)  # 显示复制成功的消息《可选》
        print("内容已复制到剪贴板。")  # messagebox.showinfo(input_content)
        #self.messageInformInWin(input_content, 2800)
        time.sleep(1)

    def cut_content(self, which_text):  # 剪切输入框的内容
        input_content = which_text.get("1.0", "end-1c")  # 将内容复制到剪贴板
        self.clearContent(which_text)
        self.clipboard_clear()
        self.clipboard_append(input_content)  # 显示复制成功的消息《可选》
        print("内容已剪切到剪贴板。")  # messagebox.showinfo(input_content)
        self.messageInformInWin(input_content, 2800)
        time.sleep(2)

    def encryptResult(self, stringValue, secretChar):
        if not isinstance(stringValue, str):
            return "输入类型错误"
        if len(stringValue) < 2:
            return stringValue
        newstringValue = secretChar + stringValue[0: len(stringValue) // 2] + secretChar + stringValue[len(
            stringValue) // 2:] + secretChar
        return newstringValue

    def getTestData(self, which_text, encryptChar_text):
        encryptChar = encryptChar_text.get()  # .get()#测试
        # 生成测试数据
        f = Faker(["zh_CN"])  # 默认en_Us，支持中文本地化zh_Ch
        fake_name = f.name()
        fake_id = self.encryptResult(generate_id_card(), encryptChar)
        fake_phone = self.encryptResult(f.phone_number(), encryptChar)
        fake_email = self.encryptResult(f.email(), encryptChar)
        plate_info = generate_ramdon_plate()
        plate_number = self.encryptResult(plate_info[0], encryptChar)
        plate_type = self.encryptResult(plate_info[1], encryptChar)
        fake_postcode = self.encryptResult(f.postcode(), encryptChar)
        vin = self.encryptResult(generate_vin(), encryptChar)
        currentTimeInfo = getCurrentTimeInfo()
        timestamp = self.encryptResult(str(currentTimeInfo[0]), encryptChar)
        formatted_time = self.encryptResult(currentTimeInfo[1], encryptChar)
        bankUser = '平安测试六零零零三四零一二四二零'
        bankName = '平安银行'
        acountCard = self.encryptResult("11006545830302", encryptChar)
        # 以JSON格式返回结果
        data = {
            "姓名": fake_name,
            "身份证号": fake_id,
            "手机号": fake_phone,
            "邮箱": fake_email,
            "车牌号": plate_number,
            "车牌类型": plate_type,
            "VIN': vin,"
            "邮编": fake_postcode,
            "开户名": bankUser,
            "银行": bankName,
            "银行卡": acountCard,
            "时间": timestamp,
            "格式化时间": formatted_time
        }
        # 打印JSON格式数据
        print(json.dumps(data, ensure_ascii=False, indent=4))
        which_text.delete(1.0, tk.END)
        which_text.insert(tk.END, json.dumps(data, ensure_ascii=False, indent=4).replace("\'", '\"'))

    def getHtmlData(self, which_text, html_content):
        tag = False
        output = ""
        res = []
        for char in html_content:
            if char == '<':
                if len(output) != 0:
                    res.append(output)
                    output = ""
                tag = True
            elif char == '>':
                tag = False
            elif not tag:
                output = output + char
        res = [r for r in res if r != '' or r != ' ' or r != '\n' or r != '\t']
        self.clearContent(which_text)
        which_text.insert('insert', '\n'.join(res))

    def clearContent(self, which_text):
        # 清空输入框的内容
        which_text.delete("1.0", "end-1c")

    def clearContent1(self, which_entry):
        # 清空输入框的内容
        which_entry.delete(0, tk.END)

    def paste(self, which_win, which_text):
        # 清空输入框的内容
        which_text.delete("1.0", "end-1c")
        # 获取煎贴板的内容
        clipboard_content = which_win.clipboard_get()
        # 将内容插入到输入框
        which_text.insert('insert', clipboard_content)

    def paste1(self, which_win, which_entry):
        # 清空输入框的内容
        which_entry.delete(0, tk.END)
        # 获取煎贴板的内容
        clipboard_content = which_win.clipboard_get()
        # 将内容插入到输入框
        which_entry.insert('insert', clipboard_content)

    def copy_value_of_key(self, which_win, which_text, targetkey):  # 获取输入框的内容
        input_content = which_text.get("1.0", "end-1c")
        # 将内容复制到剪贴板
        which_win.clipboard_clear()
        try:
            res = self.try_parse_json(input_content)  # <class  dict'>print(type(res))
            value = find_value_of_key_in_nested_dict(res, targetkey)
            which_win.clipboard_append(value)
            # 显示复制成功的消息《可选)
            print("内容已复制到剪贴板。")
            # messagebox.showinfo(value)messageInformInWin(value,lastingTime: 500)
            self.messageInformInWin(value, 2800)
        except Exception as e:
            #which_win.clipboard_append('0000')  # 显示复制成功的消息《可选》
            print(f"内容复制到剪贴板失败，具体原因为:{e}")
            which_win.clipboard_append(f"内容复制到剪贴板失败，具体原因为:{e}")  # 显示复制成功的消息《可选》
            # messagebox.showinfo(f"内容复制到剪贴板失败，具体原因为:{e}")messaneinformTnWincintorncontimplePostmanApp, find file_to_fill record()>try>with open(path, 'r") as file中"肉客更制到前贴板牛哈，且休厦因头
            self.messageInformInWin(f"内容复制到剪贴板失败，具体原因为:{e}", 2800)

            # messagebox.showinfo(f"内容复制到剪贴板失败，具体原因为:{e}")
            # messageInformInWin(infomContent,f"内容复制到剪贴板失败，具体原因为: {e}",500)

    def load_records(self):
        records = list_request_records()
        self.record_combobox['values'] = records

    def load_qa_records(self,which_combo):
        records = list_qa_records()
        which_combo['values'] = records

    def load_llm_records(self,which_combo,which_para):
        records = list_llm_records(which_para)
        which_combo['values'] = records

    ## 发送请求的函数
    def send_request(self):
        # 获取用户输入
        url = self.url_entry.get()
        method = self.method_combobox.get()
        headers_input = self.headers_text.get("1.0", tk.END)
        body_input = self.body_text.get("1.0", tk.END)

        # 尝试将输入转换为字典
        headers = self.try_parse_json(headers_input)
        body = self.try_parse_json(body_input)

        # 发送请求
        try:
            if method == 'GET':
                response = requests.request(method, url=url, params=body, headers=headers)
                # 展示响应
                self.response_text.delete('1.0', tk.END)
                self.response_text.insert(tk.END, self.format_json(response.text))
                save_request_data(url, headers, method, body, self.try_parse_json(response.text))
                self.load_records()
            elif method == 'POST':
                response = requests.post(url=url, data=json.dumps(body), headers=headers)
                # 展示响应
                self.response_text.delete('1.0', tk.END)
                self.response_text.insert(tk.END, self.format_json(response.text))
                save_request_data(url, headers, method, body, self.try_parse_json(response.text))
                self.load_records()
            else:
                response = requests.request(method, url=url, params=body, headers=headers)
                # 展示响应
                self.response_text.delete('1.0', tk.END)
                self.response_text.insert(tk.END, self.format_json(response.text))
                save_request_data(url, headers, method, body, self.try_parse_json(response.text))
                self.load_records()

        except Exception as e:
            self.response_text.delete('1.0', tk.END)
            self.response_text.insert(tk.END, f"Error: {e}")

    # 尝试解析JSON字符串
    def try_parse_json(self, json_input):
        try:
            return json.loads(json_input)
        except json.JSONDecodeError:
            return {}

    # 格式化Headers输入
    def format_headers(self):
        formatted = self.format_json(self.headers_text.get("1.0", tk.END))
        self.headers_text.delete('1.0', tk.END)
        self.headers_text.insert(tk.END, formatted)

    # 格式化Body输入
    def format_body(self):
        formatted = self.format_json(self.body_text.get("1.0", tk.END))
        self.body_text.delete('1.0', tk.END)
        self.body_text.insert(tk.END, formatted)

    # 格式化Response输入
    def format_response(self):
        formatted = self.format_json(self.response_text.get("1.0", tk.END))
        self.response_text.delete('1.0', tk.END)
        self.response_text.insert(tk.END, formatted)

    # 格式化JSON字符串
    def format_json(self, json_input):
        try:
            parsed = json.loads(json_input)
            formatted = json.dumps(parsed, ensure_ascii=False, indent=4)
            return formatted
        except json.JSONDecodeError:
            return "{}"  # 返回一个空的JSON对象，如果输入无法解析

    def fill_record(self, event):
        """通过选择的记录填充表单"""
        record_file = self.record_combobox.get()
        path = f'{current_script_path}/{logFileName}/{record_file}'
        with open(path, 'r',encoding=encodingType) as file:
            data = json.load(file)
            request_data = data['request']
            response_data = data['response']

            self.url_entry.delete(0, tk.END)
            self.url_entry.insert(0, data['url'])
            self.method_combobox.set(data['method'])
            self.headers_text.delete(1.0, tk.END)
            self.headers_text.insert(tk.END,
                                     json.dumps(data['headers'], ensure_ascii=False, indent=4).replace('\'', '\"'))
            self.body_text.delete(1.0, tk.END)
            self.body_text.insert(tk.END, json.dumps(request_data, ensure_ascii=False, indent=4).replace('\'', '\"'))
            self.response_text.delete(1.0, tk.END)
            self.response_text.insert(tk.END,
                                      json.dumps(response_data, ensure_ascii=False, indent=4).replace('\'', '\"'))

    def fill_QA_record(self, event):
        """通过选择的记录填充表单"""
        record_file = self.qa_record_combo.get() #self.record_combobox.get()
        qtext = self.left_output_text
        atext = self.right_output_text
        path = f'{current_script_path}/{QALogFileName}/{record_file}'
        try:
            res = []
            with open(path, 'r') as file:
                lines = file.readlines()
                for line in lines:
                    res.append(line)

            qa = ''.join(res)
            q = qa.split('******')[0]
            a = qa.split('******')[1]

            qtext.delete(1.0, tk.END)
            qtext.insert(tk.END, q)
            atext.delete(1.0, tk.END)
            atext.insert(tk.END, a)

        except:
            pass

    def copy_headers_text(self):
        # 获取输入框内容
        input_content = self.headers_text.get('1.0', 'end-1c')
        self.clipboard_clear()
        self.clipboard_append(input_content)
        # 显示复制成功消息（可选）
        print('内容已复制到剪贴板。')

    def copy_body_text(self):
        # 获取输入框内容
        input_content = self.body_text.get('1.0', 'end-1c')
        self.clipboard_clear()
        self.clipboard_append(input_content)
        # 显示复制成功消息（可选）
        print('内容已复制到剪贴板。')

    def copy_response_text(self):
        # 获取输入框内容
        input_content = self.response_text.get('1.0', 'end-1c')
        self.clipboard_clear()
        self.clipboard_append(input_content)
        # 显示复制成功消息（可选）
        print('内容已复制到剪贴板。')

    def copy_response_key(self):
        key = 'data'
        # 获取输入框内容
        input_content = self.response_text.get('1.0', 'end-1c')
        # 将内容复制到剪贴板
        self.clipboard_clear()
        try:
            res = self.try_parse_json(input_content)  # <class 'dict'>
            print(type(res))
            self.clipboard_append(find_value_in_nested_dict(res, key))
            # 显示复制成功的消息（可选）
            print('内容已复制到剪贴板。')
        except Exception as e:
            self.clipboard_append('0000')
            # 显示复制成功的消息（可选）
            print(f'内容复制到薄贴板失败，具体原因为：{e}')

    def search_text(self, search_term, which_text):
        # 清除之前高亮
        which_text.tag_remove('match', '1.0', 'end')
        # 如录有裂索内容，则进行搜索
        if search_term:
            start_pos = "1.0"
            while True:
                # 搜索第一个匹配项
                pos = which_text.search(search_term, start_pos, stopindex='end')
                # 如果没有找到匹配项，退出循环
                if not pos:
                    break
                # 标记匹配项
                end_pos = f"{pos}+{len(search_term)}c"
                which_text.tag_add("match", pos, end_pos)
                # 设置新的搜索起始位置
                start_pos = end_pos


if __name__ == '__main__':
    app = SimplePostmanApp()
    app.mainloop()

"""

export PATH="/usr/local/bin/python3.7:$PATH"

Windows打包应用
#pyinstaller --onefile --windowed D:\TestDev\PostmanByTk1.py

Mac打包应用
一、pyinstaller 主要用于 Python3
which -a python3.7--看python3.7安装在哪里
pip show pyinstaller--看pyinstaller安装在哪里
/usr/local/bin/python3.7 -m pip install pyinstaller--指定版本安装
/opt/homebrew/bin/python3.12 -m pip install pyinstaller--指定版本安装
pip3.7 show pyinstaller--检查是否安装成功 

# PyInstaller --onefile --windowed /Users/ketangchen/Documents/UiAutoOfApp20240703/UiAutoOfApp/utils/PostmanByTk1.py
python3.7 -m PyInstaller --onefile --windowed /Users/ketangchen/Documents/UiAutoOfApp20240703/UiAutoOfApp/utils/PostmanByTk1.py ##重点看python3.7这里，python3.7下安装了PyInstaller

#dist路径
/Users/ketangchen/Documents/UiAutoOfApp/utils/dist

brew install create-dmg # arch -arm64 brew install create-dmg
create-dmg \
  --volname "AboutPostman" \
  --window-size 800 600 \
  --app-drop-link 400 200 \
  /Users/ketangchen/Desktop/AboutPostman.dmg  \
  /Users/ketangchen/Documents/UiAutoOfApp/utils/dist/aboutPostman.app


未能打开磁盘映像
pip3 install scholarly  -i https://pypi.tuna.tsinghua.edu.cn/simple
/usr/local/bin/python3.7 -m pip install pyinstaller -i https://pypi.tuna.tsinghua.edu.cn/simple
/usr/local/bin/python3.7 -m pip install --upgrade pyinstaller -i https://pypi.tuna.tsinghua.edu.cn/simple
=======
python3.7 -m PyInstaller --onefile --windowed /Users/ketangchen/Documents/UiAutoOfApp20240703/UiAutoOfApp/utils/PostmanByTk1.py ##重点看python3.7这里，python3.7下安装了PyInstaller
/usr/local/bin/python3.7 pip -m install Faker  -i https://pypi.tuna.tsinghua.edu.cn/simple

二、py2app
python3 -m pip uninstall setuptools==69.5.1

pip install --upgrade setuptools
/usr/local/bin/python3.7 -m pip install setuptools
pip install --upgrade setuptools
/usr/local/bin/python3.7 -m pip install setuptools
创建你的setup.py脚本。若无，则可以使用py2applet创建一个。在终端中运行
py2applet --make-setup aboutPostman.py
运行setup.py脚本来创建应用
python3.7 setup.py py2app

python3.7 setup.py py2app --debug

chmod +x /path/to/your/app/YourApp.appYourApp
chmod x /Users/ketangchen/Documents/UiAutoOfApp/utils/dist/aboutPostman.app/Contents/MacOS/aboutPostman

dist路径: /Users/ketangchen/Documents/UiAutoOfApp20240703/UiAutoOfApp/utils



打包依赖库
pip freeze > requirements.txt
安装依赖库
pip install -r /Users/ketangchen/Documents/UiAutoOfApp20240703/UiAutoOfApp/utils/requirements.txt



xattr -r -d com.apple.quarantine  /Users/ketangchen/Desktop/AboutPostman.dmg
xattr -r -d com.apple.quarantine  /Users/ketangchen/Documents/UiAutoOfApp/utils/dist/aboutPostman.app

arch -x86_64 python3 PostmanByTk1.py
arch -x86_64 python3 /Users/ketangchen/Documents/UiAutoOfApp20240703/UiAutoOfApp/utils/aboutOcr_PaddleOcr.py
什么原因造成的？如何解决？



操作系统版本（uname -m）：x86_64
Python版本（python --version）:3.9
PaddlePaddle和PaddleOCR的版本（pip list | grep paddle）:

pip怎么查看一个三方库有哪些版本？
pip index versions <package>
pip index versions paddlepaddle
pip index versions paddleocr

conda创建虚拟环境
conda create -n my_env1 python=3.9
conda activate my_env1
conda deactivate
conda env list
file $(which python3)


import paddle
print(paddle.__file__) # /Users/ketangchen/miniconda3/envs/my_env1/lib/python3.9/site-packages/paddle/__init__.py


cp -r /Users/ketangchen/miniconda3/envs/my_env1/lib/python3.9/site-packages/paddle/libs ./dist/PostmanByTk1.app/Contents/MacOS/ 
"""

